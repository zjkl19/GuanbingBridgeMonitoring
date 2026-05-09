#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Generate a bridge point configuration and CSV acquisition workbook.

The script compares three sources:
  1. design/device list xlsx,
  2. JSON configuration points/groups,
  3. actual daily-export CSV files under data_<bridge>_YYYY-MM-DD.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


MODULE_ALIASES: dict[str, list[str]] = {
    "temperature": ["temperature"],
    "humidity": ["humidity"],
    "wind": ["wind", "wind_speed", "wind_direction"],
    "earthquake": ["earthquake", "eq"],
    "deflection": ["deflection"],
    "bearing_displacement": ["bearing_displacement"],
    "tilt": ["tilt"],
    "acceleration": ["acceleration"],
    "accel_spectrum": ["accel_spectrum", "acceleration"],
    "cable_accel": ["cable_accel"],
    "cable_accel_spectrum": ["cable_accel_spectrum", "cable_accel"],
    "strain": ["strain", "strain_timeseries"],
    "dynamic_strain_highpass": ["dynamic_strain", "strain"],
    "dynamic_strain_lowpass": ["dynamic_strain_lowpass", "dynamic_strain", "strain"],
    "gnss": ["gnss"],
    "rainfall": ["rainfall"],
}


MODULE_LABELS: dict[str, str] = {
    "temperature": "温度",
    "humidity": "湿度",
    "wind": "风速风向",
    "earthquake": "地震动",
    "deflection": "挠度",
    "bearing_displacement": "支座/梁端位移",
    "tilt": "倾角",
    "acceleration": "振动加速度",
    "accel_spectrum": "振动频谱",
    "cable_accel": "索力加速度",
    "cable_accel_spectrum": "索力加速度频谱",
    "strain": "应变",
    "dynamic_strain_highpass": "动应变（高通）",
    "dynamic_strain_lowpass": "动应变（低通）",
    "gnss": "GNSS",
    "rainfall": "雨量",
}


def flatten_points(value: Any) -> list[str]:
    out: list[str] = []
    if value is None:
        return out
    if isinstance(value, str):
        return [value] if value else []
    if isinstance(value, (list, tuple, set)):
        for item in value:
            out.extend(flatten_points(item))
        return out
    if isinstance(value, dict):
        for item in value.values():
            out.extend(flatten_points(item))
        return out
    return out


def unique_keep_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for v in values:
        s = str(v).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def point_candidates(point_id: str) -> list[str]:
    candidates = [point_id]
    candidates.append(re.sub(r"[-_][XYZ]$", "", point_id))
    candidates.append(re.sub(r"[-_][XYZ][-_]?", "-", point_id))
    candidates.append(re.sub(r"[-_][XYZ][-_]([^\\/]+)$", r"-\1", point_id))
    return unique_keep_order(candidates)


def config_points_by_module(cfg: dict[str, Any]) -> dict[str, list[str]]:
    points = cfg.get("points") or {}
    groups = cfg.get("groups") or {}
    result: dict[str, list[str]] = {}
    for module, aliases in MODULE_ALIASES.items():
        vals: list[str] = []
        for alias in aliases:
            if alias in points:
                vals.extend(flatten_points(points[alias]))
            if alias in groups:
                vals.extend(flatten_points(groups[alias]))
        result[module] = unique_keep_order(vals)
    return result


def collect_csv(root: Path, start: str, end: str, prefixes: list[str]) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    for folder in sorted(root.glob("data_*_????-??-??")):
        if not folder.is_dir():
            continue
        m = re.match(r"data_([A-Za-z0-9]+)_(\d{4}-\d{2}-\d{2})$", folder.name)
        if not m:
            continue
        prefix, day = m.groups()
        if prefix not in prefixes or day < start or day > end:
            continue
        candidates = [folder / "data" / prefix / "csv", folder / "data" / "jlj" / "csv", folder / "data" / "sxh" / "csv", folder / "csv"]
        csv_dir = next((p for p in candidates if p.is_dir()), None)
        if not csv_dir:
            continue
        for file in csv_dir.glob("*.csv"):
            rec = records.setdefault(file.stem, {"point_id": file.stem, "days": set(), "files": []})
            rec["days"].add(day)
            rec["files"].append(str(file))
    for rec in records.values():
        rec["days"] = sorted(rec["days"])
    return records


def read_design_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = None
    headers: list[str] = []
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if any(str(v).strip() == "设备编号" for v in values if v is not None):
            header_row = row_idx
            headers = [str(v).strip() if v is not None else f"列{col}" for col, v in enumerate(values, 1)]
            break
    if header_row is None:
        raise ValueError(f"未找到设计表表头: {path}")

    rows: list[dict[str, Any]] = []
    carry_fields = ["设备名称", "采集频率"]
    carry_values = {name: None for name in carry_fields}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        row = {headers[col - 1]: values[col - 1] for col in range(1, len(headers) + 1)}
        for name in carry_fields:
            if name in row:
                if row[name] is None or str(row[name]).strip() == "":
                    row[name] = carry_values[name]
                else:
                    carry_values[name] = row[name]
        rows.append(row)
    return headers, rows


def infer_module(device_name: str, device_id: str) -> str:
    text = f"{device_name or ''} {device_id or ''}"
    rules = [
        ("温湿", "temperature"),
        ("温度", "temperature"),
        ("风速", "wind"),
        ("风向", "wind"),
        ("地震", "earthquake"),
        ("加速度", "acceleration"),
        ("振动", "acceleration"),
        ("索力", "cable_accel"),
        ("吊杆", "cable_accel"),
        ("应变", "strain"),
        ("挠度", "deflection"),
        ("位移", "bearing_displacement"),
        ("伸缩", "bearing_displacement"),
        ("倾角", "tilt"),
        ("GNSS", "gnss"),
        ("雨", "rainfall"),
    ]
    for key, module in rules:
        if key in text:
            return module
    return "other"


def find_match(point_id: str, actual_ids: set[str]) -> str:
    for candidate in point_candidates(point_id):
        if candidate in actual_ids:
            return candidate
    return ""


def apply_table_style(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border


def autosize(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row, col).value
            if value is not None:
                width = max(width, min(max_width, len(str(value)) * 1.8))
        ws.column_dimensions[letter].width = width


def build_workbook(
    cfg_path: Path,
    data_root: Path,
    design_path: Path,
    output_path: Path,
    start: str,
    end: str,
) -> None:
    cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    vendor = str(cfg.get("vendor") or "")
    prefixes = ["sxh"] if vendor in {"shuixianhua", "sxh"} else ["jlj", "sxh"]
    configured = config_points_by_module(cfg)
    actual = collect_csv(data_root, start, end, prefixes)
    actual_ids = set(actual)
    design_headers, design_rows = read_design_rows(design_path)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "汇总"
    ws_cfg = wb.create_sheet("配置测点获取情况")
    ws_design = wb.create_sheet("设计表核对")
    ws_csv = wb.create_sheet("实际CSV清单")

    summary_rows = []
    for module, pts in configured.items():
        if not pts:
            continue
        found = sum(1 for p in pts if find_match(p, actual_ids))
        missing = len(pts) - found
        summary_rows.append([
            module,
            MODULE_LABELS.get(module, module),
            len(pts),
            found,
            missing,
            found / len(pts) if pts else 0,
        ])

    ws_sum.append(["模块代码", "模块", "配置测点数", "实际获取测点数", "缺失测点数", "获取率"])
    for row in summary_rows:
        ws_sum.append(row)
    for cell in ws_sum["F"][1:]:
        cell.number_format = "0.0%"

    ws_cfg.append(["模块代码", "模块", "配置测点", "是否获取CSV", "匹配CSV测点", "获取天数", "备注"])
    for module, pts in configured.items():
        for point in pts:
            matched = find_match(point, actual_ids)
            days = actual.get(matched, {}).get("days", []) if matched else []
            ws_cfg.append([
                module,
                MODULE_LABELS.get(module, module),
                point,
                "是" if matched else "否",
                matched,
                len(days),
                "" if matched else "配置测点未找到对应CSV",
            ])

    design_out_headers = design_headers + ["推断模块", "是否在配置中", "是否获取CSV", "匹配CSV测点", "备注"]
    ws_design.append(design_out_headers)
    all_config_points = {p for pts in configured.values() for p in pts}
    all_config_candidates = {c for p in all_config_points for c in point_candidates(p)}
    for row in design_rows:
        point_id = str(row.get("设备编号") or "").strip()
        module = infer_module(str(row.get("设备名称") or ""), point_id)
        matched = find_match(point_id, actual_ids)
        in_config = point_id in all_config_points or any(c in all_config_candidates for c in point_candidates(point_id))
        values = [row.get(h) for h in design_headers]
        values += [MODULE_LABELS.get(module, module), "是" if in_config else "否", "是" if matched else "否", matched, "" if matched else "设计测点未找到对应CSV"]
        ws_design.append(values)

    ws_csv.append(["CSV测点", "获取天数", "首日", "末日", "样例文件"])
    for point_id in sorted(actual):
        days = actual[point_id]["days"]
        files = actual[point_id]["files"]
        ws_csv.append([point_id, len(days), days[0] if days else "", days[-1] if days else "", files[0] if files else ""])

    for ws in [ws_sum, ws_cfg, ws_design, ws_csv]:
        apply_table_style(ws)
        autosize(ws)
    for row in ws_cfg.iter_rows(min_row=2, min_col=4, max_col=4):
        if row[0].value == "否":
            row[0].fill = PatternFill("solid", fgColor="F4CCCC")
    for row in ws_design.iter_rows(min_row=2, min_col=len(design_out_headers) - 2, max_col=len(design_out_headers) - 2):
        if row[0].value == "否":
            row[0].fill = PatternFill("solid", fgColor="F4CCCC")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    parser.add_argument("--data-root", required=True)
    parser.add_argument("--design", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--start", required=True)
    parser.add_argument("--end", required=True)
    args = parser.parse_args()
    build_workbook(
        Path(args.config),
        Path(args.data_root),
        Path(args.design),
        Path(args.output),
        args.start,
        args.end,
    )
    print(args.output)


if __name__ == "__main__":
    main()
