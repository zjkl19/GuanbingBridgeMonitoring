"""Prepare and compare read-only real-bridge high-frequency regressions.

This helper deliberately does not launch MATLAB.  ``prepare`` creates an
isolated data root whose dated folders are NTFS junctions to an existing MAT
cache tree, freezes the effective bridge configuration, pins MAT-only input,
and writes a four-module run request.  ``compare`` verifies that the source
cache did not change and compares the completed candidate with an accepted
baseline.

The source cache is treated as immutable.  The generated run request never
enables archive extraction, CSV preprocessing, resampling, cache generation,
WIM, or any analysis module other than acceleration/cable-acceleration time
histories and their spectra.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import os
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

try:
    from workbench.config_layers import config_dependency_sha256, load_layered_config
    from workbench.provenance import inspect_plot_provenance
except ModuleNotFoundError:  # direct ``python scripts/highfreq_regression.py``
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from workbench.config_layers import config_dependency_sha256, load_layered_config
    from workbench.provenance import inspect_plot_provenance


ROOT = Path(__file__).resolve().parents[1]
DATE_DIR_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
MODULES = ("acceleration", "cable_accel", "accel_spectrum", "cable_accel_spectrum")
MODULE_OPTIONS = {
    "acceleration": "doAccel",
    "cable_accel": "doCableAccel",
    "accel_spectrum": "doAccelSpectrum",
    "cable_accel_spectrum": "doCableAccelSpectrum",
}
ALL_OPTION_FIELDS = (
    "precheck_zip_count",
    "doUnzip",
    "doRenameCsv",
    "doRemoveHeader",
    "doResample",
    "doCachePrebuild",
    "doLowfreqSync",
    "doTemp",
    "doHumidity",
    "doRainfall",
    "doGNSS",
    "doWind",
    "doEq",
    "doWIM",
    "doDeflect",
    "doBearingDisplacement",
    "doTilt",
    "doAccel",
    "doCableAccel",
    "doAccelSpectrum",
    "doCableAccelSpectrum",
    "doRenameCrk",
    "doCrack",
    "doStrain",
    "doDynStrainBoxplot",
    "doDynStrainLowpassBoxplot",
)
STATS_FILES = {
    "acceleration": "accel_stats.xlsx",
    "cable_accel": "cable_accel_stats.xlsx",
    "accel_spectrum": "accel_spec_stats.xlsx",
    "cable_accel_spectrum": "cable_accel_spec_stats.xlsx",
}


@dataclass(frozen=True)
class BridgeProfile:
    key: str
    config_path: Path
    source_root: Path
    baseline_root: Path
    start_date: str
    end_date: str
    representative_acceleration: tuple[str, ...]
    representative_cable_accel: tuple[str, ...]


PROFILES = {
    "hongtang": BridgeProfile(
        key="hongtang",
        config_path=ROOT / "config" / "hongtang_config.json",
        source_root=Path(r"E:\洪塘大桥数据\2026年4-6月"),
        baseline_root=Path(
            r"E:\GuanbingLocalValidation\v1.8.0-rc2_20260713_2118"
            r"\hongtang_q2_complete_0628_0630_20260713_2225"
        ),
        start_date="2026-04-01",
        end_date="2026-06-30",
        representative_acceleration=("A1", "A10-X"),
        representative_cable_accel=("CS1", "CS6"),
    ),
    "zhishan": BridgeProfile(
        key="zhishan",
        config_path=ROOT / "config" / "zhishan_config.json",
        source_root=Path(r"E:\芝山大桥\2026年4月"),
        baseline_root=Path(
            r"E:\GuanbingLocalValidation\v1.8.0-rc3_20260714_045654\zhishan"
        ),
        start_date="2026-04-01",
        end_date="2026-04-30",
        representative_acceleration=("AZ-1",),
        representative_cable_accel=("CF-1", "CF-5", "CF-7"),
    ),
}


ARTIFACT_CATEGORIES = (
    ("acceleration_raw", "时程曲线_加速度", "acceleration", ("image", "fig", "plot_json")),
    ("acceleration_rms", "时程曲线_加速度_RMS10min", "acceleration", ("image", "fig")),
    ("cable_accel_raw", "时程曲线_索力加速度", "cable_accel", ("image", "fig", "plot_json")),
    ("cable_accel_rms", "时程曲线_索力加速度_RMS10min", "cable_accel", ("image", "fig")),
    ("cable_accel_envelope", "时程曲线_索力加速度_包络30min", "cable_accel", ("image", "fig")),
    ("acceleration_spectrum_trend", "频谱峰值曲线_加速度", "acceleration", ("image", "fig")),
    ("cable_accel_spectrum_trend", "频谱峰值曲线_索力加速度", "cable_accel", ("image", "fig")),
    ("acceleration_psd", "PSD_备查", "acceleration", ("image", "fig")),
    ("cable_accel_psd", "PSD_备查_索力加速度", "cable_accel", ("image", "fig")),
)

# Complete set of high-frequency formal-output roots used by the MATLAB
# artifact collector.  ``ARTIFACT_CATEGORIES`` is point-oriented for baseline
# comparison, while this list also covers group figures and derived cable-force
# figures so the final manifest gate can prove every formal file is enumerated.
MANIFEST_FORMAL_DIRS = (
    "时程曲线_加速度",
    "时程曲线_加速度_组图",
    "时程曲线_加速度_RMS10min",
    "时程曲线_加速度_RMS10min_组图",
    "时程曲线_索力加速度",
    "时程曲线_索力加速度_组图",
    "时程曲线_索力加速度_RMS10min",
    "时程曲线_索力加速度_RMS10min_组图",
    "时程曲线_索力加速度_包络30min",
    "频谱峰值曲线_加速度",
    "频谱峰值曲线_结构加速度_组图",
    "PSD_备查",
    "频谱峰值曲线_索力加速度",
    "索力时程图",
    "索力时程图_组图",
    "PSD_备查_索力加速度",
)


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )


def _read_json(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, dict):
        raise ValueError(f"JSON root must be an object: {path}")
    return payload


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _canonical_sha256(payload: Any) -> str:
    raw = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(raw).hexdigest().upper()


def _date_in_range(name: str, start: str, end: str) -> bool:
    if not DATE_DIR_RE.fullmatch(name):
        return False
    try:
        value = date.fromisoformat(name)
    except ValueError:
        return False
    return date.fromisoformat(start) <= value <= date.fromisoformat(end)


def discover_date_directories(source_root: Path, start: str, end: str) -> list[Path]:
    if not source_root.is_dir():
        raise FileNotFoundError(f"Source data root does not exist: {source_root}")
    result = sorted(
        (
            child
            for child in source_root.iterdir()
            if child.is_dir() and _date_in_range(child.name, start, end)
        ),
        key=lambda item: item.name,
    )
    if not result:
        raise ValueError(
            f"No dated folders in {source_root} for {start} through {end}"
        )
    return result


def _inventory_file_record(path: Path, root: Path, hash_mat: bool) -> dict[str, Any]:
    stat = path.stat()
    relative = path.relative_to(root).as_posix()
    is_meta = path.name.casefold().endswith(".mat.meta.json")
    record: dict[str, Any] = {
        "path": relative,
        "kind": "meta" if is_meta else "mat",
        "bytes": int(stat.st_size),
        "modified_ns": int(stat.st_mtime_ns),
    }
    if is_meta or hash_mat:
        record["sha256"] = _sha256_file(path)
    return record


def build_source_inventory(
    source_root: Path,
    date_directories: Sequence[Path],
    *,
    hash_mode: str = "metadata",
) -> dict[str, Any]:
    if hash_mode not in {"metadata", "sha256"}:
        raise ValueError("hash_mode must be metadata or sha256")
    records: list[dict[str, Any]] = []
    for day in date_directories:
        if day.parent.resolve() != source_root.resolve():
            raise ValueError(f"Date directory is outside source root: {day}")
        for path in sorted(
            (item for item in day.rglob("*") if item.is_file()),
            key=lambda item: item.as_posix().casefold(),
        ):
            lower = path.name.casefold()
            if path.suffix.casefold() == ".mat" or lower.endswith(".mat.meta.json"):
                records.append(
                    _inventory_file_record(path, source_root, hash_mode == "sha256")
                )
    mats = {item["path"] for item in records if item["kind"] == "mat"}
    metas = {item["path"] for item in records if item["kind"] == "meta"}
    missing_meta = sorted(f"{path}.meta.json" for path in mats if f"{path}.meta.json" not in metas)
    orphan_meta = sorted(
        path for path in metas if path.removesuffix(".meta.json") not in mats
    )
    if missing_meta or orphan_meta:
        raise ValueError(
            "MAT/meta cache pairing is not closed: "
            f"missing_meta={len(missing_meta)}, orphan_meta={len(orphan_meta)}"
        )
    identity = {
        "root": str(source_root.resolve()),
        "hash_mode": hash_mode,
        "date_directories": [item.name for item in date_directories],
        "files": records,
    }
    return {
        "schema_version": 1,
        "inventory_type": "read_only_mat_cache",
        "created_at": _now_iso(),
        **identity,
        "summary": {
            "date_count": len(date_directories),
            "mat_count": len(mats),
            "meta_count": len(metas),
            "mat_bytes": sum(item["bytes"] for item in records if item["kind"] == "mat"),
            "meta_bytes": sum(item["bytes"] for item in records if item["kind"] == "meta"),
            "missing_meta_count": 0,
            "orphan_meta_count": 0,
        },
        "inventory_sha256": _canonical_sha256(identity),
    }


def compare_inventories(before: dict[str, Any], after: dict[str, Any]) -> dict[str, Any]:
    before_records = {item["path"]: item for item in before.get("files", [])}
    after_records = {item["path"]: item for item in after.get("files", [])}
    added = sorted(set(after_records) - set(before_records))
    removed = sorted(set(before_records) - set(after_records))
    changed = sorted(
        path
        for path in set(before_records) & set(after_records)
        if before_records[path] != after_records[path]
    )
    passed = not (added or removed or changed)
    return {
        "status": "ok" if passed else "failed",
        "before_sha256": before.get("inventory_sha256", ""),
        "after_sha256": after.get("inventory_sha256", ""),
        "added": added,
        "removed": removed,
        "changed": changed,
        "passed": passed,
    }


def _create_ntfs_junction(link: Path, target: Path) -> None:
    if os.name != "nt":
        raise OSError("NTFS junction preparation is only supported on Windows")
    if link.exists() or link.is_symlink():
        raise FileExistsError(f"Junction destination already exists: {link}")
    completed = subprocess.run(
        ["cmd.exe", "/d", "/c", "mklink", "/J", str(link), str(target)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=False,
    )
    if completed.returncode != 0 or not link.is_dir():
        message = (completed.stderr or completed.stdout or "mklink failed").strip()
        raise RuntimeError(f"Unable to create junction {link} -> {target}: {message}")


def _positive_number(value: Any, default: float) -> float:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
        if math.isfinite(number) and number > 0:
            return number
    return float(default)


def _effective_dynamic_contract(config: dict[str, Any]) -> dict[str, dict[str, Any]]:
    common = config.get("plot_common") if isinstance(config.get("plot_common"), dict) else {}
    overrides = (
        common.get("dynamic_raw_modules")
        if isinstance(common.get("dynamic_raw_modules"), dict)
        else {}
    )
    result: dict[str, dict[str, Any]] = {}
    for module in ("acceleration", "cable_accel"):
        override = overrides.get(module) if isinstance(overrides.get(module), dict) else {}
        sampling_mode = str(
            override.get("sampling_mode", common.get("dynamic_raw_sampling_mode", ""))
        ).strip().casefold()
        policy_default = "all_vertices" if sampling_mode == "full" else "peak_preserving"
        full_render_policy = str(
            override.get(
                "full_render_policy",
                common.get("dynamic_raw_full_render_policy", policy_default),
            )
        ).strip().casefold()
        base_max = _positive_number(common.get("fig_max_points"), 50000)
        render_default = max(1200000.0, base_max)
        render_max_points = _positive_number(
            override.get(
                "render_max_points",
                common.get("dynamic_raw_fig_max_points"),
            ),
            render_default,
        )
        min_points_per_day = _positive_number(
            override.get(
                "min_points_per_day",
                common.get("dynamic_raw_min_points_per_day"),
            ),
            12000,
        )
        result[module] = {
            "sampling_mode": sampling_mode,
            "render_mode": str(
                override.get("render_mode", common.get("dynamic_raw_render_mode", ""))
            ).strip().casefold(),
            "gap_mode": str(
                override.get("gap_mode", common.get("gap_mode", ""))
            ).strip().casefold(),
            "full_render_policy": full_render_policy,
            "render_max_points": int(round(render_max_points)),
            "min_points_per_day": int(round(min_points_per_day)),
        }
    return result


def _require_full_connect(config: dict[str, Any], label: str) -> dict[str, dict[str, Any]]:
    contract = _effective_dynamic_contract(config)
    expected = {
        "sampling_mode": "full",
        "render_mode": "line",
        "gap_mode": "connect",
        "full_render_policy": "peak_preserving",
        "render_max_points": 1200000,
        "min_points_per_day": 12000,
    }
    for module, values in contract.items():
        if values != expected:
            raise ValueError(f"{label} must keep {module} full+line+connect, got {values}")
    return contract


def validate_config_binding(plan: dict[str, Any]) -> dict[str, Any]:
    """Bind the frozen snapshot, run request, and effective render contract."""

    snapshot_path = Path(str(plan.get("config_snapshot_path", ""))).expanduser().resolve()
    request_path = Path(str(plan.get("run_request_path", ""))).expanduser().resolve()
    result: dict[str, Any] = {
        "status": "failed",
        "passed": False,
        "snapshot_path": str(snapshot_path),
        "request_path": str(request_path),
        "checks": [],
    }
    try:
        snapshot = _read_json(snapshot_path)
        request = _read_json(request_path)
        actual_sha = config_dependency_sha256(snapshot_path).upper()
        contract = _require_full_connect(snapshot, "Candidate configuration snapshot")
        request_contract = _require_full_connect(
            request.get("config") if isinstance(request.get("config"), dict) else {},
            "Embedded run-request configuration",
        )
    except Exception as exc:
        result["message"] = str(exc)
        return result

    checks: list[dict[str, Any]] = result["checks"]

    def record(name: str, passed: bool, detail: Any) -> None:
        checks.append({"name": name, "passed": bool(passed), "detail": detail})

    plan_sha = str(plan.get("config_snapshot_sha256", "")).upper()
    request_sha = str(request.get("config_sha256", "")).upper()
    request_config_path = Path(str(request.get("config_path", ""))).expanduser().resolve()
    record("plan_snapshot_sha256", plan_sha == actual_sha, {"expected": plan_sha, "actual": actual_sha})
    record("request_snapshot_sha256", request_sha == actual_sha, {"request": request_sha, "actual": actual_sha})
    record(
        "request_config_path",
        request_config_path == snapshot_path,
        {"request": str(request_config_path), "snapshot": str(snapshot_path)},
    )
    record("embedded_contract", request_contract == contract, request_contract)
    record("enabled_modules", request.get("enabled_modules") == list(MODULES), request.get("enabled_modules"))
    passed = bool(checks) and all(item["passed"] for item in checks)
    result.update(
        {
            "status": "ok" if passed else "failed",
            "passed": passed,
            "snapshot_sha256": actual_sha,
            "dynamic_plot_contract": contract,
            "message": "configuration binding passed" if passed else "configuration binding failed",
        }
    )
    return result


def _selected_points(
    profile: BridgeProfile, config: dict[str, Any], mode: str
) -> dict[str, list[str]]:
    points = config.get("points")
    if not isinstance(points, dict):
        raise ValueError("Configuration must contain a points object")
    if mode == "representative":
        acceleration = list(profile.representative_acceleration)
        cable = list(profile.representative_cable_accel)
    elif mode == "full":
        acceleration = [str(item) for item in points.get("acceleration", [])]
        cable = [str(item) for item in points.get("cable_accel", [])]
    else:
        raise ValueError("mode must be representative or full")
    if not acceleration or not cable:
        raise ValueError("Selected acceleration and cable-acceleration point sets must be non-empty")
    for key, selected in (
        ("acceleration", acceleration),
        ("accel_spectrum", acceleration),
        ("cable_accel", cable),
        ("cable_accel_spectrum", cable),
    ):
        configured = [str(item) for item in points.get(key, [])]
        missing = [item for item in selected if item not in configured]
        if missing:
            raise ValueError(f"Selected points are absent from points.{key}: {missing}")
    return {"acceleration": acceleration, "cable_accel": cable}


def _scoped_config(
    config: dict[str, Any], selected: dict[str, list[str]], mode: str
) -> dict[str, Any]:
    scoped = copy.deepcopy(config)
    adapter = scoped.setdefault("data_adapter", {})
    if not isinstance(adapter, dict):
        raise ValueError("data_adapter must be an object")
    time_series = adapter.setdefault("time_series", {})
    if not isinstance(time_series, dict):
        raise ValueError("data_adapter.time_series must be an object")
    time_series["source_mode"] = "mat_only"
    if mode == "representative":
        points = scoped.setdefault("points", {})
        points["acceleration"] = list(selected["acceleration"])
        points["accel_spectrum"] = list(selected["acceleration"])
        points["cable_accel"] = list(selected["cable_accel"])
        points["cable_accel_spectrum"] = list(selected["cable_accel"])
        plot_common = scoped.setdefault("plot_common", {})
        if not isinstance(plot_common, dict):
            raise ValueError("plot_common must be an object")
        # Group plots are outside this representative point regression and may
        # otherwise pull non-selected points back into memory.
        plot_common["dynamic_group_sampling_mode"] = "off"
    return scoped


def _analysis_options() -> dict[str, Any]:
    options: dict[str, Any] = {field: False for field in ALL_OPTION_FIELDS}
    for module in MODULES:
        options[MODULE_OPTIONS[module]] = True
    options["input_mode"] = "mat_only"
    return options


def prepare_job(
    bridge: str,
    output_root: Path,
    *,
    source_root: Path | None = None,
    config_path: Path | None = None,
    baseline_root: Path | None = None,
    mode: str = "representative",
    run_id: str | None = None,
    hash_mode: str = "metadata",
    junction_creator: Callable[[Path, Path], None] | None = None,
) -> dict[str, Any]:
    if bridge not in PROFILES:
        raise ValueError(f"Unknown bridge profile: {bridge}")
    profile = PROFILES[bridge]
    source_root = (source_root or profile.source_root).expanduser().resolve()
    config_path = (config_path or profile.config_path).expanduser().resolve()
    baseline_root = (baseline_root or profile.baseline_root).expanduser().resolve()
    output_root = output_root.expanduser().resolve()
    if not config_path.is_file():
        raise FileNotFoundError(f"Configuration does not exist: {config_path}")
    if output_root == source_root or source_root in output_root.parents:
        raise ValueError("Regression output must not be inside the source data root")
    run_id = run_id or (
        f"{bridge}_{profile.start_date.replace('-', '')}_{profile.end_date.replace('-', '')}_"
        f"{mode}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )
    task_root = output_root / run_id
    if task_root.exists():
        raise FileExistsError(f"Regression task root already exists: {task_root}")
    task_root.mkdir(parents=True)

    date_dirs = discover_date_directories(source_root, profile.start_date, profile.end_date)
    inventory = build_source_inventory(source_root, date_dirs, hash_mode=hash_mode)
    control_dir = task_root / "run_logs" / "highfreq_regression"
    inventory_path = control_dir / "source_inventory_before.json"
    _write_json(inventory_path, inventory)

    create_junction = junction_creator or _create_ntfs_junction
    created_links: list[Path] = []
    try:
        for source_day in date_dirs:
            link = task_root / source_day.name
            create_junction(link, source_day)
            created_links.append(link)
    except Exception:
        # Only remove the junction entries created by this invocation.  Never
        # recurse into them and never touch their source targets.
        for link in reversed(created_links):
            if os.name == "nt":
                subprocess.run(
                    ["cmd.exe", "/d", "/c", "rmdir", str(link)],
                    capture_output=True,
                    check=False,
                )
            elif link.is_symlink():
                link.unlink(missing_ok=True)
        raise

    config, dependencies = load_layered_config(config_path)
    source_contract = _require_full_connect(config, "Source configuration")
    selected = _selected_points(profile, config, mode)
    scoped = _scoped_config(config, selected, mode)
    snapshot_path = task_root / "validation_config" / f"{bridge}_highfreq_{mode}.json"
    scoped["source"] = str(snapshot_path)
    _write_json(snapshot_path, scoped)
    snapshot_hash = config_dependency_sha256(snapshot_path)
    frozen_contract = _require_full_connect(scoped, "Frozen configuration")

    status_path = control_dir / "analysis_status.json"
    stop_path = control_dir / "stop.flag"
    request_path = control_dir / "run_request.json"
    async_run_id = f"local-highfreq-{bridge}-{run_id}"
    request = {
        "project_root": str(ROOT),
        "data_root": str(task_root),
        "start_date": profile.start_date,
        "end_date": profile.end_date,
        "config_path": str(snapshot_path),
        "config_sha256": snapshot_hash,
        "options": _analysis_options(),
        "config": scoped,
        "enabled_modules": list(MODULES),
        "async_run_id": async_run_id,
        "stop_file": str(stop_path),
        "async_status_file": str(status_path),
    }
    _write_json(request_path, request)
    _write_json(
        status_path,
        {"status": "prepared", "async_run_id": async_run_id, "launch_performed": False},
    )

    requested_dates = [
        item.date().isoformat()
        for item in _date_range(profile.start_date, profile.end_date)
    ]
    available_dates = [item.name for item in date_dirs]
    plan = {
        "schema_version": 1,
        "plan_type": "local_real_bridge_high_frequency_regression",
        "created_at": _now_iso(),
        "launch_performed": False,
        "bridge": bridge,
        "mode": mode,
        "project_root": str(ROOT),
        "task_root": str(task_root),
        "data_root": str(task_root),
        "source_root": str(source_root),
        "source_read_only_contract": True,
        "baseline_root": str(baseline_root),
        "start_date": profile.start_date,
        "end_date": profile.end_date,
        "requested_dates": requested_dates,
        "available_dates": available_dates,
        "missing_dates": sorted(set(requested_dates) - set(available_dates)),
        "date_junction_count": len(created_links),
        "source_inventory_path": str(inventory_path),
        "source_inventory_sha256": inventory["inventory_sha256"],
        "inventory_hash_mode": hash_mode,
        "source_config_path": str(config_path),
        "source_config_dependencies": [str(item) for item in dependencies],
        "config_snapshot_path": str(snapshot_path),
        "config_snapshot_sha256": snapshot_hash,
        "dynamic_plot_contract": frozen_contract,
        "source_dynamic_plot_contract": source_contract,
        "selected_points": selected,
        "enabled_modules": list(MODULES),
        "run_request_path": str(request_path),
        "analysis_status_path": str(status_path),
        "runner_command": [
            str(ROOT / "bin" / "BridgeAnalysisRunner" / "BridgeAnalysisRunner.exe"),
            str(request_path),
        ],
        "comparison_contract": {
            "stats": "cell_by_cell_selected_points",
            "candidate_plot_provenance": "v2_full_analysis_render_only_reduction_gate",
            "artifact_counts": "selected_point_exact_vs_accepted_baseline",
            "input_inventory": "before_after_exact",
            "visual_review": "paired_candidate_baseline_images_manual_review_required",
        },
    }
    plan_path = task_root / "regression_plan.json"
    plan["plan_path"] = str(plan_path)
    _write_json(plan_path, plan)
    return plan


def _date_range(start: str, end: str) -> Iterable[datetime]:
    from datetime import timedelta

    current = datetime.combine(date.fromisoformat(start), datetime.min.time())
    terminal = datetime.combine(date.fromisoformat(end), datetime.min.time())
    while current <= terminal:
        yield current
        current += timedelta(days=1)


def _cell_equal(left: Any, right: Any, abs_tol: float, rel_tol: float) -> bool:
    if isinstance(left, bool) or isinstance(right, bool):
        return left == right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        if not (math.isfinite(float(left)) and math.isfinite(float(right))):
            return left == right
        return math.isclose(float(left), float(right), abs_tol=abs_tol, rel_tol=rel_tol)
    return left == right


def _compare_values(
    label: str,
    baseline_values: Sequence[Sequence[Any]],
    candidate_values: Sequence[Sequence[Any]],
    *,
    abs_tol: float,
    rel_tol: float,
    max_mismatches: int = 100,
) -> list[dict[str, Any]]:
    mismatches: list[dict[str, Any]] = []
    rows = max(len(baseline_values), len(candidate_values))
    for row_index in range(rows):
        left_row = baseline_values[row_index] if row_index < len(baseline_values) else ()
        right_row = candidate_values[row_index] if row_index < len(candidate_values) else ()
        columns = max(len(left_row), len(right_row))
        for column_index in range(columns):
            left = left_row[column_index] if column_index < len(left_row) else None
            right = right_row[column_index] if column_index < len(right_row) else None
            if not _cell_equal(left, right, abs_tol, rel_tol):
                mismatches.append(
                    {
                        "scope": label,
                        "row": row_index + 1,
                        "column": column_index + 1,
                        "baseline": _json_safe(left),
                        "candidate": _json_safe(right),
                    }
                )
                if len(mismatches) >= max_mismatches:
                    return mismatches
    return mismatches


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def compare_workbook(
    baseline_path: Path,
    candidate_path: Path,
    selected_points: Sequence[str],
    *,
    row_oriented: bool,
    abs_tol: float = 0.0,
    rel_tol: float = 0.0,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required for workbook comparison") from exc
    if not baseline_path.is_file() or not candidate_path.is_file():
        return {
            "status": "failed",
            "baseline": str(baseline_path),
            "candidate": str(candidate_path),
            "message": "baseline or candidate workbook is missing",
            "mismatches": [],
        }
    baseline = load_workbook(baseline_path, read_only=True, data_only=False)
    candidate = load_workbook(candidate_path, read_only=True, data_only=False)
    mismatches: list[dict[str, Any]] = []
    try:
        if row_oriented:
            if not baseline.sheetnames or not candidate.sheetnames:
                raise ValueError("Row-oriented workbook has no worksheet")
            baseline_sheet = baseline[baseline.sheetnames[0]]
            candidate_sheet = candidate[candidate.sheetnames[0]]
            baseline_rows = list(baseline_sheet.iter_rows(values_only=True))
            candidate_rows = list(candidate_sheet.iter_rows(values_only=True))
            if not baseline_rows or not candidate_rows:
                raise ValueError("Row-oriented workbook is empty")
            if tuple(baseline_rows[0]) != tuple(candidate_rows[0]):
                mismatches.append(
                    {
                        "scope": "header",
                        "baseline": [_json_safe(item) for item in baseline_rows[0]],
                        "candidate": [_json_safe(item) for item in candidate_rows[0]],
                    }
                )
            left_by_point = {str(row[0]): tuple(row) for row in baseline_rows[1:] if row and row[0] is not None}
            right_by_point = {str(row[0]): tuple(row) for row in candidate_rows[1:] if row and row[0] is not None}
            if set(right_by_point) != set(selected_points):
                mismatches.append(
                    {
                        "scope": "candidate_point_set",
                        "expected": list(selected_points),
                        "actual": sorted(right_by_point),
                    }
                )
            for point in selected_points:
                left = [left_by_point[point]] if point in left_by_point else []
                right = [right_by_point[point]] if point in right_by_point else []
                mismatches.extend(
                    _compare_values(
                        point, left, right, abs_tol=abs_tol, rel_tol=rel_tol
                    )
                )
        else:
            if set(candidate.sheetnames) != set(selected_points):
                mismatches.append(
                    {
                        "scope": "candidate_sheet_set",
                        "expected": list(selected_points),
                        "actual": candidate.sheetnames,
                    }
                )
            for point in selected_points:
                if point not in baseline.sheetnames or point not in candidate.sheetnames:
                    mismatches.append(
                        {
                            "scope": point,
                            "message": "selected worksheet is missing from baseline or candidate",
                        }
                    )
                    continue
                left = list(baseline[point].iter_rows(values_only=True))
                right = list(candidate[point].iter_rows(values_only=True))
                mismatches.extend(
                    _compare_values(
                        point, left, right, abs_tol=abs_tol, rel_tol=rel_tol
                    )
                )
    finally:
        baseline.close()
        candidate.close()
    return {
        "status": "ok" if not mismatches else "failed",
        "baseline": str(baseline_path),
        "candidate": str(candidate_path),
        "selected_points": list(selected_points),
        "mismatch_count": len(mismatches),
        "mismatches": mismatches[:100],
        "passed": not mismatches,
    }


def _point_pattern(point: str) -> re.Pattern[str]:
    return re.compile(rf"(?<![A-Za-z0-9]){re.escape(point)}(?![A-Za-z0-9])", re.IGNORECASE)


def _artifact_kind(path: Path) -> str | None:
    lower = path.name.casefold()
    if lower.endswith(".plot.json"):
        return "plot_json"
    if path.suffix.casefold() == ".fig":
        return "fig"
    if path.suffix.casefold() in {".jpg", ".jpeg", ".png"}:
        return "image"
    return None


def _count_selected_artifacts(
    root: Path, directory: str, points: Sequence[str], kinds: Sequence[str]
) -> dict[str, dict[str, int]]:
    folder = root / directory
    result = {point: {kind: 0 for kind in kinds} for point in points}
    if not folder.is_dir():
        return result
    patterns = {point: _point_pattern(point) for point in points}
    for path in folder.rglob("*"):
        if not path.is_file():
            continue
        kind = _artifact_kind(path)
        if kind not in kinds:
            continue
        matches = [point for point, pattern in patterns.items() if pattern.search(path.stem)]
        if len(matches) == 1:
            result[matches[0]][kind] += 1
    return result


def compare_artifact_counts(
    baseline_root: Path,
    candidate_root: Path,
    selected: dict[str, list[str]],
) -> dict[str, Any]:
    categories: list[dict[str, Any]] = []
    passed = True
    for label, directory, family, kinds in ARTIFACT_CATEGORIES:
        points = selected[family]
        baseline = _count_selected_artifacts(baseline_root, directory, points, kinds)
        candidate = _count_selected_artifacts(candidate_root, directory, points, kinds)
        equal = baseline == candidate and all(
            count > 0 for values in candidate.values() for count in values.values()
        )
        categories.append(
            {
                "category": label,
                "directory": directory,
                "baseline": baseline,
                "candidate": candidate,
                "passed": equal,
            }
        )
        passed = passed and equal
    return {"status": "ok" if passed else "failed", "categories": categories, "passed": passed}


def validate_candidate_plot_provenance(
    candidate_root: Path,
    selected: dict[str, list[str]],
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    passed = True
    for module, directory, family in (
        ("acceleration", "时程曲线_加速度", "acceleration"),
        ("cable_accel", "时程曲线_索力加速度", "cable_accel"),
    ):
        folder = candidate_root / directory
        paths = sorted(folder.glob("*.plot.json")) if folder.is_dir() else []
        for point in selected[family]:
            matched = [path for path in paths if _point_pattern(point).search(path.stem)]
            if len(matched) != 1:
                rows.append(
                    {
                        "module": module,
                        "point": point,
                        "status": "failed",
                        "message": f"expected one plot.json, found {len(matched)}",
                    }
                )
                passed = False
                continue
            path = matched[0]
            payload = _read_json(path)
            schema = payload.get("schema_version")
            if not isinstance(schema, (int, float)) or int(schema) < 2:
                rows.append(
                    {
                        "module": module,
                        "point": point,
                        "path": str(path),
                        "status": "failed",
                        "message": "candidate raw plot provenance requires schema_version>=2",
                    }
                )
                passed = False
                continue
            inspected = inspect_plot_provenance(module, path)
            rows.append(
                {
                    "module": module,
                    "point": point,
                    "path": str(path),
                    "status": inspected.status,
                    "source_count": inspected.source_count,
                    "plotted_count": inspected.plotted_count,
                    "incomplete_days": list(inspected.incomplete_days),
                    "message": inspected.message,
                    "passed": inspected.closed,
                }
            )
            passed = passed and inspected.closed
    return {"status": "ok" if passed else "failed", "rows": rows, "passed": passed}


def _visual_review_pairs(
    baseline_root: Path,
    candidate_root: Path,
    selected: dict[str, list[str]],
) -> list[dict[str, str]]:
    pairs: list[dict[str, str]] = []
    for _, directory, family, _ in ARTIFACT_CATEGORIES:
        baseline_folder = baseline_root / directory
        candidate_folder = candidate_root / directory
        if not baseline_folder.is_dir() or not candidate_folder.is_dir():
            continue
        for point in selected[family]:
            pattern = _point_pattern(point)
            left = next(
                (
                    path
                    for path in sorted(baseline_folder.rglob("*.jpg"))
                    if pattern.search(path.stem)
                ),
                None,
            )
            right = next(
                (
                    path
                    for path in sorted(candidate_folder.rglob("*.jpg"))
                    if pattern.search(path.stem)
                ),
                None,
            )
            if left is not None and right is not None:
                pairs.append(
                    {
                        "category": directory,
                        "point": point,
                        "baseline": str(left),
                        "candidate": str(right),
                    }
                )
    return pairs


def validate_candidate_manifest(
    candidate_root: Path,
    expected_modules: Sequence[str] = MODULES,
    *,
    expected_config_path: Path | None = None,
    expected_config_sha256: str | None = None,
) -> dict[str, Any]:
    """Require one successful final manifest bound to the expected outputs.

    Artifact counts and provenance can look complete even when the Runner
    ultimately published a failed or partial manifest.  This gate therefore
    treats the final manifest as an independent release contract rather than
    inferring run success from files that happen to exist.
    """

    candidate_root = candidate_root.expanduser().resolve()
    expected = [str(item) for item in expected_modules]
    manifests = sorted((candidate_root / "run_logs").glob("analysis_manifest_*.json"))
    result: dict[str, Any] = {
        "status": "failed",
        "passed": False,
        "expected_modules": expected,
        "manifest_count": len(manifests),
        "manifest_paths": [str(path) for path in manifests],
        "checks": [],
    }
    if len(manifests) != 1:
        result["message"] = (
            f"expected exactly one final analysis manifest, found {len(manifests)}"
        )
        return result

    manifest_path = manifests[0]
    try:
        manifest = _read_json(manifest_path)
    except Exception as exc:
        result["message"] = f"cannot read final analysis manifest: {exc}"
        return result

    checks: list[dict[str, Any]] = result["checks"]

    def record(name: str, passed: bool, detail: Any) -> None:
        checks.append({"name": name, "passed": bool(passed), "detail": detail})

    record("schema_version", manifest.get("schema_version") == 3, manifest.get("schema_version"))
    record("status", manifest.get("status") == "ok", manifest.get("status"))
    enabled = manifest.get("enabled_modules")
    record("enabled_modules", enabled == expected, enabled)
    manifest_request = manifest.get("run_request")
    if expected_config_path is not None:
        expected_config_path = expected_config_path.expanduser().resolve()
        top_path = Path(str(manifest.get("config_path", ""))).expanduser().resolve()
        nested_path = None
        if isinstance(manifest_request, dict):
            nested_path = Path(
                str(manifest_request.get("config_path", ""))
            ).expanduser().resolve()
        record(
            "config_path",
            top_path == expected_config_path and nested_path == expected_config_path,
            {
                "manifest": str(top_path),
                "run_request": str(nested_path) if nested_path is not None else None,
                "expected": str(expected_config_path),
            },
        )
    if expected_config_sha256 is not None:
        nested_sha = ""
        if isinstance(manifest_request, dict):
            nested_sha = str(manifest_request.get("config_sha256", "")).upper()
        expected_sha = str(expected_config_sha256).upper()
        record(
            "config_sha256",
            nested_sha == expected_sha,
            {"manifest_run_request": nested_sha, "expected": expected_sha},
        )
    record(
        "missing_expected_stats",
        manifest.get("missing_expected_stats", []) == [],
        manifest.get("missing_expected_stats"),
    )
    record(
        "missing_stats_files",
        manifest.get("missing_stats_files", []) == [],
        manifest.get("missing_stats_files"),
    )

    module_results = manifest.get("module_results")
    if not isinstance(module_results, list):
        record("module_results", False, "module_results is not a list")
        module_results = []
    selected_results = [
        item
        for item in module_results
        if isinstance(item, dict) and str(item.get("key", "")) in expected
    ]
    keys = [str(item.get("key", "")) for item in selected_results]
    record(
        "module_result_keys",
        len(keys) == len(expected) and set(keys) == set(expected) and len(keys) == len(set(keys)),
        keys,
    )

    module_checks: list[dict[str, Any]] = []
    manifested_paths: set[Path] = set()
    for module in expected:
        matches = [item for item in selected_results if item.get("key") == module]
        if len(matches) != 1:
            module_checks.append(
                {"module": module, "passed": False, "message": f"found {len(matches)} results"}
            )
            continue
        item = matches[0]
        artifacts = item.get("artifacts")
        if not isinstance(artifacts, list):
            artifacts = []
        artifact_rows: list[dict[str, Any]] = []
        artifacts_ok = True
        for artifact in artifacts:
            path_text = artifact.get("path") if isinstance(artifact, dict) else None
            artifact_path = Path(path_text).expanduser().resolve() if path_text else None
            if artifact_path is not None:
                manifested_paths.add(artifact_path)
            inside_root = bool(
                artifact_path is not None
                and (artifact_path == candidate_root or candidate_root in artifact_path.parents)
            )
            exists = bool(artifact_path is not None and artifact_path.is_file())
            actual_bytes = artifact_path.stat().st_size if exists else None
            expected_bytes = artifact.get("bytes") if isinstance(artifact, dict) else None
            size_matches = bool(
                exists
                and isinstance(expected_bytes, (int, float))
                and int(expected_bytes) == actual_bytes
            )
            row_ok = inside_root and exists and size_matches and artifact.get("exists") is True
            artifacts_ok = artifacts_ok and row_ok
            artifact_rows.append(
                {
                    "path": path_text,
                    "inside_candidate_root": inside_root,
                    "exists": exists,
                    "expected_bytes": expected_bytes,
                    "actual_bytes": actual_bytes,
                    "passed": row_ok,
                }
            )
        expected_stats = (candidate_root / "stats" / STATS_FILES[module]).resolve()
        stats_path_text = item.get("stats_path")
        stats_path = Path(stats_path_text).expanduser().resolve() if stats_path_text else None
        artifact_count_matches = item.get("artifact_count") == len(artifacts)
        module_ok = bool(
            item.get("status") == "ok"
            and item.get("stats_exists") is True
            and stats_path == expected_stats
            and expected_stats.is_file()
            and artifact_count_matches
            and artifacts
            and artifacts_ok
        )
        module_checks.append(
            {
                "module": module,
                "status": item.get("status"),
                "stats_path": stats_path_text,
                "expected_stats_path": str(expected_stats),
                "artifact_count": item.get("artifact_count"),
                "actual_artifact_count": len(artifacts),
                "artifact_count_matches": artifact_count_matches,
                "artifacts": artifact_rows,
                "passed": module_ok,
            }
        )
    result["module_checks"] = module_checks
    record("modules", bool(module_checks) and all(row["passed"] for row in module_checks), module_checks)

    # A self-consistent artifact_count does not prove that every formal output
    # was enumerated.  Scan the high-frequency output directories independently
    # so a newly added figure family (for example the 30-minute cable envelope)
    # cannot silently exist on disk while remaining outside the final manifest.
    formal_suffixes = {".jpg", ".jpeg", ".png", ".emf", ".fig"}
    formal_files: set[Path] = set()
    for directory in MANIFEST_FORMAL_DIRS:
        output_dir = candidate_root / directory
        if not output_dir.is_dir():
            continue
        for path in output_dir.rglob("*"):
            if not path.is_file():
                continue
            lower_name = path.name.lower()
            if path.suffix.lower() in formal_suffixes or lower_name.endswith(".plot.json"):
                formal_files.add(path.resolve())
    missing_formal = sorted(str(path) for path in formal_files - manifested_paths)
    record(
        "formal_artifact_coverage",
        not missing_formal,
        {
            "formal_file_count": len(formal_files),
            "manifested_formal_file_count": len(formal_files & manifested_paths),
            "missing": missing_formal,
        },
    )

    passed = bool(checks) and all(item["passed"] for item in checks)
    result.update(
        {
            "status": "ok" if passed else "failed",
            "passed": passed,
            "manifest_path": str(manifest_path),
            "manifest_sha256": _sha256_file(manifest_path),
            "message": "final manifest contract passed" if passed else "final manifest contract failed",
        }
    )
    return result


def compare_job(
    prepared_root: Path,
    *,
    baseline_root: Path | None = None,
    abs_tol: float = 0.0,
    rel_tol: float = 0.0,
) -> dict[str, Any]:
    prepared_root = prepared_root.expanduser().resolve()
    plan = _read_json(prepared_root / "regression_plan.json")
    candidate_root = Path(plan["data_root"]).expanduser().resolve()
    baseline_root = (
        baseline_root.expanduser().resolve()
        if baseline_root is not None
        else Path(plan["baseline_root"]).expanduser().resolve()
    )
    if not baseline_root.is_dir():
        raise FileNotFoundError(f"Accepted baseline does not exist: {baseline_root}")
    selected = plan.get("selected_points")
    if not isinstance(selected, dict):
        raise ValueError("Regression plan lacks selected_points")
    selected = {
        "acceleration": [str(item) for item in selected.get("acceleration", [])],
        "cable_accel": [str(item) for item in selected.get("cable_accel", [])],
    }

    config_binding = validate_config_binding(plan)
    if not config_binding["passed"]:
        config_contract = config_binding.get("dynamic_plot_contract", {})
    else:
        config_contract = config_binding["dynamic_plot_contract"]
    source_root = Path(plan["source_root"]).expanduser().resolve()
    date_dirs = [source_root / name for name in plan["available_dates"]]
    before = _read_json(Path(plan["source_inventory_path"]))
    after = build_source_inventory(
        source_root, date_dirs, hash_mode=str(plan["inventory_hash_mode"])
    )
    after_path = prepared_root / "run_logs" / "highfreq_regression" / "source_inventory_after.json"
    _write_json(after_path, after)
    inventory_result = compare_inventories(before, after)
    manifest_result = validate_candidate_manifest(
        candidate_root,
        expected_config_path=Path(plan["config_snapshot_path"]),
        expected_config_sha256=str(config_binding.get("snapshot_sha256", "")),
    )

    workbook_results = []
    for module, filename in STATS_FILES.items():
        family = "acceleration" if module in {"acceleration", "accel_spectrum"} else "cable_accel"
        workbook_results.append(
            {
                "module": module,
                **compare_workbook(
                    baseline_root / "stats" / filename,
                    candidate_root / "stats" / filename,
                    selected[family],
                    row_oriented=module in {"acceleration", "cable_accel"},
                    abs_tol=abs_tol,
                    rel_tol=rel_tol,
                ),
            }
        )

    provenance_result = validate_candidate_plot_provenance(candidate_root, selected)
    artifact_result = compare_artifact_counts(baseline_root, candidate_root, selected)
    visual_pairs = _visual_review_pairs(baseline_root, candidate_root, selected)
    passed = (
        config_binding["passed"]
        and manifest_result["passed"]
        and inventory_result["passed"]
        and all(item["passed"] for item in workbook_results)
        and provenance_result["passed"]
        and artifact_result["passed"]
    )
    result = {
        "schema_version": 2,
        "comparison_type": "local_real_bridge_high_frequency_regression",
        "created_at": _now_iso(),
        "status": "ok" if passed else "failed",
        "passed": passed,
        "prepared_root": str(prepared_root),
        "candidate_root": str(candidate_root),
        "baseline_root": str(baseline_root),
        "selected_points": selected,
        "numeric_tolerance": {"absolute": abs_tol, "relative": rel_tol},
        "config_contract": config_contract,
        "config_binding": config_binding,
        "analysis_manifest": manifest_result,
        "input_inventory": inventory_result,
        "workbooks": workbook_results,
        "plot_provenance": provenance_result,
        "artifact_counts": artifact_result,
        "visual_review_required": True,
        "visual_review_pairs": visual_pairs,
        "note": (
            "Candidate v2 render vertex counts are intentionally not compared with the v1 "
            "baseline. Numerical statistics are compared cell by cell; visual similarity "
            "must be approved from the paired images before release."
        ),
    }
    output_path = prepared_root / "run_logs" / "highfreq_regression" / "regression_compare.json"
    _write_json(output_path, result)
    result["result_path"] = str(output_path)
    return result


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    commands = parser.add_subparsers(dest="command", required=True)
    prepare = commands.add_parser("prepare", help="prepare an isolated, non-launched regression")
    prepare.add_argument("--bridge", choices=sorted(PROFILES), required=True)
    prepare.add_argument("--output-root", type=Path, required=True)
    prepare.add_argument("--source-root", type=Path)
    prepare.add_argument("--config", type=Path)
    prepare.add_argument("--baseline-root", type=Path)
    prepare.add_argument("--mode", choices=("representative", "full"), default="representative")
    prepare.add_argument("--run-id")
    prepare.add_argument("--inventory-hash", choices=("metadata", "sha256"), default="metadata")

    compare = commands.add_parser("compare", help="compare a completed regression to its baseline")
    compare.add_argument("--prepared-root", type=Path, required=True)
    compare.add_argument("--baseline-root", type=Path)
    compare.add_argument("--abs-tol", type=float, default=0.0)
    compare.add_argument("--rel-tol", type=float, default=0.0)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    try:
        if args.command == "prepare":
            result = prepare_job(
                args.bridge,
                args.output_root,
                source_root=args.source_root,
                config_path=args.config,
                baseline_root=args.baseline_root,
                mode=args.mode,
                run_id=args.run_id,
                hash_mode=args.inventory_hash,
            )
            summary = {
                "status": "prepared",
                "launch_performed": False,
                "task_root": result["task_root"],
                "run_request_path": result["run_request_path"],
                "date_junction_count": result["date_junction_count"],
                "source_inventory_sha256": result["source_inventory_sha256"],
            }
        else:
            result = compare_job(
                args.prepared_root,
                baseline_root=args.baseline_root,
                abs_tol=args.abs_tol,
                rel_tol=args.rel_tol,
            )
            summary = {
                "status": result["status"],
                "passed": result["passed"],
                "result_path": result["result_path"],
            }
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0 if summary.get("passed", True) else 2
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
