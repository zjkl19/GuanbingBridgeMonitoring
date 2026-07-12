from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


EXPECTED_MODULES = {
    "wind",
    "earthquake",
    "acceleration",
    "cable_accel",
    "accel_spectrum",
    "cable_accel_spectrum",
}
EXPECTED_PROVENANCE_COUNTS = {
    "wind": 4,
    "earthquake": 3,
    "acceleration": 12,
    "cable_accel": 24,
}
EXPECTED_STATS_FILES = {
    "wind_stats.xlsx",
    "eq_stats.xlsx",
    "accel_stats.xlsx",
    "accel_spec_stats.xlsx",
    "cable_accel_stats.xlsx",
    "cable_accel_spec_stats.xlsx",
}
EXPECTED_ACCEL_POINTS = {
    *(f"A{i}" for i in range(1, 9)),
    "A9-X",
    "A9-Y",
    "A10-X",
    "A10-Y",
}
EXPECTED_CABLE_POINTS = {
    *(f"CS{i}" for i in range(1, 13)),
    *(f"CX{i}" for i in range(1, 13)),
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def norm(path: str | Path) -> str:
    return os.path.normcase(os.path.abspath(os.fspath(path)))


def is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def workbook_signature(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, object] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = [
            tuple(row)
            for row in sheet.iter_rows(values_only=True)
            if any(value is not None and str(value).strip() for value in row)
        ]
        sheets[sheet_name] = {
            "nonempty_row_count": len(rows),
            "header": list(rows[0]) if rows else [],
            "first_column_values": [
                str(row[0]) for row in rows[1:] if row and row[0] is not None
            ],
        }
    workbook.close()
    return {"sheet_names": list(sheets), "sheets": sheets}


def latest_manifest(root: Path) -> Path:
    candidates = sorted(
        (root / "run_logs").glob("analysis_manifest_*.json"),
        key=lambda path: path.stat().st_mtime,
    )
    if not candidates:
        raise FileNotFoundError(f"No analysis manifest under {root / 'run_logs'}")
    return candidates[-1]


def validate(root: Path, manifest_path: Path, repo_root: Path) -> dict[str, object]:
    sys.path.insert(0, str(repo_root / "reporting"))
    from locked_docx_media import validate_full_plot_provenance

    manifest = json.loads(manifest_path.read_text(encoding="utf-8-sig"))
    modules = manifest.get("module_results") or []
    module_statuses = {
        str(module.get("key")): str(module.get("status")) for module in modules
    }
    errors: list[str] = []
    warnings: list[str] = []

    if manifest.get("status") != "ok":
        errors.append(f"manifest status is {manifest.get('status')!r}")
    missing_modules = EXPECTED_MODULES - module_statuses.keys()
    if missing_modules:
        errors.append(f"missing expected modules: {sorted(missing_modules)}")
    failed_modules = {
        key: status
        for key, status in module_statuses.items()
        if key in EXPECTED_MODULES and status != "ok"
    }
    if failed_modules:
        errors.append(f"non-ok expected modules: {failed_modules}")

    sampling = ((manifest.get("run_request") or {}).get("plot_sampling") or {})
    expected_sampling = {
        "mode": "full",
        "group_mode": "full",
        "render_mode": "line",
        "raw_emf_disabled": True,
    }
    for key, expected in expected_sampling.items():
        if sampling.get(key) != expected:
            errors.append(
                f"plot_sampling.{key}={sampling.get(key)!r}, expected {expected!r}"
            )

    plot_rows: list[dict[str, object]] = []
    module_provenance_counts: dict[str, int] = {}
    point_ids: set[str] = set()
    stats_paths: dict[str, Path] = {}

    for module in modules:
        module_key = str(module.get("key"))
        artifacts = module.get("artifacts") or []
        artifact_paths = {
            norm(item.get("path")) for item in artifacts if item.get("path")
        }
        provenance_items = [
            item for item in artifacts if item.get("kind") == "plot_provenance"
        ]
        if provenance_items:
            module_provenance_counts[module_key] = len(provenance_items)

        for artifact in artifacts:
            raw_path = artifact.get("path")
            if not raw_path:
                continue
            artifact_path = Path(str(raw_path))
            if module_key in EXPECTED_MODULES and not is_within(artifact_path, root):
                errors.append(f"artifact escapes event root: {artifact_path}")
            if artifact.get("kind") == "stats":
                stats_paths[artifact_path.name] = artifact_path

        for artifact in provenance_items:
            provenance_path = Path(str(artifact.get("path")))
            jpg_path = provenance_path.with_name(
                provenance_path.name[: -len(".plot.json")] + ".jpg"
            )
            if norm(jpg_path) not in artifact_paths:
                errors.append(
                    f"JPG is not bound to the same module record: {jpg_path}"
                )
                continue
            require_source = module_key in {"acceleration", "cable_accel"}
            try:
                series_count = validate_full_plot_provenance(
                    provenance_path,
                    jpg_path,
                    require_source_provenance=require_source,
                )
            except Exception as exc:
                errors.append(f"provenance failed for {provenance_path}: {exc}")
                continue

            payload = json.loads(provenance_path.read_text(encoding="utf-8-sig"))
            series = payload.get("series") or []
            if isinstance(series, dict):
                series = [series]
            if len(series) != series_count:
                errors.append(f"series count mismatch: {provenance_path}")
            series_rows: list[dict[str, object]] = []
            for item in series:
                point_id = str(item.get("point_id") or "")
                if point_id:
                    point_ids.add(point_id)
                input_count = int(item.get("input_count", -1))
                finite_count = int(item.get("finite_count", -1))
                plotted_count = int(item.get("plotted_finite_count", -1))
                if input_count <= 0 or finite_count <= 0 or plotted_count <= 0:
                    errors.append(f"empty plot series in {provenance_path}")
                if finite_count != plotted_count:
                    errors.append(
                        f"finite/plotted mismatch for {point_id}: {provenance_path}"
                    )
                row: dict[str, object] = {
                    "point_id": point_id,
                    "input_count": input_count,
                    "finite_count": finite_count,
                    "plotted_finite_count": plotted_count,
                }
                if require_source:
                    source = item.get("source") or {}
                    source_count = int(source.get("source_sample_count", -1))
                    if source_count != input_count:
                        errors.append(
                            f"source/input mismatch for {point_id}: {provenance_path}"
                        )
                    if int(source.get("conflicting_timestamp_count", -1)) != 0:
                        errors.append(
                            f"timestamp conflicts for {point_id}: {provenance_path}"
                        )
                    row.update(
                        {
                            "source_sample_count": source_count,
                            "coverage_start": source.get("coverage_start"),
                            "coverage_end": source.get("coverage_end"),
                            "incomplete_days": source.get("incomplete_days") or [],
                            "missing_required_sources": source.get(
                                "missing_required_sources"
                            )
                            or [],
                        }
                    )
                series_rows.append(row)

            try:
                with Image.open(jpg_path) as image:
                    width, height = image.size
                    image_format = image.format
                    image.verify()
            except Exception as exc:
                errors.append(f"invalid image {jpg_path}: {exc}")
                continue
            if width < 1200 or height < 650:
                errors.append(f"image is unexpectedly small {width}x{height}: {jpg_path}")
            plot_rows.append(
                {
                    "module": module_key,
                    "jpg": str(jpg_path),
                    "provenance": str(provenance_path),
                    "jpg_sha256": sha256_file(jpg_path),
                    "provenance_sha256": sha256_file(provenance_path),
                    "width_px": width,
                    "height_px": height,
                    "format": image_format,
                    "series": series_rows,
                }
            )

    for module_key, expected in EXPECTED_PROVENANCE_COUNTS.items():
        actual = module_provenance_counts.get(module_key)
        if actual != expected:
            errors.append(
                f"{module_key} provenance count is {actual}, expected {expected}"
            )

    expected_points = EXPECTED_ACCEL_POINTS | EXPECTED_CABLE_POINTS
    missing_points = sorted(expected_points - point_ids)
    if missing_points:
        errors.append(f"missing formal high-frequency points: {missing_points}")

    missing_stats = EXPECTED_STATS_FILES - stats_paths.keys()
    if missing_stats:
        errors.append(f"missing stats workbooks: {sorted(missing_stats)}")
    stats_rows: list[dict[str, object]] = []
    for name in sorted(EXPECTED_STATS_FILES & stats_paths.keys()):
        path = stats_paths[name]
        if not path.is_file():
            errors.append(f"missing stats workbook: {path}")
            continue
        signature = workbook_signature(path)
        if any(
            sheet["nonempty_row_count"] <= 1
            for sheet in signature["sheets"].values()
        ):
            errors.append(f"empty stats sheet: {path}")
        stats_rows.append(
            {
                "path": str(path),
                "sha256": sha256_file(path),
                "signature": signature,
            }
        )

    return {
        "ok": not errors,
        "event_root": str(root),
        "manifest": str(manifest_path),
        "manifest_sha256": sha256_file(manifest_path),
        "manifest_status": manifest.get("status"),
        "module_statuses": module_statuses,
        "plot_sampling": sampling,
        "module_provenance_counts": module_provenance_counts,
        "plot_count": len(plot_rows),
        "point_ids": sorted(point_ids),
        "plots": plot_rows,
        "stats": stats_rows,
        "warnings": warnings,
        "errors": errors,
    }


def parser() -> argparse.ArgumentParser:
    value = argparse.ArgumentParser(
        description="Validate a Hongtang typhoon high-frequency analysis root"
    )
    value.add_argument("--root", type=Path, required=True)
    value.add_argument("--manifest", type=Path)
    value.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parents[1])
    value.add_argument("--output", type=Path)
    return value


def main() -> None:
    args = parser().parse_args()
    manifest_path = args.manifest or latest_manifest(args.root)
    result = validate(args.root, manifest_path, args.repo_root)
    output = args.output or args.root / "run_logs" / "typhoon_highfreq_validation.json"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "ok": result["ok"],
                "manifest": result["manifest"],
                "module_provenance_counts": result["module_provenance_counts"],
                "plot_count": result["plot_count"],
                "stats_count": len(result["stats"]),
                "errors": result["errors"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    raise SystemExit(0 if result["ok"] else 2)


if __name__ == "__main__":
    main()
