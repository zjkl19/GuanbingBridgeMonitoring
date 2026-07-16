"""Offline helpers for recovering high-memory analysis modules.

The commands in this module only prepare requests or consolidate already
completed outputs.  They never launch MATLAB/BridgeAnalysisRunner.  All
runtime and artifact paths are constrained to an explicitly supplied RC tree.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import math
import ntpath
import os
import re
import shutil
import sys
import tempfile
from copy import copy as copy_style
from datetime import datetime, timezone
from pathlib import Path, PureWindowsPath
from typing import Any, Iterable, Sequence

try:
    from workbench.provenance import inspect_plot_provenance
except ModuleNotFoundError:  # direct ``python scripts/high_memory_recovery.py``
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
    from workbench.provenance import inspect_plot_provenance

SUCCESS_STATES = {"ok", "completed", "complete", "success", "succeeded"}
RECOVERY_MODULES = ("cable_accel", "accel_spectrum", "cable_accel_spectrum")
RECOVERY_AUXILIARY_MODULES = {"offset_correction_report"}
MODULE_OPTION_FIELDS = {
    "temperature": "doTemp",
    "humidity": "doHumidity",
    "rainfall": "doRainfall",
    "gnss": "doGNSS",
    "wind": "doWind",
    "earthquake": "doEq",
    "deflection": "doDeflect",
    "bearing_displacement": "doBearingDisplacement",
    "tilt": "doTilt",
    "crack": "doCrack",
    "strain": "doStrain",
    "acceleration": "doAccel",
    "cable_accel": "doCableAccel",
    "accel_spectrum": "doAccelSpectrum",
    "cable_accel_spectrum": "doCableAccelSpectrum",
}


def parse_recovery_started_at(value: str) -> datetime:
    """Parse a recovery start instant across PowerShell/native CLI boundaries.

    Windows PowerShell 5.1 normally removes command-line quoting before Python
    receives an argument.  Some wrapper/launcher combinations can instead
    preserve one matching pair of single or double quotes.  Accept exactly
    that compatibility form while continuing to require an explicit UTC
    offset and returning a normalized UTC datetime.
    """

    text = str(value).strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in {"'", '"'}:
        text = text[1:-1].strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    # .NET's round-trip "o" format emits seven fractional-second digits,
    # while Python 3.10 datetime.fromisoformat accepts at most six. Preserve
    # microsecond precision and discard only the unsupported sub-microsecond
    # remainder before parsing.
    text = re.sub(r"(?<=\.\d{6})\d+(?=[+-]\d{2}:\d{2}$)", "", text)
    if not text:
        raise ValueError("Recovery start instant is empty")
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError as exc:
        raise ValueError(f"Invalid recovery start instant: {value!r}") from exc
    if parsed.tzinfo is None:
        raise ValueError("Recovery start instant must contain a UTC offset")
    return parsed.astimezone(timezone.utc)


def _json_object(path: Path | str, label: str = "JSON") -> dict[str, Any]:
    source = Path(path)
    try:
        value = json.loads(source.read_text(encoding="utf-8-sig"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise ValueError(f"{label} must be valid UTF-8 JSON: {source}: {exc}") from exc
    if not isinstance(value, dict):
        raise ValueError(f"{label} root must be an object: {source}")
    return value


def _json_bytes(value: Any, *, pretty: bool = True) -> bytes:
    if pretty:
        text = json.dumps(value, ensure_ascii=False, indent=2) + "\n"
    else:
        text = json.dumps(value, ensure_ascii=False, separators=(",", ":")) + "\n"
    return text.encode("utf-8")


def _atomic_write(path: Path, raw: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(handle, "wb") as stream:
            stream.write(raw)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(name, path)
    finally:
        Path(name).unlink(missing_ok=True)


def _sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest().upper()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _windows_normal(path: str) -> str:
    return ntpath.normcase(ntpath.normpath(str(PureWindowsPath(path))))


def _windows_within(path: str, root: str) -> bool:
    path_n = _windows_normal(path)
    root_n = _windows_normal(root)
    try:
        return ntpath.commonpath([path_n, root_n]) == root_n
    except ValueError:
        return False


def _require_remote_path(path: str, rc_root: str, label: str) -> None:
    if not PureWindowsPath(path).is_absolute() or not _windows_within(path, rc_root):
        raise ValueError(f"{label} must be an absolute path inside RC root {rc_root}: {path}")


def _require_local_path(path: Path | str, root: Path, label: str) -> Path:
    resolved = Path(path).expanduser().resolve()
    root = root.expanduser().resolve()
    try:
        resolved.relative_to(root)
    except ValueError as exc:
        raise ValueError(f"{label} is outside allowed RC root {root}: {resolved}") from exc
    return resolved


def _iter_absolute_windows_strings(value: Any, prefix: str = "") -> Iterable[tuple[str, str]]:
    if isinstance(value, dict):
        for key, item in value.items():
            child = f"{prefix}.{key}" if prefix else str(key)
            yield from _iter_absolute_windows_strings(item, child)
    elif isinstance(value, list):
        for index, item in enumerate(value):
            yield from _iter_absolute_windows_strings(item, f"{prefix}[{index}]")
    elif isinstance(value, str) and PureWindowsPath(value).is_absolute():
        yield prefix, value


def _sanitize_inactive_preprocessing_paths(config: dict[str, Any], rc_root: str) -> None:
    preprocessing = config.get("preprocessing")
    if not isinstance(preprocessing, dict):
        return

    def walk(container: Any) -> None:
        if isinstance(container, dict):
            for key in list(container):
                value = container[key]
                if isinstance(value, str) and PureWindowsPath(value).is_absolute():
                    if not _windows_within(value, rc_root):
                        del container[key]
                else:
                    walk(value)
        elif isinstance(container, list):
            for item in container:
                walk(item)

    walk(preprocessing)


def _module_options(base: dict[str, Any], enabled: str) -> dict[str, Any]:
    result = copy.deepcopy(base)
    for key in list(result):
        normalized = str(key).casefold()
        if normalized.startswith("precheck") or normalized.startswith("do"):
            result[key] = False
    option = {
        "cable_accel": "doCableAccel",
        "accel_spectrum": "doAccelSpectrum",
        "cable_accel_spectrum": "doCableAccelSpectrum",
    }[enabled]
    result[option] = True
    result["input_mode"] = "mat_only"
    return result


def _cable_points(config: dict[str, Any]) -> list[str]:
    points = config.get("points")
    if not isinstance(points, dict):
        raise ValueError("Config must contain a points object")
    for key in ("cable_accel", "cable_force"):
        value = points.get(key)
        if isinstance(value, list) and value:
            normalized = [str(item).strip() for item in value]
            if any(not item for item in normalized) or len(set(normalized)) != len(normalized):
                raise ValueError(f"Config points.{key} contains blank or duplicate point IDs")
            return normalized
    raise ValueError("Config must contain non-empty points.cable_accel or points.cable_force")


def _require_mat_only_config(config: dict[str, Any], label: str) -> None:
    adapter = config.get("data_adapter")
    time_series = adapter.get("time_series") if isinstance(adapter, dict) else None
    source_mode = time_series.get("source_mode") if isinstance(time_series, dict) else None
    if str(source_mode or "").strip().casefold() != "mat_only":
        raise ValueError(f"{label} must pin data_adapter.time_series.source_mode=mat_only")


def _require_jiulongjiang_config(config: dict[str, Any], label: str) -> None:
    if str(config.get("vendor") or "").strip().casefold() != "jiulongjiang":
        raise ValueError(f"{label} must pin vendor=jiulongjiang")


def _dynamic_plot_contract(config: dict[str, Any]) -> dict[str, dict[str, str]]:
    common = config.get("plot_common") if isinstance(config.get("plot_common"), dict) else {}
    modules = common.get("dynamic_raw_modules") \
        if isinstance(common.get("dynamic_raw_modules"), dict) else {}
    contract: dict[str, dict[str, str]] = {}
    for module in ("acceleration", "cable_accel"):
        override = modules.get(module) if isinstance(modules.get(module), dict) else {}
        sampling = override.get("sampling_mode", common.get("dynamic_raw_sampling_mode", "capped"))
        gap = override.get("gap_mode", common.get("gap_mode", "connect"))
        contract[module] = {
            "sampling_mode": str(sampling or "").strip().casefold(),
            "gap_mode": str(gap or "").strip().casefold(),
        }
    return contract


def _require_full_connect_contract(
    config: dict[str, Any], label: str,
    *, expected: dict[str, dict[str, str]] | None = None,
) -> dict[str, dict[str, str]]:
    contract = _dynamic_plot_contract(config)
    for module, values in contract.items():
        if values != {"sampling_mode": "full", "gap_mode": "connect"}:
            raise ValueError(f"{label} must keep {module} full+connect: {values}")
    relevant_keys = {
        "acceleration": {"acceleration", "acceleration_raw"},
        "cable_accel": {"cable_accel", "cable_accel_raw", "cable_force"},
    }
    for container_name in ("plot_styles",):
        container = config.get(container_name)
        if not isinstance(container, dict):
            continue
        for module, keys in relevant_keys.items():
            for key in keys:
                value = container.get(key)
                if isinstance(value, dict) and "gap_mode" in value \
                        and str(value.get("gap_mode") or "").strip().casefold() != "connect":
                    raise ValueError(
                        f"{label} has a non-connect {container_name}.{key}.gap_mode"
                    )
    per_point = config.get("per_point")
    if isinstance(per_point, dict):
        for module, keys in relevant_keys.items():
            for key in keys:
                points = per_point.get(key)
                if not isinstance(points, dict):
                    continue
                for point_id, point_cfg in points.items():
                    plot = point_cfg.get("plot") if isinstance(point_cfg, dict) else None
                    if isinstance(plot, dict) and "gap_mode" in plot \
                            and str(plot.get("gap_mode") or "").strip().casefold() != "connect":
                        raise ValueError(
                            f"{label} has a non-connect per_point.{key}.{point_id}.plot.gap_mode"
                        )
    if expected is not None and contract != expected:
        raise ValueError(f"{label} dynamic plot contract differs from baseline")
    return contract


def _normalize_group_map(raw: Any) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = {}
    if isinstance(raw, dict):
        items = raw.items()
    elif isinstance(raw, list):
        items = ((f"G{index}", value) for index, value in enumerate(raw, start=1))
    else:
        return groups
    for name, value in items:
        values = value if isinstance(value, list) else [value]
        points: list[str] = []
        for item in values:
            if isinstance(item, str) and item.strip() and item.strip() not in points:
                points.append(item.strip())
        if points:
            groups[str(name)] = points
    return groups


def _effective_cable_groups(config: dict[str, Any]) -> dict[str, list[str]]:
    groups = config.get("groups") if isinstance(config.get("groups"), dict) else {}
    # MATLAB only falls back to cable_force when groups.cable_accel is absent;
    # an explicitly present-but-empty cable_accel block suppresses fallback.
    if "cable_accel" in groups:
        return _normalize_group_map(groups.get("cable_accel"))
    return _normalize_group_map(groups.get("cable_force"))


def _point_list(config: dict[str, Any], keys: Sequence[str], label: str) -> list[str]:
    """Mirror SpectrumConfigService's first-nonempty point-key lookup."""

    points = config.get("points")
    if not isinstance(points, dict):
        raise ValueError(f"{label} requires a points object")
    for key in keys:
        raw = points.get(key)
        if not isinstance(raw, list) or not raw:
            continue
        values = [str(item).strip() for item in raw]
        if any(not value for value in values) or len(values) != len(set(values)):
            raise ValueError(f"{label} points.{key} contains blank or duplicate IDs")
        return values
    raise ValueError(
        f"{label} must resolve an explicit non-empty point list from {list(keys)}"
    )


def _point_config(config: dict[str, Any], section: str, point_id: str) -> dict[str, Any]:
    per_point = config.get("per_point")
    values = per_point.get(section) if isinstance(per_point, dict) else None
    if not isinstance(values, dict):
        return {}
    candidates = [point_id, point_id.replace("-", "_")]
    name_map = config.get("name_map_global")
    if isinstance(name_map, dict):
        candidates.extend(
            str(key) for key, value in name_map.items() if str(value) == point_id
        )
    for candidate in dict.fromkeys(candidates):
        value = values.get(candidate)
        if isinstance(value, dict):
            return value
    return {}


def _valid_number(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    try:
        return math.isfinite(float(value))
    except (TypeError, ValueError):
        return False


def _valid_peak_order_count(raw: Any) -> int:
    if isinstance(raw, dict):
        raw = [raw]
    if not isinstance(raw, list):
        return 0
    count = 0
    for item in raw:
        if not isinstance(item, dict):
            continue
        minimum = item.get("search_min_hz", item.get("min_hz", item.get("lower_hz")))
        maximum = item.get("search_max_hz", item.get("max_hz", item.get("upper_hz")))
        center = next(
            (
                item.get(key) for key in (
                    "search_center_hz", "target_hz", "frequency_hz", "freq_hz",
                    "theoretical_hz", "theor_hz",
                ) if _valid_number(item.get(key))
            ),
            None,
        )
        has_range = (
            _valid_number(minimum) and _valid_number(maximum)
            and float(maximum) > float(minimum)
        )
        if center is not None or has_range:
            count += 1
    return count


def _numeric_list_count(raw: Any) -> int:
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return int(_valid_number(raw))
    if not isinstance(raw, list):
        return 0
    return sum(_valid_number(value) for value in raw)


def _spectrum_peak_count(config: dict[str, Any], module: str, point_id: str) -> int:
    if module == "accel_spectrum":
        params_key, point_section = "accel_spectrum_params", "accel_spectrum"
    elif module == "cable_accel_spectrum":
        params_key, point_section = "cable_accel_spectrum_params", "cable_accel"
    else:  # pragma: no cover - private API guard
        raise ValueError(f"Unsupported spectrum module: {module}")
    params = config.get(params_key) if isinstance(config.get(params_key), dict) else {}
    count = _valid_peak_order_count(params.get("peak_orders"))
    if not count:
        count = _numeric_list_count(params.get("target_freqs"))
    if not count:
        count = 3  # SpectrumAnalysisPipeline's current default target-frequency count.
    point = _point_config(config, point_section, point_id)
    point_count = _valid_peak_order_count(point.get("peak_orders"))
    if not point_count:
        point_count = _numeric_list_count(point.get("target_freqs"))
    return point_count or count


def _cable_force_parameter_evidence(
    config: dict[str, Any], point_id: str,
) -> dict[str, Any]:
    point = _point_config(config, "cable_accel", point_id)
    rho = point.get("rho")
    length = point.get("L")
    code_valid = _valid_number(rho) and _valid_number(length)
    raw_status = str(
        point.get("force_parameter_status")
        or point.get("force_parameters_status")
        or ""
    ).strip().casefold()
    verified_statuses = {"verified", "engineering_verified", "approved"}
    engineering_valid = code_valid and raw_status in verified_statuses
    if not code_valid:
        status = "missing_or_invalid_parameters"
    elif (
        float(rho) == 1.0 and float(length) == 1.0
        and not engineering_valid
    ):
        status = "placeholder_parameters"
    elif engineering_valid:
        status = "engineering_verified"
    else:
        status = "unverified_parameters"
    return {
        "point_id": point_id,
        "rho": rho if _valid_number(rho) else None,
        "L": length if _valid_number(length) else None,
        "code_valid": code_valid,
        "engineering_valid": engineering_valid,
        "parameter_status": status,
    }


def _safe_figure_name(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", str(value))


def _scope_tokens(start_date: str, end_date: str) -> tuple[str, str]:
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("Spectrum recovery scope must use YYYY-MM-DD dates") from exc
    if end < start:
        raise ValueError("Spectrum recovery end_date precedes start_date")
    return start.strftime("%Y%m%d"), end.strftime("%Y%m%d")


def _style_groups(config: dict[str, Any], style_key: str, fallback_key: str) -> dict[str, list[str]]:
    styles = config.get("plot_styles")
    style = styles.get(style_key) if isinstance(styles, dict) else None
    raw: Any = None
    if isinstance(style, dict):
        for key in ("groups", "group_points"):
            candidate = style.get(key)
            if _normalize_group_map(candidate):
                raw = candidate
                break
    if raw is None:
        groups = config.get("groups") if isinstance(config.get("groups"), dict) else {}
        raw = groups.get(fallback_key)
    return _normalize_group_map(raw)


def _styled_group_name(
    config: dict[str, Any], style_key: str, group_name: str, points: Sequence[str],
) -> str:
    styles = config.get("plot_styles")
    style = styles.get(style_key) if isinstance(styles, dict) else None
    labels = style.get("group_labels") if isinstance(style, dict) else None
    if isinstance(labels, dict) and str(labels.get(group_name) or "").strip():
        return str(labels[group_name])
    return "-".join(points) if len(points) <= 4 else group_name


def _spectrum_output_contract(
    config: dict[str, Any], module: str, start_date: str, end_date: str,
) -> dict[str, Any]:
    """Derive every formal spectrum sheet/figure expected from the pinned config."""

    start_token, end_token = _scope_tokens(start_date, end_date)
    if module == "accel_spectrum":
        points = _point_list(
            config, ("accel_spectrum", "acceleration"), "Acceleration spectrum",
        )
        groups = _style_groups(config, "accel_spectrum", "acceleration")
        spec_stubs = [
            f"SpecFreq_{point}_{start_token}_{end_token}" for point in points
        ]
        group_stubs: list[str] = []
        group_contract: list[dict[str, Any]] = []
        for name, configured_points in groups.items():
            members = [point for point in configured_points if point in points]
            if not members:
                raise ValueError(
                    f"Acceleration spectrum group {name} has no configured spectrum points"
                )
            peak_count = max(
                _spectrum_peak_count(config, module, point) for point in members
            )
            display = _safe_figure_name(
                _styled_group_name(config, "accel_spectrum", name, members)
            )
            stubs = [
                f"SpecFreq_{display}_Group{'' if index == 1 else f'_P{index}'}_"
                f"{start_token}_{end_token}"
                for index in range(1, peak_count + 1)
            ]
            group_stubs.extend(stubs)
            group_contract.append({
                "group_name": name,
                "point_ids": members,
                "expected_figure_stubs": stubs,
            })
        return {
            "configured_points": points,
            "expected_stat_sheets": points,
            "expected_specfreq_stubs": spec_stubs,
            "valid_force_points": [],
            "expected_force_stubs": [],
            "groups": group_contract,
            "expected_group_stubs": group_stubs,
            "expected_formal_figure_stubs": [*spec_stubs, *group_stubs],
        }
    if module != "cable_accel_spectrum":  # pragma: no cover - private API guard
        raise ValueError(f"Unsupported spectrum module: {module}")

    points = _point_list(
        config, ("cable_accel_spectrum", "cable_accel", "cable_force"),
        "Cable acceleration spectrum",
    )
    spec_stubs = [f"SpecFreq_{point}_{start_token}_{end_token}" for point in points]
    force_parameter_evidence = [
        _cable_force_parameter_evidence(config, point) for point in points
    ]
    valid_force_points = [
        item["point_id"] for item in force_parameter_evidence if item["code_valid"]
    ]
    force_stubs = [
        f"CableForce_{point}_{start_token}_{end_token}" for point in valid_force_points
    ]
    groups_cfg = config.get("groups") if isinstance(config.get("groups"), dict) else {}
    # SpectrumPlotService calls getGroups(cfg, 'cable_force').  The registry
    # resolves an explicit cable_force block first, then the canonical
    # cable_accel group alias when cable_force is absent.
    groups = _normalize_group_map(
        groups_cfg.get("cable_force")
        if "cable_force" in groups_cfg else groups_cfg.get("cable_accel")
    )
    group_stubs: list[str] = []
    group_contract: list[dict[str, Any]] = []
    for name, configured_points in groups.items():
        members = [point for point in configured_points if point in valid_force_points]
        if not members:
            raise ValueError(f"Cable force group {name} has no point with valid rho/L")
        display = _safe_figure_name("-".join(members) if len(members) <= 4 else name)
        stub = f"CableForce_{display}_{start_token}_{end_token}"
        group_stubs.append(stub)
        group_contract.append({
            "group_name": name,
            "point_ids": members,
            "expected_figure_stubs": [stub],
        })
    engineering_valid = bool(force_parameter_evidence) and all(
        item["engineering_valid"] for item in force_parameter_evidence
    )
    parameter_statuses = {
        item["parameter_status"] for item in force_parameter_evidence
    }
    if engineering_valid:
        engineering_status = "engineering_verified"
    elif "placeholder_parameters" in parameter_statuses:
        engineering_status = "placeholder_parameters"
    else:
        engineering_status = "unverified_parameters"
    return {
        "configured_points": points,
        "expected_stat_sheets": points,
        "expected_specfreq_stubs": spec_stubs,
        "valid_force_points": valid_force_points,
        "force_parameter_evidence": force_parameter_evidence,
        "cable_force_engineering_valid": engineering_valid,
        "cable_force_engineering_status": engineering_status,
        "expected_force_stubs": force_stubs,
        "groups": group_contract,
        "expected_group_stubs": group_stubs,
        "expected_formal_figure_stubs": [*spec_stubs, *force_stubs, *group_stubs],
    }


def _restrict_cable_point(config: dict[str, Any], point_id: str) -> None:
    points = config.setdefault("points", {})
    points["cable_accel"] = [point_id]
    points["cable_force"] = [point_id]
    # DynamicAccelerationSeriesService performs configured/fallback group
    # plotting after the point loop.  Without this explicit override a
    # seemingly point-isolated request can reload the whole cable group and
    # reproduce the original OOM.  Group figures require a separate audited
    # recovery step and are never produced by these point jobs.
    plot_common = config.setdefault("plot_common", {})
    if not isinstance(plot_common, dict):
        raise ValueError("Config plot_common must be an object")
    plot_common["dynamic_group_sampling_mode"] = "off"
    per_point = config.get("per_point")
    if isinstance(per_point, dict):
        for key in ("cable_accel", "cable_force"):
            values = per_point.get(key)
            if isinstance(values, dict):
                per_point[key] = {point_id: values[point_id]} if point_id in values else {}


def prepare_recovery_requests(
    base_request_path: Path | str,
    output_dir: Path | str,
    *,
    rc_root: str,
    remote_bundle_root: str,
    base_config_path: Path | str | None = None,
    expected_cable_points: int = 15,
) -> dict[str, Any]:
    """Create 15 point-isolated cable requests and two spectrum requests.

    ``output_dir`` is a local/offline bundle mirror.  Every path embedded in
    generated JSON points at ``remote_bundle_root`` inside ``rc_root``.  No
    process is started.
    """

    _require_remote_path(remote_bundle_root, rc_root, "Remote bundle root")
    base_request = _json_object(base_request_path, "Base run request")
    if base_config_path is not None:
        base_config = _json_object(base_config_path, "Base config")
    else:
        embedded = base_request.get("config")
        if not isinstance(embedded, dict):
            raise ValueError("Base request must embed config or --base-config must be supplied")
        base_config = copy.deepcopy(embedded)

    _require_mat_only_config(base_config, "Base recovery config")
    _require_jiulongjiang_config(base_config, "Base recovery config")
    plot_contract = _require_full_connect_contract(base_config, "Base recovery config")
    effective_cable_groups = _effective_cable_groups(base_config)
    start_date = str(base_request.get("start_date") or "").strip()
    end_date = str(base_request.get("end_date") or "").strip()
    accel_spectrum_contract = _spectrum_output_contract(
        base_config, "accel_spectrum", start_date, end_date
    )
    cable_spectrum_contract = _spectrum_output_contract(
        base_config, "cable_accel_spectrum", start_date, end_date
    )

    for field in ("project_root", "data_root"):
        value = str(base_request.get(field) or "")
        _require_remote_path(value, rc_root, f"Base request {field}")
    points = _cable_points(base_config)
    if len(points) != expected_cable_points:
        raise ValueError(
            f"Expected {expected_cable_points} cable points, found {len(points)}: {points}"
        )

    destination = Path(output_dir).expanduser().resolve()
    if destination.exists() and any(destination.iterdir()):
        raise FileExistsError(f"Recovery bundle output must be empty: {destination}")
    (destination / "configs").mkdir(parents=True, exist_ok=True)
    (destination / "requests").mkdir(parents=True, exist_ok=True)

    jobs: list[dict[str, Any]] = []
    definitions = [
        (f"cable_accel_{index:02d}", "cable_accel", point)
        for index, point in enumerate(points, start=1)
    ] + [
        ("accel_spectrum", "accel_spectrum", None),
        ("cable_accel_spectrum", "cable_accel_spectrum", None),
    ]

    for job_name, module, point_id in definitions:
        config = copy.deepcopy(base_config)
        _sanitize_inactive_preprocessing_paths(config, rc_root)
        if point_id is not None:
            _restrict_cable_point(config, point_id)
        remote_config = ntpath.join(remote_bundle_root, "configs", f"{job_name}.json")
        config["source"] = remote_config
        config_raw = _json_bytes(config)
        config_sha = _sha256_bytes(config_raw)
        _atomic_write(destination / "configs" / f"{job_name}.json", config_raw)

        request = copy.deepcopy(base_request)
        request["config"] = config
        request["config_path"] = remote_config
        request["config_sha256"] = config_sha
        request["options"] = _module_options(
            base_request.get("options") if isinstance(base_request.get("options"), dict) else {},
            module,
        )
        request["enabled_modules"] = [module]
        request["async_run_id"] = f"high-memory-recovery-{job_name}"
        request["stop_file"] = ntpath.join(remote_bundle_root, "runtime", job_name, "stop.flag")
        request["async_status_file"] = ntpath.join(
            remote_bundle_root, "runtime", job_name, "analysis_status.json"
        )
        for json_path, value in _iter_absolute_windows_strings(request):
            if not _windows_within(value, rc_root):
                raise ValueError(
                    f"Generated request path {json_path} escapes RC root {rc_root}: {value}"
                )
        request_raw = _json_bytes(request)
        _atomic_write(destination / "requests" / f"{job_name}.json", request_raw)
        job: dict[str, Any] = {
            "job_name": job_name,
            "module": module,
            "point_id": point_id or "",
            "request_path": ntpath.join(remote_bundle_root, "requests", f"{job_name}.json"),
            "request_sha256": _sha256_bytes(request_raw),
            "config_path": remote_config,
            "config_sha256": config_sha,
            "runtime_dir": ntpath.join(remote_bundle_root, "runtime", job_name),
            "launch": False,
        }
        if point_id is not None:
            job["snapshot_dir"] = ntpath.join(
                remote_bundle_root, "cable_stats_snapshots", f"{len(jobs)+1:02d}_{point_id}"
            )
        jobs.append(job)

    plan = {
        "schema_version": 1,
        "plan_type": "high_memory_recovery_plan",
        "launch": False,
        "allowed_rc_root": rc_root,
        "remote_bundle_root": remote_bundle_root,
        "data_root": str(base_request["data_root"]),
        "start_date": start_date,
        "end_date": end_date,
        "cable_point_order": points,
        "cable_group_plot_policy": {
            "point_jobs": "disabled",
            "config_field": "plot_common.dynamic_group_sampling_mode",
            "value": "off",
            "resolution_required": (
                "Group figures must be recovered in a separate process or bound to "
                "independently audited existing artifacts before composite-manifest publication."
            ),
        },
        "dynamic_plot_contract": plot_contract,
        "cable_group_recovery": {
            "required": bool(effective_cable_groups),
            "launch": False,
            "resolved_groups": effective_cable_groups,
            "strategy": "separate_group_only_process",
            "automatic_request_generated": False,
            "blocking_reason": (
                "The current MATLAB module has no safe group-only entry point; a dedicated "
                "group-only process or independently audited existing group artifacts are required."
                if effective_cable_groups else ""
            ),
        },
        "jobs": jobs,
    }
    canonical_base_config_sha = _sha256_bytes(_json_bytes(base_config))
    recovery_expectations = {
        "schema_version": 1,
        "expectation_type": "high_memory_recovery_plot_buckets",
        "baseline_config_canonical_sha256": canonical_base_config_sha,
        "cable_points": [
            {"point_id": point, "expected_plot_provenance_count": 1}
            for point in points
        ],
        "cable_group": {
            "resolved_groups": effective_cable_groups,
            "expected_plot_provenance_count": 0 if not effective_cable_groups else None,
            "operator_action_required": bool(effective_cable_groups),
        },
        # Current spectrum writers do not emit formal plot provenance.  These
        # explicit zero buckets prevent their absence from being hidden by an
        # excess in another module; change only with code-backed evidence.
        "accel_spectrum": {
            "expected_plot_provenance_count": 0,
            **accel_spectrum_contract,
        },
        "cable_accel_spectrum": {
            "expected_plot_provenance_count": 0,
            **cable_spectrum_contract,
        },
    }
    _atomic_write(destination / "recovery_plan.json", _json_bytes(plan))
    _atomic_write(
        destination / "recovery_plot_expectations.json", _json_bytes(recovery_expectations)
    )
    group_evidence = {
        "schema_version": 1,
        "evidence_type": "cable_accel_group_plot_resolution",
        "status": "pending" if effective_cable_groups else "ok",
        "mode": "separate_recovery" if effective_cable_groups else "not_applicable",
        "reason_code": "" if effective_cable_groups else "no_effective_cable_accel_groups",
        "baseline_config_canonical_sha256": canonical_base_config_sha,
        "resolved_groups": effective_cable_groups,
        "artifacts": [],
    }
    _atomic_write(
        destination / "cable_group_evidence.json", _json_bytes(group_evidence)
    )
    return plan


def _nonempty_rows(ws: Any) -> list[int]:
    result: list[int] = []
    for row_index in range(1, ws.max_row + 1):
        if any(ws.cell(row_index, column).value is not None for column in range(1, ws.max_column + 1)):
            result.append(row_index)
    return result


def _single_row_workbook(path: Path, expected_point: str) -> tuple[Any, Any, list[Any], int]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("openpyxl is required to merge recovery statistics") from exc
    workbook = load_workbook(path, data_only=False)
    if len(workbook.worksheets) != 1:
        workbook.close()
        raise ValueError(f"Expected one worksheet in {path}")
    sheet = workbook.worksheets[0]
    rows = _nonempty_rows(sheet)
    if rows != [1, 2]:
        workbook.close()
        raise ValueError(f"Expected exactly one header and one data row in {path}; rows={rows}")
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    if not headers or str(headers[0]).strip().casefold().replace("_", "") not in {
        "pointid",
        "测点编号",
    }:
        workbook.close()
        raise ValueError(f"First statistics column must identify PointID in {path}: {headers}")
    actual_point = str(sheet.cell(2, 1).value or "").strip()
    if actual_point != expected_point:
        workbook.close()
        raise ValueError(
            f"Statistics point mismatch for {path}: expected {expected_point}, got {actual_point}"
        )
    return workbook, sheet, headers, 2


def merge_cable_accel_stats(
    ordered_points: Sequence[str],
    input_files: Sequence[Path | str],
    output_path: Path | str,
    *,
    allowed_root: Path | str,
    receipt_path: Path | str | None = None,
) -> dict[str, Any]:
    """Merge 15 one-row workbooks in exact configured point order."""

    root = Path(allowed_root).expanduser().resolve()
    points = [str(item).strip() for item in ordered_points]
    if len(points) != 15 or len(set(points)) != 15:
        raise ValueError("Cable statistics merge requires exactly 15 unique configured points")
    if len(input_files) != len(points):
        raise ValueError(f"Expected {len(points)} input workbooks, got {len(input_files)}")
    inputs = [_require_local_path(item, root, "Cable statistics input") for item in input_files]
    if len(set(inputs)) != len(inputs):
        raise ValueError("Cable statistics inputs must be unique files")
    target = _require_local_path(output_path, root, "Merged statistics output")
    receipt = _require_local_path(
        receipt_path or target.with_name(f"{target.stem}.merge_receipt.json"),
        root,
        "Merge receipt",
    )
    if target.exists() or receipt.exists():
        raise FileExistsError(f"Merged output/receipt already exists: {target}, {receipt}")

    source_records: list[dict[str, Any]] = []
    base_workbook = None
    base_sheet = None
    expected_headers: list[Any] | None = None
    try:
        for index, (point, source) in enumerate(zip(points, inputs)):
            if not source.is_file():
                raise FileNotFoundError(f"Cable statistics input does not exist: {source}")
            workbook, sheet, headers, data_row = _single_row_workbook(source, point)
            if expected_headers is None:
                expected_headers = headers
                base_workbook, base_sheet = workbook, sheet
            else:
                if headers != expected_headers:
                    workbook.close()
                    raise ValueError(
                        f"Column structure mismatch for {source}: {headers} != {expected_headers}"
                    )
                destination_row = 2 + index
                for column in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(data_row, column)
                    target_cell = base_sheet.cell(destination_row, column, source_cell.value)
                    target_cell._style = copy_style(source_cell._style)
                    if source_cell.has_style:
                        target_cell.number_format = source_cell.number_format
                    target_cell.font = copy_style(source_cell.font)
                    target_cell.fill = copy_style(source_cell.fill)
                    target_cell.border = copy_style(source_cell.border)
                    target_cell.alignment = copy_style(source_cell.alignment)
                    target_cell.protection = copy_style(source_cell.protection)
                workbook.close()
            source_records.append(
                {
                    "order": index + 1,
                    "point_id": point,
                    "path": str(source),
                    "bytes": source.stat().st_size,
                    "sha256": _sha256_file(source),
                }
            )
        if base_workbook is None or base_sheet is None or expected_headers is None:
            raise ValueError("No cable statistics inputs were supplied")
        target.parent.mkdir(parents=True, exist_ok=True)
        temporary = target.with_name(f".{target.name}.{os.getpid()}.tmp.xlsx")
        base_workbook.save(temporary)
        base_workbook.close()
        os.replace(temporary, target)
    finally:
        if base_workbook is not None:
            base_workbook.close()

    # Re-open the published workbook to make the order/row count a hard gate.
    try:
        from openpyxl import load_workbook
        checked = load_workbook(target, read_only=True, data_only=False)
        sheet = checked.worksheets[0]
        actual = [str(sheet.cell(row, 1).value or "").strip() for row in range(2, 17)]
        row_count = len(_nonempty_rows(sheet))
        checked.close()
    except Exception:
        target.unlink(missing_ok=True)
        raise
    if actual != points or row_count != 16:
        target.unlink(missing_ok=True)
        raise RuntimeError(f"Merged cable statistics failed order/row validation: {actual}")

    result = {
        "schema_version": 1,
        "receipt_type": "cable_accel_stats_merge",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "ordered_points": points,
        "columns": expected_headers,
        "inputs": source_records,
        "output": {
            "path": str(target),
            "bytes": target.stat().st_size,
            "sha256": _sha256_file(target),
            "rows": 15,
        },
    }
    _atomic_write(receipt, _json_bytes(result))
    result["receipt_path"] = str(receipt)
    result["receipt_sha256"] = _sha256_file(receipt)
    return result


def _module_records(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    raw = manifest.get("module_results") or manifest.get("module_logs") or []
    if isinstance(raw, dict):
        raw = [raw]
    if not isinstance(raw, list) or any(not isinstance(item, dict) for item in raw):
        raise ValueError("Analysis manifest module_results must be an object array")
    return [copy.deepcopy(item) for item in raw]


def _successful_record(manifest: dict[str, Any], module: str) -> dict[str, Any]:
    if str(manifest.get("status") or "").strip().casefold() not in SUCCESS_STATES:
        raise ValueError(f"Recovery manifest is not successful for {module}")
    records = _module_records(manifest)
    keys = [str(item.get("key") or "") for item in records]
    if len(keys) != len(set(keys)):
        raise ValueError(f"Recovery manifest contains duplicate module results: {keys}")
    foreign = [key for key in keys if key not in {module, *RECOVERY_AUXILIARY_MODULES}]
    if foreign:
        raise ValueError(
            f"Recovery manifest for {module} contains unexpected module results: {foreign}"
        )
    failed_auxiliary = [
        key for key, item in zip(keys, records)
        if key in RECOVERY_AUXILIARY_MODULES
        and str(item.get("status") or "").strip().casefold() not in SUCCESS_STATES
    ]
    if failed_auxiliary:
        raise ValueError(f"Recovery manifest has failed auxiliary results: {failed_auxiliary}")
    matches = [item for item in records if str(item.get("key") or "") == module]
    if len(matches) != 1:
        raise ValueError(f"Recovery manifest must contain exactly one {module} result")
    record = matches[0]
    if str(record.get("status") or "").strip().casefold() not in SUCCESS_STATES:
        raise ValueError(f"Recovery module {module} is not successful")
    return record


def _file_binding(path: Path, role: str) -> dict[str, Any]:
    return {"role": role, "path": str(path), "bytes": path.stat().st_size, "sha256": _sha256_file(path)}


def _validated_request_config(
    manifest: dict[str, Any], root: Path, *, expected_module: str,
    expected_point: str | None = None,
    expected_scope: tuple[Path, str, str] | None = None,
    expected_plot_contract: dict[str, dict[str, str]] | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    request = manifest.get("run_request")
    if not isinstance(request, dict):
        raise ValueError("Recovery manifest must contain run_request evidence")
    data_root = _require_local_path(
        str(request.get("data_root") or ""), root, "Recovery data root"
    )
    if not data_root.is_dir():
        raise FileNotFoundError(f"Recovery data root does not exist: {data_root}")
    if str(request.get("project_root") or "").strip():
        _require_local_path(str(request["project_root"]), root, "Recovery project root")
    start_date = str(request.get("start_date") or "").strip()
    end_date = str(request.get("end_date") or "").strip()
    if not start_date or not end_date:
        raise ValueError("Recovery run request must pin start_date and end_date")
    if expected_scope is not None and (data_root, start_date, end_date) != expected_scope:
        raise ValueError(
            "Recovery run request scope differs from baseline: "
            f"{(data_root, start_date, end_date)} != {expected_scope}"
        )
    enabled = request.get("enabled_modules")
    if enabled != [expected_module]:
        raise ValueError(
            f"Recovery run request must enable only {expected_module}: {enabled}"
        )
    options = request.get("options")
    if isinstance(options, dict):
        if str(options.get("input_mode") or "").strip().casefold() != "mat_only":
            raise ValueError("Recovery options must pin input_mode=mat_only")
        expected_option = MODULE_OPTION_FIELDS[expected_module]
        if options.get(expected_option) is not True:
            raise ValueError(f"Recovery options did not enable {expected_option}")
        unexpected = [
            field for module, field in MODULE_OPTION_FIELDS.items()
            if module != expected_module and options.get(field) is True
        ]
        if unexpected:
            raise ValueError(f"Recovery options enable unexpected modules: {unexpected}")
    config_path = _require_local_path(
        str(request.get("config_path") or ""), root, "Recovery config"
    )
    if not config_path.is_file():
        raise FileNotFoundError(f"Recovery config does not exist: {config_path}")
    expected_hash = str(request.get("config_sha256") or "").strip().upper()
    actual_hash = _sha256_file(config_path)
    if not expected_hash or expected_hash != actual_hash:
        raise ValueError(
            f"Recovery config SHA-256 mismatch: expected {expected_hash}, got {actual_hash}"
        )
    config = _json_object(config_path, "Recovery config")
    _require_mat_only_config(config, "Recovery config")
    _require_jiulongjiang_config(config, "Recovery config")
    _require_full_connect_contract(
        config, "Recovery config", expected=expected_plot_contract
    )
    if expected_point is not None:
        configured = _cable_points(config)
        if configured != [expected_point]:
            raise ValueError(
                f"Cable point config is not isolated to {expected_point}: {configured}"
            )
        mode = config.get("plot_common", {}).get("dynamic_group_sampling_mode") \
            if isinstance(config.get("plot_common"), dict) else None
        if str(mode or "").strip().casefold() != "off":
            raise ValueError(
                f"Cable point config must disable dynamic group plots for {expected_point}"
            )
    return request, config


def _plot_stub(path: Path) -> str:
    suffix = ".plot.json"
    name = path.name
    if not name.casefold().endswith(suffix):
        raise ValueError(f"Not a plot provenance path: {path}")
    return name[: -len(suffix)]


def _validate_plot_artifacts(
    module: str, artifacts: Sequence[dict[str, Any]], *, require_any: bool = True
) -> int:
    """Validate formal plot JSONs and their inventoried figure companions.

    Formal plot provenance must be complete enough for the workbench's strict
    source/input/finite/plotted closure gate.  A provenance file is not accepted
    merely because it parses: its ``file_stub`` and a same-directory, same-stub
    figure must also be present in the audited artifact list.
    """

    figures: set[tuple[Path, str]] = set()
    provenances: list[Path] = []
    for artifact in artifacts:
        path = Path(str(artifact.get("path") or "")).resolve()
        lower = path.name.casefold()
        if lower.endswith(".plot.json"):
            provenances.append(path)
        elif path.suffix.casefold() in {".jpg", ".jpeg", ".png", ".emf", ".fig"}:
            figures.add((path.parent, path.stem.casefold()))
    if require_any and not provenances:
        raise ValueError(f"Module {module} has no formal plot provenance")
    for path in provenances:
        payload = _json_object(path, f"{module} plot provenance")
        stub = _plot_stub(path)
        if str(payload.get("file_stub") or "").strip() != stub:
            raise ValueError(
                f"Plot provenance file_stub does not match its filename for {module}: {path}"
            )
        if (path.parent, stub.casefold()) not in figures:
            raise ValueError(
                f"Plot provenance has no inventoried same-stub figure for {module}: {path}"
            )
        row = inspect_plot_provenance(module, path)
        if not row.closed:
            raise ValueError(
                f"Plot provenance counts/source do not close for {module}: {path}: {row.message}"
            )
    return len(provenances)


def _plot_stubs(artifacts: Sequence[dict[str, Any]]) -> list[str]:
    return sorted(
        _plot_stub(Path(str(item.get("path") or "")))
        for item in artifacts
        if str(item.get("path") or "").casefold().endswith(".plot.json")
    )


def _validate_plot_bucket(
    label: str, artifacts: Sequence[dict[str, Any]], expectation: dict[str, Any],
    *, module: str,
) -> int:
    expected_count = expectation.get("expected_plot_provenance_count")
    if not isinstance(expected_count, int) or expected_count < 0:
        raise ValueError(f"{label} must declare a non-negative expected plot count")
    actual_count = _validate_plot_artifacts(
        module, artifacts, require_any=expected_count > 0
    )
    if actual_count != expected_count:
        raise ValueError(f"{label} plot count differs: {actual_count} != {expected_count}")
    if "expected_plot_stubs" in expectation:
        raw_stubs = expectation.get("expected_plot_stubs")
        if not isinstance(raw_stubs, list) or any(
            not isinstance(value, str) or not value.strip() for value in raw_stubs
        ):
            raise ValueError(f"{label} expected_plot_stubs must be a string array")
        expected_stubs = sorted(value.strip() for value in raw_stubs)
        if len(expected_stubs) != len(set(expected_stubs)):
            raise ValueError(f"{label} expected_plot_stubs contains duplicates")
        actual_stubs = _plot_stubs(artifacts)
        if actual_stubs != expected_stubs:
            raise ValueError(f"{label} plot stubs differ: {actual_stubs} != {expected_stubs}")
    return actual_count


def _validate_spectrum_stats(
    stats_path: Path, module: str, contract: dict[str, Any],
) -> None:
    """Require one structurally valid statistics sheet per configured point."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("openpyxl is required to validate spectrum statistics") from exc
    expected = contract.get("expected_stat_sheets")
    if not isinstance(expected, list) or not expected or any(
        not isinstance(value, str) or not value.strip() for value in expected
    ):
        raise ValueError(f"{module} expected_stat_sheets must be a non-empty string array")
    workbook = load_workbook(stats_path, read_only=True, data_only=False)
    try:
        if workbook.sheetnames != expected:
            raise ValueError(
                f"{module} statistics sheets differ from configured points: "
                f"{workbook.sheetnames} != {expected}"
            )
        for point in expected:
            sheet = workbook[point]
            if sheet.max_row < 2:
                raise ValueError(f"{module} statistics sheet {point} has no daily rows")
            headers = [
                str(sheet.cell(1, column).value or "").strip()
                for column in range(1, sheet.max_column + 1)
            ]
            if "Date" not in headers:
                raise ValueError(f"{module} statistics sheet {point} has no Date column")
            if not any(value.startswith("Freq_") for value in headers):
                raise ValueError(f"{module} statistics sheet {point} has no Freq_* column")
            if not any(value.startswith("Amp_") for value in headers):
                raise ValueError(f"{module} statistics sheet {point} has no Amp_* column")
            if module == "cable_accel_spectrum" and "CableForce_kN" not in headers:
                raise ValueError(
                    f"{module} statistics sheet {point} has no CableForce_kN column"
                )
    finally:
        workbook.close()


def _validate_spectrum_figures(
    artifacts: Sequence[dict[str, Any]], module: str, contract: dict[str, Any],
) -> None:
    """Require the exact configured set of report-facing spectrum figure bundles."""

    from collections import Counter

    expected = contract.get("expected_formal_figure_stubs")
    if not isinstance(expected, list) or not expected or any(
        not isinstance(value, str) or not value.strip() for value in expected
    ):
        raise ValueError(
            f"{module} expected_formal_figure_stubs must be a non-empty string array"
        )
    raster_extensions = {".jpg", ".jpeg", ".png"}
    all_figure_extensions = {*raster_extensions, ".fig", ".emf"}
    bundles: dict[tuple[Path, str], set[str]] = {}
    for artifact in artifacts:
        path = Path(str(artifact.get("path") or "")).resolve()
        suffix = path.suffix.casefold()
        if suffix not in all_figure_extensions:
            continue
        if not (
            path.stem.startswith("SpecFreq_") or path.stem.startswith("CableForce_")
        ):
            continue
        bundles.setdefault((path.parent, path.stem), set()).add(suffix)
    if not bundles:
        raise ValueError(f"{module} has no formal SpecFreq_/CableForce_ figures")
    for (_parent, stub), extensions in bundles.items():
        if not extensions.intersection(raster_extensions):
            raise ValueError(f"{module} formal figure {stub} has no raster image")
    actual = Counter(stub for _parent, stub in bundles)
    required = Counter(expected)
    if actual != required:
        raise ValueError(
            f"{module} formal figure bundles differ from the config-derived contract: "
            f"{dict(actual)} != {dict(required)}"
        )


def build_baseline_evidence(
    run_request_path: Path | str,
    inventory_path: Path | str,
    output_path: Path | str,
    *,
    expected_modules: Sequence[str],
    allowed_root: Path | str,
) -> dict[str, Any]:
    """Rebuild the first-12 evidence from live files, never a failed manifest.

    The inventory is an explicit audit boundary.  Each module entry supplies
    ``stats_path``, an exact ``expected_plot_provenance_count``, and one or
    more ``artifact_paths`` and/or ``artifact_globs``.
    Relative glob patterns are evaluated below ``allowed_root``.  This avoids
    heuristically accepting stale files from neighboring runs.
    """

    root = Path(allowed_root).expanduser().resolve()
    request_path = _require_local_path(run_request_path, root, "Baseline run request")
    inventory_file = _require_local_path(inventory_path, root, "Baseline inventory")
    output = _require_local_path(output_path, root, "Baseline evidence output")
    if output.exists():
        raise FileExistsError(f"Baseline evidence output already exists: {output}")
    request = _json_object(request_path, "Baseline run request")
    _require_local_path(str(request.get("data_root") or ""), root, "Baseline data root")
    if str(request.get("project_root") or "").strip():
        _require_local_path(str(request["project_root"]), root, "Baseline project root")
    config_path = _require_local_path(str(request.get("config_path") or ""), root, "Baseline config")
    if not config_path.is_file():
        raise FileNotFoundError(f"Baseline config does not exist: {config_path}")
    config_hash = _sha256_file(config_path)
    if str(request.get("config_sha256") or "").strip().upper() != config_hash:
        raise ValueError("Baseline run request config SHA-256 does not match the pinned config")
    config = _json_object(config_path, "Baseline config")
    _require_mat_only_config(config, "Baseline config")
    _require_jiulongjiang_config(config, "Baseline config")
    plot_contract = _require_full_connect_contract(config, "Baseline config")
    options = request.get("options") if isinstance(request.get("options"), dict) else {}
    modules = [str(item) for item in expected_modules]
    if len(modules) != 12 or len(set(modules)) != 12:
        raise ValueError("Baseline evidence requires exactly 12 unique expected modules")
    for module in modules:
        option = MODULE_OPTION_FIELDS.get(module)
        if not option or options.get(option) is not True:
            raise ValueError(f"Baseline run request did not enable expected module {module}")

    inventory = _json_object(inventory_file, "Baseline artifact inventory")
    raw_entries = inventory.get("modules")
    if not isinstance(raw_entries, list) or any(not isinstance(item, dict) for item in raw_entries):
        raise ValueError("Baseline inventory modules must be an object array")
    by_key = {str(item.get("key") or ""): item for item in raw_entries}
    if list(by_key) != modules or len(by_key) != len(raw_entries):
        raise ValueError(
            f"Baseline inventory must contain the exact module order {modules}: {list(by_key)}"
        )

    records: list[dict[str, Any]] = []
    claimed: set[Path] = set()
    provenance_count = 0
    for module in modules:
        entry = by_key[module]
        stats_path = _require_local_path(str(entry.get("stats_path") or ""), root, f"{module} stats")
        if not stats_path.is_file():
            raise FileNotFoundError(f"Baseline stats do not exist for {module}: {stats_path}")
        if stats_path in claimed:
            raise ValueError(f"Baseline statistics file is shared by multiple modules: {stats_path}")
        claimed.add(stats_path)
        candidates: list[Path] = []
        for value in entry.get("artifact_paths") or []:
            candidates.append(_require_local_path(str(value), root, f"{module} artifact"))
        for pattern in entry.get("artifact_globs") or []:
            pattern_text = str(pattern).strip()
            if not pattern_text or PureWindowsPath(pattern_text).is_absolute() or Path(pattern_text).is_absolute():
                raise ValueError(f"Baseline artifact glob must be relative to RC root: {pattern_text}")
            candidates.extend(path.resolve() for path in root.glob(pattern_text) if path.is_file())
        unique = sorted(set(candidates), key=lambda item: item.as_posix().casefold())
        if not unique:
            raise ValueError(f"Baseline inventory found no artifacts for {module}")
        artifacts = [
            {"kind": "stats", "role": "stats", "path": str(stats_path), "exists": True,
             "bytes": stats_path.stat().st_size, "sha256": _sha256_file(stats_path)}
        ]
        for path in unique:
            _require_local_path(path, root, f"{module} artifact")
            if not path.is_file():
                raise FileNotFoundError(f"Baseline artifact does not exist: {path}")
            if path in claimed or path == stats_path:
                raise ValueError(f"Baseline artifact is duplicated across modules: {path}")
            claimed.add(path)
            lower = path.name.casefold()
            if lower.endswith(".plot.json"):
                kind, role = "plot_provenance", "plot_provenance"
                _json_object(path, "Plot provenance")
            elif path.suffix.casefold() in {".jpg", ".jpeg", ".png", ".emf", ".fig"}:
                kind, role = "figure", _infer_artifact_role(path)
            elif lower.endswith("_summary.txt"):
                kind, role = "summary", _infer_artifact_role(path)
            else:
                raise ValueError(f"Unsupported baseline artifact type for {module}: {path}")
            artifacts.append({"kind": kind, "role": role, "path": str(path), "exists": True,
                              "bytes": path.stat().st_size, "sha256": _sha256_file(path)})
        expected_provenance_count = entry.get("expected_plot_provenance_count")
        if not isinstance(expected_provenance_count, int) or expected_provenance_count < 0:
            raise ValueError(
                f"Baseline module {module} must declare expected_plot_provenance_count"
            )
        bucket_expectation = {
            "expected_plot_provenance_count": expected_provenance_count,
        }
        if "expected_plot_stubs" in entry:
            bucket_expectation["expected_plot_stubs"] = entry["expected_plot_stubs"]
        actual_provenance_count = _validate_plot_bucket(
            f"Baseline module {module}", artifacts, bucket_expectation, module=module
        )
        provenance_count += actual_provenance_count
        records.append({
            "key": module,
            "label": module,
            "status": "ok",
            "message": "Rebuilt from pinned run request/config and explicit live-file inventory.",
            "stats_path": str(stats_path),
            "stats_exists": True,
            "artifacts": artifacts,
            "artifact_count": len(artifacts),
            "figure_paths": [item["path"] for item in artifacts if item["kind"] == "figure"],
            "figure_count": sum(item["kind"] == "figure" for item in artifacts),
            "plot_provenance_count": actual_provenance_count,
        })
    request_bridge_profile = (
        request.get("bridge_profile")
        if isinstance(request.get("bridge_profile"), dict)
        else {}
    )
    bridge_id = str(
        request_bridge_profile.get("bridge_id") or config.get("vendor") or ""
    ).strip()
    evidence = {
        "schema_version": 2,
        "manifest_type": "baseline_artifact_reconstruction",
        "status": "ok",
        "written_at": datetime.now(timezone.utc).isoformat(),
        "run_request": request,
        "config_path": str(config_path),
        "config_sha256": config_hash,
        "config_vendor": str(config.get("vendor") or ""),
        "dynamic_plot_contract": plot_contract,
        "module_results": records,
        "module_logs": records,
        "module_artifacts": [
            {"key": item["key"], "label": item["label"], "artifacts": item["artifacts"]}
            for item in records
        ],
        "artifact_count": sum(len(item["artifacts"]) for item in records),
        "plot_provenance_count": provenance_count,
        "module_status_counts": {"ok": 12, "fail": 0, "skip": 0, "missing": 0, "other": 0},
        "source_chain": {
            "run_request": _file_binding(request_path, "run_request"),
            "config": _file_binding(config_path, "config"),
            "inventory": _file_binding(inventory_file, "explicit_artifact_inventory"),
            "failed_manifest_used": False,
        },
    }
    if bridge_id:
        # A manually prepared recovery request can predate the workbench's
        # bridge_profile field.  Keep the reconstructed manifest bound to the
        # effective vendor/bridge so the normal report gate need not weaken
        # its bridge-identity check for recovered production results.
        evidence["bridge_profile"] = {
            **request_bridge_profile,
            "bridge_id": bridge_id,
        }
    _atomic_write(output, _json_bytes(evidence, pretty=False))
    _json_object(output, "Published baseline evidence")
    return {"path": str(output), "bytes": output.stat().st_size,
            "sha256": _sha256_file(output), "module_count": 12,
            "artifact_count": evidence["artifact_count"]}


def _infer_artifact_role(path: Path) -> str:
    """Mirror ``ArtifactCollector.inferRole`` for rebuilt baseline evidence.

    Recovery manifests are report-authoritative, so generic ``time_history``
    and ``summary`` labels are insufficient for semantic plot directories.
    Only the immediate parent plus filename participate, matching the MATLAB
    collector and avoiding accidental matches from arbitrary ancestors.
    """

    text = f"{path.parent.name}/{path.name}".casefold()
    if "rms10" in text or "rms_10" in text:
        return "rms10min"
    if "envelope30" in text or "包络" in text:
        return "envelope30min"
    if any(token in text for token in ("specfreq", "spectrum", "psd", "频谱")):
        return "spectrum"
    if "boxplot" in text or "箱线" in text:
        return "boxplot"
    if "windrose" in text or "风玫瑰" in text:
        return "wind_rose"
    if "freq" in text or "频率" in text or "频次" in text:
        return "frequency_distribution"
    if "speed10min" in text or "10min" in text:
        return "wind_speed10min"
    if "filt" in text or "滤波" in text:
        return "filtered"
    if "orig" in text or "原始" in text:
        return "raw"
    return "time_history"


def _rehash_module_record(
    record: dict[str, Any], root: Path, *, stats_override: Path | None = None,
    ignore_stats: bool = False,
) -> dict[str, Any]:
    result = copy.deepcopy(record)
    if str(result.get("status") or "").strip().casefold() not in SUCCESS_STATES:
        raise ValueError(f"Module is not successful: {result.get('key')}")
    artifacts = result.get("artifacts") or []
    if isinstance(artifacts, dict):
        artifacts = [artifacts]
    if not isinstance(artifacts, list):
        raise ValueError(f"Module artifacts must be an array: {result.get('key')}")
    normalized: list[dict[str, Any]] = []
    seen: set[Path] = set()
    for index, artifact in enumerate(artifacts):
        if not isinstance(artifact, dict):
            raise ValueError(f"Module artifact {index} is not an object: {result.get('key')}")
        if (stats_override is not None or ignore_stats) and str(artifact.get("kind") or "").casefold() == "stats":
            continue
        path = _require_local_path(str(artifact.get("path") or ""), root, "Module artifact")
        if path in seen:
            raise ValueError(f"Duplicate artifact in module {result.get('key')}: {path}")
        seen.add(path)
        if not path.is_file():
            raise FileNotFoundError(f"Module artifact does not exist: {path}")
        updated = copy.deepcopy(artifact)
        updated.update({"path": str(path), "exists": True, "bytes": path.stat().st_size, "sha256": _sha256_file(path)})
        normalized.append(updated)
    stats_path = stats_override
    if not ignore_stats and stats_path is None and str(result.get("stats_path") or "").strip():
        stats_path = _require_local_path(str(result["stats_path"]), root, "Module statistics")
    if stats_path is not None:
        stats_path = _require_local_path(stats_path, root, "Module statistics")
        if not stats_path.is_file():
            raise FileNotFoundError(f"Module statistics do not exist: {stats_path}")
        if stats_path not in seen:
            normalized.insert(0, {"kind": "stats", "role": "stats", "path": str(stats_path), "exists": True, "bytes": stats_path.stat().st_size, "sha256": _sha256_file(stats_path)})
        result["stats_path"] = str(stats_path)
        result["stats_exists"] = True
    elif ignore_stats:
        result["stats_path"] = ""
        result["stats_exists"] = False
    result["artifacts"] = normalized
    result["artifact_count"] = len(normalized)
    result["figure_paths"] = [item["path"] for item in normalized if str(item.get("kind") or "").casefold() == "figure"]
    result["figure_count"] = len(result["figure_paths"])
    return result


def _point_token(value: str) -> str:
    return re.sub(r"[^0-9a-z]+", "", str(value).casefold())


def _artifact_belongs_to_point(artifact: dict[str, Any], path: Path, point_id: str) -> bool:
    if _point_token(point_id) not in _point_token(path.name):
        return False
    if path.name.casefold().endswith(".plot.json"):
        payload = _json_object(path, "Cable point plot provenance")
        series = payload.get("series")
        if isinstance(series, dict):
            series = [series]
        if not isinstance(series, list) or not series:
            raise ValueError(f"Cable plot provenance has no series: {path}")
        identifiers = {
            str(item.get("point_id") or "").strip()
            for item in series
            if isinstance(item, dict) and str(item.get("point_id") or "").strip()
        }
        if identifiers != {point_id}:
            raise ValueError(
                f"Cable plot provenance point IDs do not match {point_id}: {identifiers}: {path}"
            )
    return True


def _validate_reporting_contract(
    contract_path: Path | str,
    root: Path,
    records: Sequence[dict[str, Any]],
    baseline_config: dict[str, Any],
) -> tuple[Path, dict[str, Any]]:
    """Validate a MATLAB-produced report contract before manifest recovery.

    Recovery manifests are report-authoritative, so silently falling back to
    a separately discovered ``analysis_reporting_contract_*.json`` would
    break the pinned provenance chain.  The contract is therefore required,
    checked against the recovered module/statistics inventory, and embedded
    in the composite manifest while its source file is hash-bound below.
    """

    path = _require_local_path(contract_path, root, "Analysis reporting contract")
    if not path.is_file():
        raise FileNotFoundError(f"Analysis reporting contract does not exist: {path}")
    contract = _json_object(path, "Analysis reporting contract")
    if contract.get("schema_version") != 1:
        raise ValueError("Analysis reporting contract schema_version must be 1")
    if contract.get("contract_type") != "analysis_reporting_contract":
        raise ValueError(
            "Analysis reporting contract has an unexpected contract_type"
        )

    expected_bridge_id = str(baseline_config.get("vendor") or "").strip()
    profile = contract.get("profile")
    if not isinstance(profile, dict):
        raise ValueError("Analysis reporting contract profile must be an object")
    contract_bridge_id = str(profile.get("bridge_id") or "").strip()
    if not expected_bridge_id or contract_bridge_id.casefold() != expected_bridge_id.casefold():
        raise ValueError(
            "Analysis reporting contract bridge profile differs from the pinned config: "
            f"{contract_bridge_id or '<empty>'} != {expected_bridge_id or '<empty>'}"
        )
    contract_vendor = str(profile.get("vendor") or "").strip()
    if contract_vendor and contract_vendor.casefold() != expected_bridge_id.casefold():
        raise ValueError(
            "Analysis reporting contract vendor differs from the pinned config: "
            f"{contract_vendor} != {expected_bridge_id}"
        )

    modules = contract.get("modules")
    if not isinstance(modules, list) or any(not isinstance(item, dict) for item in modules):
        raise ValueError("Analysis reporting contract modules must be an object array")
    keys = [str(item.get("key") or "").strip() for item in modules]
    if any(not key for key in keys) or len(keys) != len(set(keys)):
        raise ValueError("Analysis reporting contract module keys must be non-empty and unique")
    expected_by_key = {str(item.get("key") or ""): item for item in records}
    if set(keys) != set(expected_by_key) or len(keys) != len(expected_by_key):
        raise ValueError(
            "Analysis reporting contract module set differs from the recovered manifest: "
            f"{keys} != {list(expected_by_key)}"
        )

    summary = contract.get("summary")
    if not isinstance(summary, dict):
        raise ValueError("Analysis reporting contract summary must be an object")
    module_count = summary.get("module_count")
    if isinstance(module_count, bool) or not isinstance(module_count, int) \
            or module_count != len(modules):
        raise ValueError(
            "Analysis reporting contract summary.module_count does not match its modules"
        )

    total_points = 0
    total_groups = 0
    for module in modules:
        key = str(module["key"])
        expected_stats = Path(str(expected_by_key[key].get("stats_path") or "")).name
        contract_stats = str(module.get("stats_file") or "").strip()
        if not expected_stats or contract_stats.casefold() != expected_stats.casefold():
            raise ValueError(
                f"Analysis reporting contract stats_file differs for {key}: "
                f"{contract_stats or '<empty>'} != {expected_stats or '<empty>'}"
            )

        points = module.get("points")
        point_count = module.get("point_count")
        if not isinstance(points, list) or isinstance(point_count, bool) \
                or not isinstance(point_count, int) or point_count != len(points):
            raise ValueError(
                f"Analysis reporting contract point count is inconsistent for {key}"
            )
        if any(not isinstance(point, str) or not point.strip() for point in points):
            raise ValueError(
                f"Analysis reporting contract points must be non-empty strings for {key}"
            )
        if len(points) != len(set(points)):
            raise ValueError(f"Analysis reporting contract points are duplicated for {key}")
        total_points += point_count

        groups = module.get("groups")
        group_count = module.get("group_count")
        if not isinstance(groups, list) or isinstance(group_count, bool) \
                or not isinstance(group_count, int) or group_count != len(groups):
            raise ValueError(
                f"Analysis reporting contract group count is inconsistent for {key}"
            )
        total_groups += group_count

        output_dirs = module.get("output_dirs")
        output_records = module.get("output_dir_records")
        if not isinstance(output_dirs, list) or any(
            not isinstance(value, str) or not value.strip() for value in output_dirs
        ):
            raise ValueError(
                f"Analysis reporting contract output_dirs must be a string array for {key}"
            )
        # MATLAB jsonencode emits a scalar struct as a JSON object but emits a
        # struct array as a JSON array.  Treat the scalar representation as the
        # one-record form of the same contract and normalize it before the
        # remaining validation and before embedding the contract in the
        # composite manifest.  An empty struct represents no records.
        if isinstance(output_records, dict):
            output_records = [output_records] if output_records else []
            module["output_dir_records"] = output_records
        if not isinstance(output_records, list) or any(
            not isinstance(item, dict) for item in output_records
        ):
            raise ValueError(
                f"Analysis reporting contract output_dir_records must be an object array for {key}"
            )
        record_dirs: list[str] = []
        seen_records: set[tuple[str, str]] = set()
        for record in output_records:
            field = str(record.get("field") or "").strip()
            directory = str(record.get("dir") or "").strip()
            role = str(record.get("role") or "").strip()
            if not field or not directory or not role:
                raise ValueError(
                    f"Analysis reporting contract output-dir record is incomplete for {key}"
                )
            win_path = PureWindowsPath(directory)
            if win_path.is_absolute() or Path(directory).is_absolute() \
                    or ".." in win_path.parts:
                raise ValueError(
                    f"Analysis reporting contract output dir must be relative for {key}: {directory}"
                )
            identity = (field.casefold(), directory.casefold())
            if identity in seen_records:
                raise ValueError(
                    f"Analysis reporting contract output-dir record is duplicated for {key}"
                )
            seen_records.add(identity)
            record_dirs.append(directory)
        expected_dirs = list(dict.fromkeys(record_dirs))
        if output_dirs != expected_dirs:
            raise ValueError(
                f"Analysis reporting contract output_dirs differ from output_dir_records for {key}"
            )

    for field, actual in (
        ("point_count", total_points),
        ("group_count", total_groups),
    ):
        value = summary.get(field)
        if isinstance(value, bool) or not isinstance(value, int) or value != actual:
            raise ValueError(
                f"Analysis reporting contract summary.{field} is inconsistent"
            )
    return path, contract


def compose_recovery_manifest(
    baseline_manifest_path: Path | str,
    *,
    expected_baseline_modules: Sequence[str],
    cable_point_manifest_paths: Sequence[Path | str],
    cable_merge_receipt_path: Path | str,
    cable_group_evidence_path: Path | str,
    recovery_expectations_path: Path | str,
    reporting_contract_path: Path | str,
    accel_spectrum_manifest_path: Path | str,
    cable_spectrum_manifest_path: Path | str,
    output_path: Path | str,
    allowed_root: Path | str,
    expected_plot_provenance_count: int,
) -> dict[str, Any]:
    """Create one valid manifest from 12 baseline and recovered module evidence."""

    root = Path(allowed_root).expanduser().resolve()
    baseline_path = _require_local_path(baseline_manifest_path, root, "Baseline manifest")
    output = _require_local_path(output_path, root, "Composite manifest")
    if output.exists():
        raise FileExistsError(f"Composite manifest already exists: {output}")
    baseline = _json_object(baseline_path, "Baseline evidence manifest")
    if baseline.get("manifest_type") != "baseline_artifact_reconstruction":
        raise ValueError(
            "Baseline evidence must be rebuilt from run_request/config and explicit artifact inventory; "
            "the failed/truncated analysis manifest is not accepted"
        )
    if str(baseline.get("status") or "").strip().casefold() not in SUCCESS_STATES:
        raise ValueError("Baseline evidence manifest must be successful")
    baseline_records = _module_records(baseline)
    expected = [str(item) for item in expected_baseline_modules]
    actual = [str(item.get("key") or "") for item in baseline_records]
    if len(expected) != 12 or len(set(expected)) != 12 or actual != expected:
        raise ValueError(f"Baseline module order must exactly match 12 expected modules: {actual} != {expected}")
    records = [_rehash_module_record(item, root) for item in baseline_records]
    rebuilt_baseline_provenance_count = sum(
        _validate_plot_artifacts(
            str(record.get("key") or ""), record.get("artifacts") or [], require_any=False
        )
        for record in records
    )
    published_baseline_provenance_count = baseline.get("plot_provenance_count")
    if (
        not isinstance(published_baseline_provenance_count, int)
        or rebuilt_baseline_provenance_count != published_baseline_provenance_count
    ):
        raise ValueError(
            "Baseline formal plot provenance inventory changed after evidence publication"
        )
    baseline_request = baseline.get("run_request")
    if not isinstance(baseline_request, dict):
        raise ValueError("Baseline evidence is missing its pinned run_request")
    baseline_scope = (
        _require_local_path(
            str(baseline_request.get("data_root") or ""), root, "Baseline data root"
        ),
        str(baseline_request.get("start_date") or "").strip(),
        str(baseline_request.get("end_date") or "").strip(),
    )
    if not baseline_scope[1] or not baseline_scope[2]:
        raise ValueError("Baseline run request must pin start_date and end_date")
    baseline_config_path = _require_local_path(
        str(baseline.get("config_path") or ""), root, "Baseline config"
    )
    if _sha256_file(baseline_config_path) != str(baseline.get("config_sha256") or "").upper():
        raise ValueError("Baseline config changed after evidence publication")
    baseline_config = _json_object(baseline_config_path, "Baseline config")
    baseline_plot_contract = _require_full_connect_contract(
        baseline_config, "Baseline config"
    )
    effective_cable_groups = _effective_cable_groups(baseline_config)
    expectations_path = _require_local_path(
        recovery_expectations_path, root, "Recovery plot expectations"
    )
    expectations = _json_object(expectations_path, "Recovery plot expectations")
    if expectations.get("expectation_type") != "high_memory_recovery_plot_buckets":
        raise ValueError("Recovery plot expectations have an unexpected expectation_type")
    canonical_config_sha = _sha256_bytes(_json_bytes(baseline_config))
    if str(expectations.get("baseline_config_canonical_sha256") or "").upper() \
            != canonical_config_sha:
        raise ValueError("Recovery plot expectations do not bind the baseline config")
    derived_spectrum_contracts = {
        module: _spectrum_output_contract(
            baseline_config, module, baseline_scope[1], baseline_scope[2]
        )
        for module in ("accel_spectrum", "cable_accel_spectrum")
    }
    for module, contract in derived_spectrum_contracts.items():
        expected_payload = {
            "expected_plot_provenance_count": 0,
            **contract,
        }
        if expectations.get(module) != expected_payload:
            raise ValueError(
                f"Recovery expectations for {module} differ from the pinned config/scope"
            )

    receipt_path = _require_local_path(cable_merge_receipt_path, root, "Cable merge receipt")
    receipt = _json_object(receipt_path, "Cable merge receipt")
    if receipt.get("receipt_type") != "cable_accel_stats_merge":
        raise ValueError("Cable merge receipt has an unexpected receipt_type")
    points = receipt.get("ordered_points")
    if not isinstance(points, list) or len(points) != 15 or len(set(points)) != 15:
        raise ValueError("Cable merge receipt must list 15 unique ordered points")
    point_expectations = expectations.get("cable_points")
    if not isinstance(point_expectations, list) or any(
        not isinstance(item, dict) for item in point_expectations
    ):
        raise ValueError("Recovery plot expectations cable_points must be an object array")
    expectation_point_ids = [str(item.get("point_id") or "") for item in point_expectations]
    if expectation_point_ids != points:
        raise ValueError(
            f"Recovery point expectation order differs from merge receipt: {expectation_point_ids}"
        )
    merged = receipt.get("output")
    if not isinstance(merged, dict):
        raise ValueError("Cable merge receipt is missing output binding")
    merged_path = _require_local_path(str(merged.get("path") or ""), root, "Merged cable statistics")
    if not merged_path.is_file() or int(merged.get("bytes") or -1) != merged_path.stat().st_size or str(merged.get("sha256") or "").upper() != _sha256_file(merged_path):
        raise ValueError("Merged cable statistics no longer match their receipt")

    cable_paths = [_require_local_path(item, root, "Cable point recovery manifest") for item in cable_point_manifest_paths]
    if len(cable_paths) != 15 or len(set(cable_paths)) != 15:
        raise ValueError("Exactly 15 unique cable point recovery manifests are required")
    cable_artifacts: list[dict[str, Any]] = []
    cable_source_chain: list[dict[str, Any]] = []
    artifact_paths: set[Path] = set()
    warnings: list[Any] = []
    for point, point_expectation, manifest_path in zip(points, point_expectations, cable_paths):
        manifest = _json_object(manifest_path, f"Cable point manifest {point}")
        record = _successful_record(manifest, "cable_accel")
        _request, _config = _validated_request_config(
            manifest, root, expected_module="cable_accel", expected_point=point,
            expected_scope=baseline_scope, expected_plot_contract=baseline_plot_contract,
        )
        checked = _rehash_module_record(record, root, ignore_stats=True)
        ignored = 0
        matched = 0
        for artifact in checked["artifacts"]:
            if str(artifact.get("kind") or "").casefold() == "stats":
                continue
            artifact_path = Path(artifact["path"]).resolve()
            if not _artifact_belongs_to_point(artifact, artifact_path, point):
                ignored += 1
                continue
            if artifact_path in artifact_paths:
                raise ValueError(f"Cable point manifests contain duplicate artifact: {artifact_path}")
            artifact_paths.add(artifact_path)
            cable_artifacts.append(artifact)
            matched += 1
        if matched == 0:
            raise ValueError(f"Cable point manifest has no point-specific artifacts for {point}")
        point_artifacts = [
            item for item in cable_artifacts
            if Path(str(item.get("path") or "")).resolve() in artifact_paths
            and _point_token(point) in _point_token(Path(str(item.get("path") or "")).name)
        ]
        _validate_plot_bucket(
            f"Cable point {point}", point_artifacts, point_expectation,
            module="cable_accel",
        )
        warnings.extend(checked.get("warnings") or [])
        cable_source_chain.append({
            "point_id": point,
            "ignored_neighbor_artifact_count": ignored,
            **_file_binding(manifest_path, "cable_point_manifest"),
        })

    group_evidence_path = _require_local_path(
        cable_group_evidence_path, root, "Cable group-plot resolution evidence"
    )
    group_evidence = _json_object(group_evidence_path, "Cable group-plot resolution evidence")
    if group_evidence.get("evidence_type") != "cable_accel_group_plot_resolution":
        raise ValueError("Cable group evidence has an unexpected evidence_type")
    if str(group_evidence.get("status") or "").casefold() not in SUCCESS_STATES:
        raise ValueError("Cable group-plot resolution evidence is not successful")
    group_mode = str(group_evidence.get("mode") or "").strip()
    if group_mode not in {"separate_recovery", "audited_existing", "not_applicable"}:
        raise ValueError("Cable group evidence mode is not recognized")
    group_entries = group_evidence.get("artifacts") or []
    if not isinstance(group_entries, list) or any(not isinstance(item, dict) for item in group_entries):
        raise ValueError("Cable group evidence artifacts must be an object array")
    group_expectation = expectations.get("cable_group")
    if not isinstance(group_expectation, dict):
        raise ValueError("Recovery plot expectations must contain cable_group")
    if group_expectation.get("resolved_groups") != effective_cable_groups:
        raise ValueError("Cable group expectation does not match baseline fallback-group resolution")
    if group_evidence.get("resolved_groups") != effective_cable_groups:
        raise ValueError("Cable group evidence does not match baseline fallback-group resolution")
    if str(group_evidence.get("baseline_config_canonical_sha256") or "").upper() \
            != canonical_config_sha:
        raise ValueError("Cable group evidence does not bind the baseline config")
    if group_mode == "not_applicable":
        if effective_cable_groups:
            raise ValueError("Cable groups are configured; not_applicable evidence is forbidden")
        if group_evidence.get("reason_code") != "no_effective_cable_accel_groups":
            raise ValueError(
                "not_applicable cable group evidence requires reason_code="
                "no_effective_cable_accel_groups"
            )
    if group_mode != "not_applicable" and not group_entries:
        raise ValueError("Recovered/audited cable group evidence requires artifacts")
    for entry in group_entries:
        group_path = _require_local_path(str(entry.get("path") or ""), root, "Cable group artifact")
        if not group_path.is_file():
            raise FileNotFoundError(f"Cable group artifact does not exist: {group_path}")
        if group_path in artifact_paths:
            raise ValueError(f"Duplicate cable group artifact: {group_path}")
        artifact_paths.add(group_path)
        cable_artifacts.append({
            "kind": str(entry.get("kind") or ("plot_provenance" if group_path.name.casefold().endswith(".plot.json") else "figure")),
            "role": str(entry.get("role") or "group_plot"),
            "path": str(group_path),
            "exists": True,
            "bytes": group_path.stat().st_size,
            "sha256": _sha256_file(group_path),
        })
    group_artifacts = cable_artifacts[-len(group_entries):] if group_entries else []
    _validate_plot_bucket(
        "Cable group", group_artifacts, group_expectation, module="cable_accel"
    )
    cable_artifacts.insert(0, {"kind": "stats", "role": "stats", "path": str(merged_path), "exists": True, "bytes": merged_path.stat().st_size, "sha256": _sha256_file(merged_path)})
    cable_record = {
        "key": "cable_accel",
        "label": "cable_accel",
        "status": "ok",
        "message": "Recovered as 15 point-isolated processes and merged in configured order.",
        "error_type": "",
        "stats_file": merged_path.name,
        "stats_path": str(merged_path),
        "stats_exists": True,
        "artifacts": cable_artifacts,
        "artifact_count": len(cable_artifacts),
        "figure_paths": [item["path"] for item in cable_artifacts if str(item.get("kind") or "").casefold() == "figure"],
        "warnings": warnings,
        "recovery_point_count": 15,
        "group_plot_resolution": {
            "mode": group_mode,
            "evidence_path": str(group_evidence_path),
            "artifact_count": len(group_entries),
        },
    }
    cable_record["figure_count"] = len(cable_record["figure_paths"])
    records.append(cable_record)

    recovery_bindings: list[dict[str, Any]] = []
    for module, source in (
        ("accel_spectrum", accel_spectrum_manifest_path),
        ("cable_accel_spectrum", cable_spectrum_manifest_path),
    ):
        manifest_path = _require_local_path(source, root, f"{module} recovery manifest")
        manifest = _json_object(manifest_path, f"{module} recovery manifest")
        _request, recovery_config = _validated_request_config(
            manifest, root, expected_module=module, expected_scope=baseline_scope,
            expected_plot_contract=baseline_plot_contract,
        )
        recovered = _rehash_module_record(_successful_record(manifest, module), root)
        module_expectation = expectations.get(module)
        if not isinstance(module_expectation, dict):
            raise ValueError(f"Recovery plot expectations must contain {module}")
        _validate_plot_bucket(
            f"Recovery module {module}", recovered.get("artifacts") or [],
            module_expectation, module=module,
        )
        recovery_contract = _spectrum_output_contract(
            recovery_config, module, baseline_scope[1], baseline_scope[2]
        )
        if recovery_contract != derived_spectrum_contracts[module]:
            raise ValueError(
                f"Recovery config changes the {module} point/group/force output contract"
            )
        stats_path = _require_local_path(
            str(recovered.get("stats_path") or ""), root, f"{module} statistics"
        )
        _validate_spectrum_stats(stats_path, module, recovery_contract)
        _validate_spectrum_figures(
            recovered.get("artifacts") or [], module, recovery_contract
        )
        if module == "cable_accel_spectrum":
            recovered["cable_force_engineering_valid"] = recovery_contract[
                "cable_force_engineering_valid"
            ]
            recovered["cable_force_engineering_status"] = recovery_contract[
                "cable_force_engineering_status"
            ]
            recovered["force_parameter_evidence"] = recovery_contract[
                "force_parameter_evidence"
            ]
            if not recovery_contract["cable_force_engineering_valid"]:
                recovered.setdefault("warnings", []).append(
                    "CableForce figures reproduce the configured calculation path but "
                    "the rho/L parameters are placeholders or are not engineering-verified; "
                    "do not use them for an engineering cable-force conclusion."
                )
        records.append(recovered)
        recovery_bindings.append(_file_binding(manifest_path, f"{module}_manifest"))

    keys = [str(item.get("key") or "") for item in records]
    expected_all = [*expected, *RECOVERY_MODULES]
    if keys != expected_all or len(set(keys)) != 15:
        raise ValueError(f"Composite module order/uniqueness failed: {keys} != {expected_all}")
    contract_path, reporting_contract = _validate_reporting_contract(
        reporting_contract_path, root, records, baseline_config
    )
    globally_seen: set[Path] = set()
    for record in records:
        for artifact in record.get("artifacts") or []:
            path = Path(str(artifact["path"])).resolve()
            if path in globally_seen:
                raise ValueError(f"Artifact is claimed by more than one composite module: {path}")
            globally_seen.add(path)
    if not isinstance(expected_plot_provenance_count, int) or expected_plot_provenance_count <= 0:
        raise ValueError("expected_plot_provenance_count must be a positive integer")
    recovery_bucket_expected_total = sum(
        int(item["expected_plot_provenance_count"]) for item in point_expectations
    ) + int(group_expectation["expected_plot_provenance_count"]) + sum(
        int(expectations[module]["expected_plot_provenance_count"])
        for module in ("accel_spectrum", "cable_accel_spectrum")
    )
    bucket_expected_total = (
        int(published_baseline_provenance_count) + recovery_bucket_expected_total
    )
    if bucket_expected_total != expected_plot_provenance_count:
        raise ValueError(
            "Per-bucket plot expectations do not equal the declared composite total: "
            f"{bucket_expected_total} != {expected_plot_provenance_count}"
        )
    actual_plot_provenance_count = sum(
        str(artifact.get("path") or "").casefold().endswith(".plot.json")
        for record in records for artifact in record.get("artifacts") or []
    )
    if actual_plot_provenance_count != expected_plot_provenance_count:
        raise ValueError(
            "Composite formal plot provenance count is incomplete: "
            f"{actual_plot_provenance_count} != {expected_plot_provenance_count}"
        )
    spectrum_formal_figure_stub_count = sum(
        len(contract["expected_formal_figure_stubs"])
        for contract in derived_spectrum_contracts.values()
    )
    formal_figure_stub_count = (
        actual_plot_provenance_count + spectrum_formal_figure_stub_count
    )

    composite = copy.deepcopy(baseline)
    composite.update(
        {
            "schema_version": max(2, int(baseline.get("schema_version") or 0)),
            "manifest_type": "composite_analysis_recovery",
            "status": "ok",
            "written_at": datetime.now(timezone.utc).isoformat(),
            "module_results": records,
            "module_logs": records,
            "module_artifacts": [
                {"key": item["key"], "label": item.get("label", item["key"]), "artifacts": item.get("artifacts") or []}
                for item in records
            ],
            "artifact_count": sum(len(item.get("artifacts") or []) for item in records),
            "plot_provenance_count": actual_plot_provenance_count,
            # Provenance-backed plots and spectrum formal figures are separate
            # contracts.  SpectrumPlotService currently emits no plot.json, so
            # the Jiulongjiang closure is 42 provenance-backed plots but 88
            # report-facing formal figure stubs (42 + 46), not 88 provenance.
            "formal_figure_stub_count": formal_figure_stub_count,
            "formal_figure_stub_count_by_source": {
                "plot_provenance_backed": actual_plot_provenance_count,
                "accel_spectrum": len(
                    derived_spectrum_contracts["accel_spectrum"][
                        "expected_formal_figure_stubs"
                    ]
                ),
                "cable_accel_spectrum": len(
                    derived_spectrum_contracts["cable_accel_spectrum"][
                        "expected_formal_figure_stubs"
                    ]
                ),
            },
            "cable_force_engineering_valid": derived_spectrum_contracts[
                "cable_accel_spectrum"
            ]["cable_force_engineering_valid"],
            "cable_force_engineering_status": derived_spectrum_contracts[
                "cable_accel_spectrum"
            ]["cable_force_engineering_status"],
            "cable_force_parameter_evidence": derived_spectrum_contracts[
                "cable_accel_spectrum"
            ]["force_parameter_evidence"],
            "cable_force_engineering_note": (
                "CableForce figures are retained only as process-reproduction evidence. "
                "Do not use their values or plots for engineering conclusions until rho/L "
                "parameters are explicitly engineering-verified."
                if not derived_spectrum_contracts["cable_accel_spectrum"][
                    "cable_force_engineering_valid"
                ] else "Cable-force parameters are explicitly marked engineering-verified."
            ),
            "reporting_contract": reporting_contract,
            "stats_files": [item["stats_path"] for item in records if str(item.get("stats_path") or "")],
            "module_status_counts": {"ok": 15, "fail": 0, "skip": 0, "missing": 0, "other": 0},
            "source_chain": {
                "baseline": _file_binding(baseline_path, "baseline_manifest"),
                "cable_accel_points": cable_source_chain,
                "cable_accel_merge_receipt": _file_binding(receipt_path, "cable_accel_merge_receipt"),
                "cable_accel_group_resolution": _file_binding(
                    group_evidence_path, "cable_accel_group_plot_resolution"
                ),
                "recovery_plot_expectations": _file_binding(
                    expectations_path, "recovery_plot_expectations"
                ),
                "reporting_contract": _file_binding(
                    contract_path, "analysis_reporting_contract"
                ),
                "recovery_modules": recovery_bindings,
            },
        }
    )
    raw = _json_bytes(composite, pretty=False)
    _atomic_write(output, raw)
    # Publication gate: exact bytes must decode to an object after atomic replace.
    published = _json_object(output, "Published composite manifest")
    if published.get("status") != "ok" or len(_module_records(published)) != 15:
        output.unlink(missing_ok=True)
        raise RuntimeError("Published composite manifest failed JSON/status/module validation")
    return {
        "path": str(output),
        "bytes": output.stat().st_size,
        "sha256": _sha256_file(output),
        "module_count": 15,
        "artifact_count": composite["artifact_count"],
        "plot_provenance_count": actual_plot_provenance_count,
        "formal_figure_stub_count": formal_figure_stub_count,
        "reporting_contract_sha256": _sha256_file(contract_path),
    }


def _split_csv(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    commands = parser.add_subparsers(dest="command", required=True)
    prepare = commands.add_parser("prepare-requests", help="prepare recovery requests without launching them")
    prepare.add_argument("--base-request", type=Path, required=True)
    prepare.add_argument("--base-config", type=Path)
    prepare.add_argument("--output-dir", type=Path, required=True)
    prepare.add_argument("--rc-root", required=True)
    prepare.add_argument("--remote-bundle-root", required=True)

    merge = commands.add_parser("merge-cable-stats", help="merge 15 one-row cable statistics workbooks")
    merge.add_argument("--points", required=True, help="comma-separated configured point order")
    merge.add_argument("--inputs", nargs="+", type=Path, required=True)
    merge.add_argument("--output", type=Path, required=True)
    merge.add_argument("--allowed-root", type=Path, required=True)
    merge.add_argument("--receipt", type=Path)

    baseline = commands.add_parser(
        "build-baseline-evidence",
        help="rebuild 12-module evidence from a pinned request/config and explicit file inventory",
    )
    baseline.add_argument("--run-request", type=Path, required=True)
    baseline.add_argument("--inventory", type=Path, required=True)
    baseline.add_argument("--modules", required=True)
    baseline.add_argument("--output", type=Path, required=True)
    baseline.add_argument("--allowed-root", type=Path, required=True)

    compose = commands.add_parser("compose-manifest", help="compose a verified 15-module recovery manifest")
    compose.add_argument("--baseline", type=Path, required=True)
    compose.add_argument("--baseline-modules", required=True)
    compose.add_argument("--cable-manifests", nargs="+", type=Path, required=True)
    compose.add_argument("--cable-merge-receipt", type=Path, required=True)
    compose.add_argument("--cable-group-evidence", type=Path, required=True)
    compose.add_argument("--recovery-expectations", type=Path, required=True)
    compose.add_argument("--reporting-contract", type=Path, required=True)
    compose.add_argument("--accel-spectrum-manifest", type=Path, required=True)
    compose.add_argument("--cable-spectrum-manifest", type=Path, required=True)
    compose.add_argument("--output", type=Path, required=True)
    compose.add_argument("--allowed-root", type=Path, required=True)
    compose.add_argument("--expected-plot-provenance-count", type=int, required=True)
    args = parser.parse_args(argv)

    if args.command == "prepare-requests":
        result = prepare_recovery_requests(
            args.base_request,
            args.output_dir,
            rc_root=args.rc_root,
            remote_bundle_root=args.remote_bundle_root,
            base_config_path=args.base_config,
        )
    elif args.command == "merge-cable-stats":
        result = merge_cable_accel_stats(
            _split_csv(args.points), args.inputs, args.output,
            allowed_root=args.allowed_root, receipt_path=args.receipt,
        )
    elif args.command == "build-baseline-evidence":
        result = build_baseline_evidence(
            args.run_request,
            args.inventory,
            args.output,
            expected_modules=_split_csv(args.modules),
            allowed_root=args.allowed_root,
        )
    else:
        result = compose_recovery_manifest(
            args.baseline,
            expected_baseline_modules=_split_csv(args.baseline_modules),
            cable_point_manifest_paths=args.cable_manifests,
            cable_merge_receipt_path=args.cable_merge_receipt,
            cable_group_evidence_path=args.cable_group_evidence,
            recovery_expectations_path=args.recovery_expectations,
            reporting_contract_path=args.reporting_contract,
            accel_spectrum_manifest_path=args.accel_spectrum_manifest,
            cable_spectrum_manifest_path=args.cable_spectrum_manifest,
            output_path=args.output,
            allowed_root=args.allowed_root,
            expected_plot_provenance_count=args.expected_plot_provenance_count,
        )
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
