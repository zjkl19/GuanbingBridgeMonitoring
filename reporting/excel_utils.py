from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def load_sheet_rows(
    path: Path,
    sheet: str | None = None,
    *,
    strip_headers: bool = False,
    skip_empty: bool = False,
    require_exists: bool = True,
) -> list[dict]:
    if not Path(path).exists():
        if require_exists:
            raise FileNotFoundError(path)
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    if strip_headers:
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    else:
        headers = [str(value) if value is not None else "" for value in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if strip_headers:
            item = {key: value for key, value in zip(headers, row) if key}
        else:
            item = {key: value for key, value in zip(headers, row)}
        if skip_empty and not any(value is not None for value in item.values()):
            continue
        out.append(item)
    return out
