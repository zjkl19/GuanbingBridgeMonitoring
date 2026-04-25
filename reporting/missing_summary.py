from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_HEADERS = ["序号", "报告类型", "类别", "章节/模块", "对象", "说明", "严重性", "来源"]


def missing_summary_paths(report_path: Path) -> tuple[Path, Path]:
    return (
        report_path.with_name(f"{report_path.stem}_missing_summary.txt"),
        report_path.with_name(f"{report_path.stem}_missing_summary.xlsx"),
    )


def _missing_string_to_item(text: str) -> dict[str, str]:
    raw = str(text).strip()
    if raw.startswith("section:"):
        parts = raw.split(":", 2)
        section = parts[1] if len(parts) > 1 else ""
        detail = parts[2] if len(parts) > 2 else raw
        return {
            "category": "章节内容缺失",
            "section": section,
            "item": section,
            "detail": detail,
            "severity": "warning",
            "source": raw,
        }
    if raw.startswith("warning:"):
        detail = raw.split(":", 1)[1] if ":" in raw else raw
        return {
            "category": "生成警告",
            "section": "",
            "item": "",
            "detail": detail,
            "severity": "warning",
            "source": raw,
        }
    if raw.startswith("wim:"):
        parts = raw.split(":", 2)
        section = "WIM"
        item = parts[1] if len(parts) > 1 else ""
        detail = parts[2] if len(parts) > 2 else raw
        return {
            "category": "WIM图表缺失",
            "section": section,
            "item": item,
            "detail": detail,
            "severity": "warning",
            "source": raw,
        }
    if ":" in raw:
        section, item = raw.split(":", 1)
    else:
        section, item = "", raw
    return {
        "category": "图表/资源缺失",
        "section": section,
        "item": item,
        "detail": "报告生成时未找到对应图片或资源。",
        "severity": "warning",
        "source": raw,
    }


def normalize_missing_items(items: Iterable[str | dict[str, Any]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    for item in items:
        if isinstance(item, str):
            normalized.append(_missing_string_to_item(item))
            continue
        if not isinstance(item, dict):
            normalized.append(_missing_string_to_item(str(item)))
            continue
        normalized.append(
            {
                "category": str(item.get("category", "内容缺失")),
                "section": str(item.get("section", "")),
                "item": str(item.get("item", "")),
                "detail": str(item.get("detail", "")),
                "severity": str(item.get("severity", "warning")),
                "source": str(item.get("source", "")),
            }
        )
    return normalized


def _write_txt(path: Path, report_type: str, report_path: Path, rows: list[dict[str, str]], context: dict[str, Any] | None) -> None:
    lines = [
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"报告类型: {report_type}",
        f"报告文件: {report_path}",
        f"缺失/提示数量: {len(rows)}",
    ]
    if context:
        lines.append("")
        lines.append("上下文:")
        for key, value in context.items():
            lines.append(f"- {key}: {value}")
    lines.append("")
    if rows:
        lines.append("缺失/提示明细:")
        for idx, row in enumerate(rows, start=1):
            lines.append(
                f"{idx}. [{row['severity']}] {row['category']} | "
                f"{row['section']} | {row['item']} | {row['detail']}"
            )
    else:
        lines.append("未发现缺失内容。")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_xlsx(path: Path, report_type: str, report_path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "缺失内容清单"
    ws.append(["报告类型", report_type])
    ws.append(["报告文件", str(report_path)])
    ws.append(["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["缺失/提示数量", len(rows)])
    ws.append([])
    ws.append(SUMMARY_HEADERS)

    header_row = 6
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if rows:
        for idx, row in enumerate(rows, start=1):
            ws.append(
                [
                    idx,
                    report_type,
                    row["category"],
                    row["section"],
                    row["item"],
                    row["detail"],
                    row["severity"],
                    row["source"],
                ]
            )
    else:
        ws.append([1, report_type, "无缺失", "", "", "未发现缺失内容。", "ok", ""])

    widths = [8, 18, 18, 24, 30, 58, 12, 45]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A7"
    wb.save(path)


def write_missing_summary(
    report_type: str,
    report_path: Path,
    missing_items: Iterable[str | dict[str, Any]],
    context: dict[str, Any] | None = None,
) -> tuple[Path, Path]:
    txt_path, xlsx_path = missing_summary_paths(report_path)
    rows = normalize_missing_items(missing_items)
    _write_txt(txt_path, report_type, report_path, rows, context)
    _write_xlsx(xlsx_path, report_type, report_path, rows)
    return txt_path, xlsx_path
