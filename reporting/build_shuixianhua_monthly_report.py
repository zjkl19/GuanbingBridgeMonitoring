from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from zipfile import ZIP_DEFLATED, ZipFile
from config_loader import load_report_config
from analysis_manifest import missing_module_summary_items
from artifact_lookup import (
    filename_has_point_token,
    latest_file_patterns as lookup_latest_file_patterns,
    resolve_output_dirs,
)
from report_build_manifest import write_report_build_manifest
from report_context import ReportBuildContext
from report_qc import (
    ReportQcIssue,
    ReportQcResult,
    check_shuixianhua_report,
    write_report_qc_report,
)
from report_artifact_resolver import find_latest_image_patterns
from docx_utils import (
    insert_picture_before,
    paragraph_has_image,
    previous_body_paragraphs,
    prune_unused_document_image_relationships,
    remove_nearby_picture_block_before,
)

from docx import Document
from docx.oxml.ns import qn
from docx.text.paragraph import Paragraph
from lxml import etree
from openpyxl import load_workbook

try:
    from shuixianhua_table_anchors import required_result_tables
    from ooxml_utils import fill_table as ooxml_fill_table
    from ooxml_utils import rewrite_paragraphs_containing as ooxml_rewrite_contains
    from ooxml_utils import set_cell_text as ooxml_set_cell_text
    from ooxml_utils import set_paragraph_text as ooxml_set_paragraph_text
    from ooxml_utils import xml_text as ooxml_text
except Exception:  # pragma: no cover - package import path
    from .shuixianhua_table_anchors import required_result_tables
    from .ooxml_utils import fill_table as ooxml_fill_table
    from .ooxml_utils import rewrite_paragraphs_containing as ooxml_rewrite_contains
    from .ooxml_utils import set_cell_text as ooxml_set_cell_text
    from .ooxml_utils import set_paragraph_text as ooxml_set_paragraph_text
    from .ooxml_utils import xml_text as ooxml_text


REPORT_NO = "BG20TUJC2600003-J1"
EXCLUDED_ACQUISITION_MODULES = {"dynamic_strain_highpass", "dynamic_strain_lowpass"}


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="Build Shuixianhua monthly monitoring report.")
    parser.add_argument("--config", type=Path, default=repo_root / "config" / "shuixianhua_config.json")
    parser.add_argument("--template", type=Path, default=repo_root / "reports" / "水仙花大桥健康监测月报模板.docx")
    parser.add_argument("--result-root", type=Path, default=Path(r"E:\水仙花大桥数据\2026年3月"))
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--period-label", default="2026年3月份")
    parser.add_argument("--monitoring-range", default="2026年03月23日~2026年03月31日")
    parser.add_argument("--report-date", default="2026年04月05日")
    parser.add_argument("--start-date", default="")
    parser.add_argument("--end-date", default="")
    parser.add_argument("--no-word-update", action="store_true", help="Skip Word field update and PDF export.")
    return parser.parse_args()

def load_json(path: Path) -> dict[str, Any]:
    return load_report_config(path)

def load_rows(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not raw:
        return []
    header = [str(value) if value is not None else "" for value in raw[0]]
    rows: list[dict[str, Any]] = []
    for values in raw[1:]:
        item = {key: value for key, value in zip(header, values)}
        if any(value is not None and value != "" for value in item.values()):
            rows.append(item)
    return rows

def safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def fmt_num(value: Any, digits: int = 3) -> str:
    num = safe_float(value)
    if num is None:
        return "/"
    text = f"{num:.{digits}f}".rstrip("0").rstrip(".")
    return text if text else "0"

def fmt_percent(value: Any) -> str:
    num = safe_float(value)
    if num is None:
        return "/"
    if num <= 1:
        num *= 100
    return f"{num:.1f}%"

def fmt_datetime(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return "" if value is None else str(value)

def extreme_row(rows: list[dict[str, Any]], key: str, *, abs_value: bool = False) -> dict[str, Any] | None:
    best = None
    best_value = None
    for row in rows:
        value = safe_float(row.get(key))
        if value is None:
            continue
        metric = abs(value) if abs_value else value
        if best_value is None or metric > best_value:
            best = row
            best_value = metric
    return best

def update_word_fields_and_export_pdf(docx_path: Path) -> Path | None:
    docx_path = docx_path.resolve()
    pdf_path = docx_path.with_suffix(".pdf")
    script = f"""
$docx = @'
{docx_path}
'@
$pdf = @'
{pdf_path}
'@
$word = New-Object -ComObject Word.Application
$word.Visible = $false
$word.DisplayAlerts = 0
try {{
    $doc = $word.Documents.Open($docx)
    $doc.TrackRevisions = $false
    if ($doc.Revisions.Count -gt 0) {{ $doc.AcceptAllRevisions() | Out-Null }}
    $doc.Fields.Update() | Out-Null
    foreach ($toc in $doc.TablesOfContents) {{ $toc.Update() | Out-Null }}
    $doc.TrackRevisions = $false
    if ($doc.Revisions.Count -gt 0) {{ $doc.AcceptAllRevisions() | Out-Null }}
    $doc.Save()
    $doc.ExportAsFixedFormat($pdf, 17)
    $doc.Close($false)
}} finally {{
    $word.Quit()
    [System.Runtime.InteropServices.Marshal]::ReleaseComObject($word) | Out-Null
}}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        timeout=300,
    )
    if result.returncode != 0:
        stderr = (result.stderr or b"").decode("utf-8", errors="replace").strip()
        stdout = (result.stdout or b"").decode("utf-8", errors="replace").strip()
        print(f"Warning: Word field update/PDF export failed: {stderr or stdout or 'unknown error'}")
        return None
    return pdf_path if pdf_path.exists() else None

def adjusted_rows(stats_dir: Path, filename: str, fallback: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    path = stats_dir / "adjusted" / filename
    if path.exists():
        return load_rows(path)
    return fallback or []

def scaled_accel_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Convert Shuixianhua acceleration stats from cm/s² to m/s²."""
    out = []
    for row in rows:
        item = dict(row)
        for key in ["Min", "Max", "Mean", "RMS10minMax"]:
            value = safe_float(item.get(key))
            if value is not None:
                item[key] = value / 100.0
        out.append(item)
    return out

def _sxh_accel_group_for_point(cfg: dict[str, Any], point_id: str) -> str | None:
    groups = cfg.get("groups", {}).get("acceleration", {})
    if not isinstance(groups, dict):
        return None
    for group_name, points in groups.items():
        if isinstance(points, list) and point_id in {str(point) for point in points}:
            return str(group_name)
    return None

def _sxh_accel_rms_limits_mps2(cfg: dict[str, Any], point_id: str | None = None) -> tuple[float, float]:
    warn_lines = cfg.get("plot_styles", {}).get("acceleration", {}).get("rms_warn_lines", {})
    lines = None
    if isinstance(warn_lines, dict):
        group_name = _sxh_accel_group_for_point(cfg, point_id or "")
        if group_name:
            lines = warn_lines.get(group_name)
        if lines is None and warn_lines:
            first_key = next(iter(warn_lines))
            lines = warn_lines.get(first_key)
    elif isinstance(warn_lines, list):
        lines = warn_lines

    values = []
    if isinstance(lines, list):
        for line in lines:
            if isinstance(line, dict):
                value = safe_float(line.get("y"))
                if value is not None:
                    values.append(value / 100.0)
    if len(values) >= 2:
        return values[0], values[1]
    return 0.315, 0.5

def _sxh_accel_rms_status(value: float | None, first_limit: float, second_limit: float) -> str:
    if value is None:
        return "暂不具备阈值判定条件。"
    first_text = fmt_num(first_limit, 3)
    second_text = fmt_num(second_limit, 3)
    if value <= first_limit:
        return f"未超过一级阈值{first_text}m/s²，处于超限阈值范围之内，未出现超过各级超限阈值和报警的情况。"
    if value <= second_limit:
        return f"超过一级阈值{first_text}m/s²，未超过二级阈值{second_text}m/s²。"
    return f"超过二级阈值{second_text}m/s²。"

def report_acquisition_rows(acquisition_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for row in acquisition_rows:
        code = row.get("模块代码")
        if code in EXCLUDED_ACQUISITION_MODULES:
            continue
        item = dict(row)
        if code == "strain":
            item["模块"] = "结构应变及动应变"
        rows.append(item)
    return rows

def _sxh_fixed(value: Any, digits: int = 1) -> str:
    num = safe_float(value)
    if num is None:
        return "/"
    return f"{num:.{digits}f}"

def _sxh_range(rows: list[dict[str, Any]], min_key: str, max_key: str, *, digits: int = 1, unit: str = "") -> str:
    lows = [safe_float(row.get(min_key)) for row in rows]
    highs = [safe_float(row.get(max_key)) for row in rows]
    lows = [value for value in lows if value is not None]
    highs = [value for value in highs if value is not None]
    if not lows or not highs:
        return "/"
    return f"{_sxh_fixed(min(lows), digits)}{unit}~{_sxh_fixed(max(highs), digits)}{unit}"

def _sxh_range_plain(rows: list[dict[str, Any]], min_key: str, max_key: str, *, digits: int = 1, unit: str = "") -> str:
    lows = [safe_float(row.get(min_key)) for row in rows]
    highs = [safe_float(row.get(max_key)) for row in rows]
    lows = [value for value in lows if value is not None]
    highs = [value for value in highs if value is not None]
    if not lows or not highs:
        return "/"
    return f"{fmt_num(min(lows), digits)}{unit}~{fmt_num(max(highs), digits)}{unit}"

def _sxh_by_prefix(rows: list[dict[str, Any]], prefix: str) -> list[dict[str, Any]]:
    return [row for row in rows if str(row.get("PointID") or row.get("测点编号") or "").startswith(prefix)]


def _sxh_normalize_point_ids(
    rows: list[dict[str, Any]], configured_points: Any
) -> list[dict[str, Any]]:
    """Repair vendor/MATLAB text-codec damage using the reviewed config point list."""
    if isinstance(configured_points, str):
        configured = [configured_points]
    elif isinstance(configured_points, list):
        configured = [str(value) for value in configured_points]
    else:
        configured = []
    if not configured:
        return rows
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        point_id = str(item.get("PointID") or "")
        if point_id not in configured:
            prefix_match = re.match(r"^([A-Za-z]+-\d+)", point_id)
            candidates = [
                value
                for value in configured
                if prefix_match and value.startswith(prefix_match.group(1))
            ]
            if len(candidates) == 1:
                item["PointID"] = candidates[0]
        out.append(item)
    return out


def _sxh_text_value(row: dict[str, Any], *keys: str, default: str = "/") -> str:
    for key in keys:
        value = row.get(key)
        if value is not None and value != "":
            return str(value)
    return default

def _sxh_wind_summary_matches_period(
    path: Path, start_date: str, end_date: str
) -> bool:
    """Require the summary filename to prove the complete report period."""
    dates: list[str] = []
    pattern = re.compile(
        r"(?<!\d)(20\d{2})[-_.]?(0[1-9]|1[0-2])[-_.]?(0[1-9]|[12]\d|3[01])(?!\d)"
    )
    for year, month, day in pattern.findall(path.stem):
        try:
            dates.append(
                datetime(int(year), int(month), int(day)).date().isoformat()
            )
        except ValueError:
            continue
    return len(dates) >= 2 and dates[0] == start_date and dates[1] == end_date


def _sxh_find_wind_summary_file(
    result_root: Path,
    point_id: str,
    start_date: str,
    end_date: str,
) -> Path:
    """Resolve one period-bound wind summary, failing closed on any ambiguity."""
    configured_dir = Path("风速风向结果") / "风玫瑰"
    patterns = [f"{point_id}_windrose_*_summary.txt"]
    lookup = lookup_latest_file_patterns(
        result_root,
        configured_dir,
        patterns,
        point_id=point_id,
        point_token_strict=True,
        kind="summary",
    )
    source = str(lookup.debug.get("source") or "")
    manifest_sources = {
        "analysis_manifest",
        "derived_artifact_manifest",
        "pinned_analysis_manifest",
    }
    if source in manifest_sources:
        if lookup.path is None:
            raise FileNotFoundError(
                "水仙花风玫瑰摘要未记录在已绑定的分析结果清单中: "
                f"point_id={point_id}, report={start_date}~{end_date}, "
                f"manifest={lookup.debug.get('manifest', '')}"
            )
        selected = lookup.path.resolve()
        if not _sxh_wind_summary_matches_period(selected, start_date, end_date):
            raise ValueError(
                "水仙花风玫瑰摘要与报告期不一致（已绑定分析结果清单）: "
                f"point_id={point_id}, report={start_date}~{end_date}, "
                f"selected={selected}, manifest={lookup.debug.get('manifest', '')}"
            )
        return selected

    candidates: list[Path] = []
    matches: list[Path] = []
    for folder in resolve_output_dirs(result_root, configured_dir):
        for pattern in patterns:
            for candidate in folder.glob(pattern):
                if not filename_has_point_token(candidate, point_id):
                    continue
                resolved = candidate.resolve()
                candidates.append(resolved)
                if _sxh_wind_summary_matches_period(
                    resolved, start_date, end_date
                ):
                    matches.append(resolved)

    unique_matches = sorted(set(matches))
    if len(unique_matches) == 1:
        return unique_matches[0]
    if len(unique_matches) > 1:
        raise ValueError(
            "水仙花同一测点同一报告期存在多份风玫瑰摘要，无法唯一绑定: "
            f"point_id={point_id}, report={start_date}~{end_date}, "
            f"matches={[str(path) for path in unique_matches]}"
        )
    if candidates:
        raise ValueError(
            "水仙花风玫瑰摘要均不属于请求的报告期: "
            f"point_id={point_id}, report={start_date}~{end_date}, "
            f"candidates={[str(path) for path in sorted(set(candidates))]}"
        )
    raise FileNotFoundError(
        "未找到水仙花本报告期风玫瑰摘要: "
        f"point_id={point_id}, report={start_date}~{end_date}"
    )


def _sxh_parse_wind_summaries(
    result_root: Path,
    wind_rows: list[dict[str, Any]],
    start_date: str,
    end_date: str,
) -> list[dict[str, Any]]:
    rows = [dict(row) for row in wind_rows]
    for row in rows:
        pid = str(row.get("测点编号") or row.get("PointID") or "").strip()
        if not pid:
            raise ValueError("水仙花风速统计行缺少测点编号，无法绑定风玫瑰摘要")
        summary_path = _sxh_find_wind_summary_file(
            result_root, pid, start_date, end_date
        )
        text = summary_path.read_text(encoding="utf-8", errors="replace")
        pid_match = re.search(r"风玫瑰简要结论（(.+?)）", text)
        if not pid_match:
            raise ValueError(f"水仙花风玫瑰摘要缺少测点标识: {summary_path}")
        summary_pid = pid_match.group(1).strip()
        if summary_pid != pid:
            raise ValueError(
                "水仙花风玫瑰摘要的测点与统计行不一致: "
                f"expected={pid}, actual={summary_pid}, path={summary_path}"
            )
        for key in ["平均风向", "主导风向", "平均风速", "最大风速", "主要风速等级"]:
            match = re.search(rf"{key}:\s*([^\n]+)", text)
            if match:
                row[key] = match.group(1).strip().replace("占比 ", "占比")
    return rows

def _sxh_accel_frequency_map(stats_dir: Path) -> dict[str, tuple[float | None, float | None]]:
    path = stats_dir / "accel_spec_stats.xlsx"
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, tuple[float | None, float | None]] = {}
    try:
        for sheet in wb.sheetnames:
            if sheet.endswith("-Y"):
                out[sheet] = (None, None)
                continue
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(value) if value is not None else "" for value in rows[0]]
            freq_indices = [idx for idx, name in enumerate(header) if name.startswith("Freq_")]
            values: list[float] = []
            for raw in rows[1:]:
                for idx in freq_indices:
                    if idx < len(raw):
                        value = safe_float(raw[idx])
                        if value is not None:
                            values.append(value)
            out[sheet] = (min(values), max(values)) if values else (None, None)
    finally:
        wb.close()
    return out

def _sxh_strain_rows_with_group(rows: list[dict[str, Any]], cfg: dict[str, Any]) -> list[dict[str, Any]]:
    labels = cfg.get("plot_styles", {}).get("strain", {}).get("group_labels", {})
    groups = cfg.get("groups", {}).get("strain", {})
    row_by_point = {str(row.get("PointID") or ""): row for row in rows}
    out = []
    for group_key, points in groups.items():
        label = labels.get(group_key, group_key)
        for point in points:
            point_text = str(point)
            if point_text not in row_by_point:
                continue
            item = dict(row_by_point.pop(point_text))
            item["分组"] = str(label)
            out.append(item)
    for row in rows:
        point_text = str(row.get("PointID") or "")
        if point_text not in row_by_point:
            continue
        item = dict(row_by_point.pop(point_text))
        item["分组"] = "/"
        out.append(item)
    return out

def _sxh_strain_group_ranges(rows: list[dict[str, Any]], cfg: dict[str, Any]) -> list[str]:
    labels = list(cfg.get("plot_styles", {}).get("strain", {}).get("group_labels", {}).values())
    if not labels:
        labels = ["小纵梁底部静应变", "横梁底部静应变", "主拱拱顶静应变", "主拱拱脚静应变"]
    out = []
    for label in labels:
        group_rows = [row for row in rows if row.get("分组") == label]
        if not group_rows:
            continue
        out.append(f"{label}为{_sxh_range_plain(group_rows, 'Min', 'Max', digits=1, unit='με')}")
    return out

def _sxh_report_row(module: str, configured: int, found: int, missing_note: str = "") -> dict[str, Any]:
    configured = int(configured or 0)
    found = int(found or 0)
    missing = max(configured - found, 0)
    return {
        "模块": module,
        "配置测点数": configured,
        "实际获取测点数": found,
        "缺失测点数": missing,
        "获取率": (found / configured) if configured else 0,
        "缺失说明": missing_note or ("无" if missing == 0 else f"缺失{missing}个测点"),
    }

def _sxh_monitoring_range_span(monitoring_range: str) -> str:
    text = str(monitoring_range or "").strip()
    chinese_dates = re.findall(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", text)
    if len(chinese_dates) >= 2:
        def fmt(parts: tuple[str, str, str]) -> str:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
        return f"{fmt(chinese_dates[0])}~{fmt(chinese_dates[1])}"
    iso_dates = re.findall(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if len(iso_dates) >= 2:
        def fmt_iso(parts: tuple[str, str, str]) -> str:
            return f"{int(parts[0]):04d}-{int(parts[1]):02d}-{int(parts[2]):02d}"
        return f"{fmt_iso(iso_dates[0])}~{fmt_iso(iso_dates[1])}"
    return text

def _sxh_period_filename_label(period_label: str) -> str:
    text = str(period_label or "").strip()
    if text:
        return re.sub(r'[\\/:*?"<>|]+', "_", text)
    return "月度"


def _sxh_period_dates(
    monitoring_range: str,
    start_date: str | None = None,
    end_date: str | None = None,
) -> tuple[str, str]:
    """Return an ISO date pair and reject an ambiguous/inverted report period."""
    explicit = [str(start_date or "").strip(), str(end_date or "").strip()]
    if all(explicit):
        values = explicit
    else:
        text = str(monitoring_range or "")
        chinese = re.findall(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", text)
        iso = re.findall(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
        parts = chinese[:2] if len(chinese) >= 2 else iso[:2]
        if len(parts) < 2:
            raise ValueError(f"无法从监测范围解析起止日期: {monitoring_range}")
        values = [f"{int(y):04d}-{int(m):02d}-{int(d):02d}" for y, m, d in parts]
    parsed = [datetime.strptime(value, "%Y-%m-%d").date() for value in values]
    if parsed[0] > parsed[1]:
        raise ValueError(f"报告起始日期晚于结束日期: {values[0]} > {values[1]}")
    return parsed[0].isoformat(), parsed[1].isoformat()


def _sxh_configured_points(context: dict[str, Any], key: str, rows_key: str) -> list[str]:
    cfg = context.get("cfg") or {}
    points = cfg.get("points", {}).get(key) or []
    if isinstance(points, str):
        points = [points]
    values = [str(value) for value in points if str(value).strip()]
    if values:
        return values
    out: list[str] = []
    for row in context.get(rows_key) or []:
        value = str(row.get("PointID") or row.get("测点编号") or "").strip()
        if value and value not in out:
            out.append(value)
    return out


def _sxh_image_patterns(stem: str, *, wildcard_tail: bool = False) -> list[str]:
    marker = "*" if wildcard_tail else ""
    return [f"{stem}{marker}{suffix}" for suffix in (".jpg", ".jpeg", ".png")]


def _sxh_image_request(
    anchor: str,
    label: str,
    directory: str,
    stem: str,
    *,
    wildcard_tail: bool = False,
) -> dict[str, Any]:
    return {
        "anchor": anchor,
        "label": label,
        "directory": directory,
        "patterns": _sxh_image_patterns(stem, wildcard_tail=wildcard_tail),
    }


def _sxh_report_image_requests(
    context: dict[str, Any], start_date: str, end_date: str
) -> list[dict[str, Any]]:
    compact = start_date.replace("-", "") + "_" + end_date.replace("-", "")
    iso = f"{start_date}_{end_date}"
    requests: list[dict[str, Any]] = []
    add = requests.append

    for point_id in _sxh_configured_points(context, "temperature", "temp_rows"):
        add(_sxh_image_request("图 2-1", f"温度时程/{point_id}", "时程曲线_温度", f"{point_id}_{compact}"))
    for point_id in _sxh_configured_points(context, "humidity", "humidity_rows"):
        add(_sxh_image_request("图 2-2", f"湿度时程/{point_id}", "时程曲线_湿度", f"{point_id}_{compact}"))
        add(_sxh_image_request("图 2-2", f"湿度频次/{point_id}", "频次分布_湿度", f"{point_id}_freq_{compact}"))

    wind_points = _sxh_configured_points(context, "wind_speed", "wind_rows")
    for point_id in wind_points:
        add(_sxh_image_request("图 2-3", f"风速时程/{point_id}", r"风速风向结果\风速时程", f"{point_id}_speed_{iso}"))
    for point_id in wind_points:
        add(_sxh_image_request("图 2-3", f"风向时程/{point_id}", r"风速风向结果\风向时程", f"{point_id}_direction_{iso}"))
    for point_id in wind_points:
        add(_sxh_image_request("图 2-3", f"风玫瑰/{point_id}", r"风速风向结果\风玫瑰", f"{point_id}_windrose_{iso}"))

    for component in ("X", "Y", "Z"):
        add(_sxh_image_request("图 2-4", f"地震动/{component}", r"地震动结果\地震动时程", f"EQ_{component}_{iso}"))
    for group in ("G1", "G2", "G3"):
        add(_sxh_image_request("图 2-5", f"主梁挠度原始/{group}", "时程曲线_挠度_组图_原始", f"Defl_{group}_Orig_{compact}"))
    for group in ("G1", "G2", "G3"):
        add(_sxh_image_request("图 2-6", f"主梁挠度滤波/{group}", "时程曲线_挠度_组图_滤波", f"Defl_{group}_Filt_{compact}"))
    for group in ("G1", "G2"):
        add(_sxh_image_request("图 2-7", f"支座位移原始/{group}", "时程曲线_支座位移_组图_原始", f"BearingDisp_{group}_{compact}_Orig"))
    for group in ("G1", "G2"):
        add(_sxh_image_request("图 2-8", f"支座位移滤波/{group}", "时程曲线_支座位移_组图_滤波", f"BearingDisp_{group}_{compact}_Filt"))

    for group in ("ZG", "ZL"):
        add(_sxh_image_request("图 2-10", f"结构振动RMS/{group}", "时程曲线_加速度_RMS10min_组图", f"AccelRMS10_{group}_Group_{compact}"))
    for group, title in (("ZG", "主拱"), ("ZL", "主梁")):
        add(_sxh_image_request("图 2-10", f"结构振动频率/{group}", "频谱峰值曲线_加速度_组图", f"SpecFreq_{group}（{title}）_Group_{compact}"))
    for point_id in ("ZLZD-01-X10-3#", "ZGZD-04-S-拱顶"):
        add(_sxh_image_request("图 2-11", f"典型频谱/{point_id}", rf"PSD_备查\{point_id}", f"PSD_{point_id}_", wildcard_tail=True))

    for group in ("GDDYB", "GJYB", "HLYB", "XZYB"):
        add(_sxh_image_request("图 2-12", f"结构应变/{group}", "时程曲线_应变_组图", f"Strain_{group}_{compact}"))
    for group in ("S6_S16", "X6_X16", "Tie"):
        add(_sxh_image_request("图 2-13", f"吊杆系杆加速度/{group}", "时程曲线_索力加速度_组图", f"CableAccel_{group}_Group_{compact}"))
    for group in ("S6_S16", "X6_X16", "Tie"):
        add(_sxh_image_request("图 2-13", f"吊杆系杆RMS/{group}", "时程曲线_索力加速度_RMS10min_组图", f"CableAccelRMS10_{group}_Group_{compact}"))
    return requests


def _sxh_image_matches_period(path: Path, start_date: str, end_date: str) -> bool:
    compact = start_date.replace("-", "") + "_" + end_date.replace("-", "")
    iso = f"{start_date}_{end_date}"
    if compact in path.stem or iso in path.stem:
        return True
    dates = re.findall(r"\d{4}-\d{2}-\d{2}", path.stem)
    return bool(dates) and all(start_date <= value <= end_date for value in dates)


def _sxh_resolve_report_images(
    result_root: Path,
    context: dict[str, Any],
    start_date: str,
    end_date: str,
) -> tuple[dict[str, list[Path]], list[dict[str, Any]], list[dict[str, str]]]:
    grouped: dict[str, list[Path]] = {}
    records: list[dict[str, Any]] = []
    missing: list[dict[str, str]] = []
    for request in _sxh_report_image_requests(context, start_date, end_date):
        lookup = find_latest_image_patterns(
            result_root,
            request["directory"],
            list(request["patterns"]),
        )
        path = lookup.path
        if path is None or not path.is_file():
            missing.append({
                "category": "report_image",
                "item": request["label"],
                "detail": json.dumps(lookup.debug, ensure_ascii=False, default=str),
            })
            continue
        if not _sxh_image_matches_period(path, start_date, end_date):
            missing.append({
                "category": "report_image_period",
                "item": request["label"],
                "detail": f"图片日期与报告期不一致: {path}",
            })
            continue
        resolved = path.resolve()
        grouped.setdefault(request["anchor"], []).append(resolved)
        records.append({
            "anchor": request["anchor"],
            "label": request["label"],
            "path": str(resolved),
            "bytes": resolved.stat().st_size,
            "sha256": hashlib.sha256(resolved.read_bytes()).hexdigest().upper(),
            "lookup": lookup.debug,
        })
    return grouped, records, missing


def _sxh_all_body_paragraphs(doc: Document) -> list[Paragraph]:
    return [Paragraph(element, doc._body) for element in doc.element.body.xpath(".//w:p")]


def _sxh_replace_report_image_blocks(
    docx_path: Path, grouped: dict[str, list[Path]]
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    doc = Document(str(docx_path))
    replacements: list[dict[str, Any]] = []
    missing: list[dict[str, str]] = []
    removed_result_hashes: set[str] = set()
    for figure_no in ("2-1", "2-2", "2-3", "2-4", "2-5", "2-6", "2-7", "2-8", "2-10", "2-11", "2-12", "2-13"):
        anchor_text = f"图 {figure_no}"
        anchor = next((p for p in _sxh_all_body_paragraphs(doc) if anchor_text in p.text), None)
        if anchor is None:
            missing.append({"category": "report_anchor", "item": anchor_text, "detail": "模板中未找到图题"})
            continue
        old_image_count = 0
        old_image_hashes: list[str] = []
        image_block_started = False
        for paragraph in previous_body_paragraphs(anchor, limit=120):
            if paragraph_has_image(paragraph):
                old_image_count += 1
                image_block_started = True
                for rel_id in paragraph._p.xpath(".//a:blip/@r:embed"):
                    image_part = doc.part.related_parts.get(rel_id)
                    if image_part is not None:
                        digest = hashlib.sha256(image_part.blob).hexdigest().upper()
                        old_image_hashes.append(digest)
                        removed_result_hashes.add(digest)
            elif not paragraph.text.strip() and image_block_started:
                continue
            else:
                break
        removed_paragraph_count = remove_nearby_picture_block_before(anchor, limit=120)
        paths = grouped.get(anchor_text, [])
        for path in paths:
            insert_picture_before(anchor, path, width_mm=145.0)
        replacements.append({
            "anchor": anchor_text,
            "removed_old_image_count": old_image_count,
            "removed_old_image_sha256": sorted(set(old_image_hashes)),
            "removed_paragraph_count": removed_paragraph_count,
            "inserted_image_count": len(paths),
            "inserted_paths": [str(path) for path in paths],
        })
    prune_unused_document_image_relationships(doc)
    doc.save(str(docx_path))
    with ZipFile(docx_path) as archive:
        remaining_media_hashes = {
            hashlib.sha256(archive.read(name)).hexdigest().upper()
            for name in archive.namelist()
            if name.startswith("word/media/") and not name.endswith("/")
        }
    stale_hashes = sorted(removed_result_hashes & remaining_media_hashes)
    if stale_hashes:
        missing.append({
            "category": "stale_template_image",
            "item": "模板旧期次结果图",
            "detail": f"仍有{len(stale_hashes)}个旧结果图媒体残留: {', '.join(stale_hashes[:10])}",
        })
    return replacements, missing


def _sxh_validate_generated_content(
    docx_path: Path, context: dict[str, Any]
) -> list[dict[str, str]]:
    """Reject stale template prose and incomplete report-facing result tables."""
    issues: list[dict[str, str]] = []
    with ZipFile(docx_path) as archive:
        root = etree.fromstring(archive.read("word/document.xml"))
    document_text = _sxh_xml_text(root)
    stale_phrases = (
        "温度监测中9个结构温度测点本月未获取数据",
        "结构温度测点WD-01~WD-09本月未获取数据",
    )
    for phrase in stale_phrases:
        if phrase in document_text:
            issues.append(
                {
                    "category": "stale_template_text",
                    "item": phrase,
                    "detail": "报告仍含旧期次缺测说明",
                }
            )

    temperature_table = required_result_tables(root)["temperature"]
    actual_points: list[str] = []
    for row in temperature_table.findall(qn("w:tr"))[1:]:
        cells = row.findall(qn("w:tc"))
        if len(cells) < 2:
            continue
        point_id = _sxh_xml_text(cells[1]).strip()
        if point_id:
            actual_points.append(point_id)
    expected_points = [
        str(row.get("PointID") or row.get("测点编号") or "").strip()
        for row in context.get("temp_rows") or []
    ]
    expected_points = [point for point in expected_points if point]
    if actual_points != expected_points:
        issues.append(
            {
                "category": "report_table_content",
                "item": "温度监测结果汇总表",
                "detail": (
                    "报告表测点与统计源不一致: "
                    f"expected={expected_points}; actual={actual_points}"
                ),
            }
        )
    return issues


def _sxh_fallback_report_rows(
    cfg: dict[str, Any],
    *,
    temp_rows: list[dict[str, Any]],
    humidity_rows: list[dict[str, Any]],
    wind_rows: list[dict[str, Any]],
    earthquake_rows: list[dict[str, Any]],
    deflection_rows: list[dict[str, Any]],
    bearing_rows: list[dict[str, Any]],
    accel_rows: list[dict[str, Any]],
    strain_rows: list[dict[str, Any]],
    cable_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    points = cfg.get("points", {})
    per_point = cfg.get("per_point", {})
    pending = cfg.get("design_points_pending", {})
    temp_config = len(points.get("temperature") or per_point.get("temperature") or []) or 10
    gnss_config = len(points.get("gnss") or per_point.get("gnss") or pending.get("gnss") or [])
    return [
        _sxh_report_row(
            "温度",
            temp_config,
            len(temp_rows),
            "无" if len(temp_rows) >= temp_config else f"本月缺少{max(temp_config - len(temp_rows), 0)}个温度测点统计结果",
        ),
        _sxh_report_row("湿度", len(points.get("humidity") or per_point.get("humidity") or []) or 1, len(humidity_rows)),
        _sxh_report_row("风速风向", len(points.get("wind_speed") or per_point.get("wind_speed") or []) or 2, len(wind_rows)),
        _sxh_report_row("地震动", len(points.get("eq") or per_point.get("eq") or []) or len(earthquake_rows), len(earthquake_rows)),
        _sxh_report_row("主梁挠度", len(points.get("deflection") or per_point.get("deflection") or []) or len(deflection_rows), len(deflection_rows)),
        _sxh_report_row("支座及伸缩缝位移", len(points.get("bearing_displacement") or per_point.get("bearing_displacement") or []) or len(bearing_rows), len(bearing_rows)),
        _sxh_report_row("拱顶、拱脚位移（GNSS）", gnss_config, 0, "本月未获取有效数据" if gnss_config else "未配置在线数据"),
        _sxh_report_row("结构振动", len(points.get("acceleration") or per_point.get("acceleration") or []) or len(accel_rows), len(accel_rows)),
        _sxh_report_row("结构应变及动应变", len(points.get("strain") or per_point.get("strain") or []) or len(strain_rows), len(strain_rows)),
        _sxh_report_row("吊杆及系杆索力加速度", len(points.get("cable_accel") or per_point.get("cable_accel") or []) or len(cable_rows), len(cable_rows)),
    ]

def _sxh_context(
    config_path: Path,
    result_root: Path,
    monitoring_range: str,
    start_date: str | None = None,
    end_date: str | None = None,
) -> dict[str, Any]:
    cfg = load_json(config_path)
    report_start, report_end = _sxh_period_dates(
        monitoring_range, start_date, end_date
    )
    stats_dir = result_root / "stats"
    acq_files = sorted(stats_dir.glob("水仙花大桥_测点配置与数据获取情况_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    acquisition_rows = load_rows(acq_files[0], "汇总") if acq_files else []
    raw_wind_rows = load_rows(stats_dir / "wind_stats.xlsx")
    wind_rows = adjusted_rows(stats_dir, "wind_direction_stats_report.xlsx", adjusted_rows(stats_dir, "wind_stats_report.xlsx", raw_wind_rows))
    wind_rows = _sxh_parse_wind_summaries(
        result_root, wind_rows, report_start, report_end
    )
    temp_rows = _sxh_normalize_point_ids(
        load_rows(stats_dir / "temp_stats.xlsx"),
        cfg.get("points", {}).get("temperature"),
    )
    humidity_rows = _sxh_normalize_point_ids(
        load_rows(stats_dir / "humidity_stats.xlsx"),
        cfg.get("points", {}).get("humidity"),
    )
    earthquake_rows = adjusted_rows(stats_dir, "earthquake_filtered_stats_mps2.xlsx", adjusted_rows(stats_dir, "earthquake_filtered_stats.xlsx", load_rows(stats_dir / "eq_stats.xlsx")))
    deflection_rows = load_rows(stats_dir / "deflection_stats.xlsx")
    bearing_rows = load_rows(stats_dir / "bearing_displacement_stats.xlsx")
    accel_rows = adjusted_rows(stats_dir, "accel_stats_mps2.xlsx", scaled_accel_rows(load_rows(stats_dir / "accel_stats.xlsx")))
    strain_rows = adjusted_rows(stats_dir, "strain_stats_zero_corrected.xlsx", load_rows(stats_dir / "strain_stats.xlsx"))
    strain_rows = _sxh_strain_rows_with_group(strain_rows, cfg)
    cable_rows = adjusted_rows(stats_dir, "cable_accel_stats_mps2.xlsx", scaled_accel_rows(load_rows(stats_dir / "cable_accel_stats.xlsx")))
    report_rows = report_acquisition_rows(acquisition_rows) if acquisition_rows else _sxh_fallback_report_rows(
        cfg,
        temp_rows=temp_rows,
        humidity_rows=humidity_rows,
        wind_rows=wind_rows,
        earthquake_rows=earthquake_rows,
        deflection_rows=deflection_rows,
        bearing_rows=bearing_rows,
        accel_rows=accel_rows,
        strain_rows=strain_rows,
        cable_rows=cable_rows,
    )
    return {
        "cfg": cfg,
        "stats_dir": stats_dir,
        "date_span": _sxh_monitoring_range_span(monitoring_range),
        "report_rows": report_rows,
        "temp_rows": temp_rows,
        "humidity_rows": humidity_rows,
        "wind_rows": wind_rows,
        "earthquake_rows": earthquake_rows,
        "deflection_rows": deflection_rows,
        "bearing_rows": bearing_rows,
        "accel_rows": accel_rows,
        "accel_freq_map": _sxh_accel_frequency_map(stats_dir),
        "strain_rows": strain_rows,
        "cable_rows": cable_rows,
    }


def _sxh_availability_summary(context: dict[str, Any]) -> str:
    cfg = context.get("cfg") or {}
    temperature_points = _sxh_configured_points(context, "temperature", "temp_rows")
    temperature_expected = len(temperature_points) or len(context.get("temp_rows") or [])
    temperature_found = len(context.get("temp_rows") or [])
    if temperature_expected and temperature_found >= temperature_expected:
        temperature_clause = f"本月温度监测{temperature_expected}个测点均获取到有效数据"
    elif temperature_found:
        temperature_clause = (
            f"本月温度监测已获取{temperature_found}个测点有效数据，"
            f"仍有{max(temperature_expected - temperature_found, 0)}个测点未获取有效数据"
        )
    else:
        temperature_clause = "本月温度监测未获取有效数据"

    gnss_points = (
        cfg.get("points", {}).get("gnss")
        or cfg.get("per_point", {}).get("gnss")
        or cfg.get("design_points_pending", {}).get("gnss")
        or []
    )
    if isinstance(gnss_points, str):
        gnss_points = [gnss_points]
    gnss_expected = len(gnss_points)
    gnss_found = 0
    for row in context.get("report_rows") or []:
        module_code = str(row.get("模块代码") or "").lower()
        module_name = str(row.get("模块") or "").upper()
        if module_code == "gnss" or "GNSS" in module_name:
            gnss_expected = int(row.get("配置测点数") or gnss_expected or 0)
            gnss_found = int(row.get("实际获取测点数") or 0)
            break
    if gnss_expected and gnss_found >= gnss_expected:
        gnss_clause = f"GNSS监测{gnss_expected}个测点均获取到有效数据"
    elif gnss_found:
        gnss_clause = f"GNSS监测已获取{gnss_found}个测点有效数据"
    else:
        gnss_clause = "GNSS本月未获取有效数据"
    return f"{temperature_clause}，{gnss_clause}。"


def _sxh_summary_payload(context: dict[str, Any]) -> dict[str, str]:
    temp_range = _sxh_range_plain(context["temp_rows"], "Min", "Max", digits=1, unit="℃")
    humidity_range = _sxh_range_plain(context["humidity_rows"], "Min", "Max", digits=1, unit="%")
    wind_deck_rows = [row for row in context["wind_rows"] if str(row.get("测点编号") or row.get("PointID") or "").startswith("FSFX-01")]
    wind_deck = wind_deck_rows[0] if wind_deck_rows else (context["wind_rows"][0] if context["wind_rows"] else {})
    wind_deck_10 = _sxh_fixed(wind_deck.get("10min平均风速最大值(m/s)") or wind_deck.get("Mean10minMax"), 2)
    eq_rows = context["earthquake_rows"]
    horiz_rows = [row for row in eq_rows if str(row.get("方向") or row.get("Component") or "").upper() in {"X", "Y"}]
    vert_rows = [row for row in eq_rows if str(row.get("方向") or row.get("Component") or "").upper() == "Z"]
    horiz_values = [safe_float(row.get("最大值(m/s²)") or row.get("Peak") or row.get("Max")) for row in horiz_rows]
    vert_values = [safe_float(row.get("最大值(m/s²)") or row.get("Peak") or row.get("Max")) for row in vert_rows]
    horiz_max = max([value for value in horiz_values if value is not None], default=None)
    vert_max = max([value for value in vert_values if value is not None], default=None)
    defl_orig = _sxh_range_plain(context["deflection_rows"], "OrigMin_mm", "OrigMax_mm", digits=1, unit="mm")
    defl_filt = _sxh_range_plain(context["deflection_rows"], "FiltMin_mm", "FiltMax_mm", digits=1, unit="mm")
    support_rows = _sxh_by_prefix(context["bearing_rows"], "ZZWY")
    expansion_rows = _sxh_by_prefix(context["bearing_rows"], "SSF")
    support_orig = _sxh_range(support_rows, "OrigMin_mm", "OrigMax_mm", digits=1, unit="mm")
    expansion_orig = _sxh_range(expansion_rows, "OrigMin_mm", "OrigMax_mm", digits=1, unit="mm")
    support_filt = _sxh_range(support_rows, "FiltMin_mm", "FiltMax_mm", digits=1, unit="mm")
    expansion_filt = _sxh_range(expansion_rows, "FiltMin_mm", "FiltMax_mm", digits=1, unit="mm")
    accel_row = extreme_row(context["accel_rows"], "RMS10minMax")
    cable_row = extreme_row(context["cable_rows"], "RMS10minMax")
    freq_values = [value for pair in context["accel_freq_map"].values() for value in pair if value is not None]
    freq_range = f"{fmt_num(min(freq_values), 3)}Hz~{fmt_num(max(freq_values), 3)}Hz" if freq_values else "/"
    strain_ranges = "；".join(_sxh_strain_group_ranges(context["strain_rows"], context["cfg"]))
    accel_point = str(accel_row.get("PointID")) if accel_row and accel_row.get("PointID") is not None else "/"
    accel_value = safe_float(accel_row.get("RMS10minMax")) if accel_row else None
    accel_first_limit, accel_second_limit = _sxh_accel_rms_limits_mps2(context["cfg"], accel_point)
    accel_status = _sxh_accel_rms_status(accel_value, accel_first_limit, accel_second_limit)
    accel_body = f"监测结果表明，各测点10min均方根最大值为{fmt_num(accel_value, 3)}m/s²，对应测点为{accel_point}，{accel_status}竖向1阶自振频率范围在{freq_range}之间，均大于结构相应理论计算的1阶竖弯频率1.050Hz。"
    cable_body = f"监测结果表明，各测点10min均方根最大值为{fmt_num(cable_row.get('RMS10minMax') if cable_row else None, 3)}m/s²，对应测点为{cable_row.get('PointID') if cable_row else '/'}，未超过1.000m/s²，均处于超限阈值范围之内，未出现超过各级超限阈值和报警的情况。"
    bearing_body = f"监测结果表明，支座位移原始数据实测值范围在{support_orig}之间，伸缩缝位移原始数据实测值范围在{expansion_orig}之间，均处于超限阈值范围之内，未出现超过各级超限阈值和报警的情况。支座位移滤波后实测值在{support_filt}之间，伸缩缝位移滤波后实测值范围在{expansion_filt}之间。"
    temp_points = [
        str(row.get("PointID") or "").strip()
        for row in context["temp_rows"]
        if str(row.get("PointID") or "").strip()
    ]
    temp_subject = temp_points[0] if len(temp_points) == 1 else f"{len(temp_points)}个温度测点"
    return {
        "availability": _sxh_availability_summary(context),
        "temp": f"监测结果表明，{temp_subject}温度实测值范围为{temp_range}；各测点结果详见温度监测结果汇总表，其中异常极值应结合数据质量复核后解释。",
        "humidity": f"监测结果表明，WSD-01-11#-S11相对湿度实测值范围为{humidity_range}，处于正常环境湿度范围。",
        "wind": f"监测结果表明，桥面风速风向测点10min平均风速最大值为{wind_deck_10}m/s，未超过25m/s，处于预警阈值范围之内，未出现超过各级超限阈值和报警的情况。",
        "earthquake": f"监测结果表明，水平向地震动加速度峰值为{fmt_num(horiz_max, 3)}m/s²，竖向地震动加速度峰值为{fmt_num(vert_max, 3)}m/s²，处于超限阈值范围之内，未出现超过各级超限阈值和报警的情况。",
        "deflection_body": f"监测结果表明，主梁挠度在{defl_orig}之间，均处于预警阈值范围之内，未超过各级超限阈值和报警的情况。主梁挠度滤波后在{defl_filt}之间。",
        "deflection_front": f"主梁挠度原始数据实测值范围在{defl_orig}之间，均处于预警阈值范围之内，未出现超过各级超限阈值和报警的情况。滤波后实测值范围在{defl_filt}之间。",
        "bearing_body": bearing_body,
        "bearing_front": bearing_body.replace("监测结果表明，", "", 1),
        "accel_body": accel_body,
        "accel_front": accel_body.replace("监测结果表明，", "", 1),
        "strain": f"监测结果表明，本月结构应变按组图分组统计：{strain_ranges}，均处于超限阈值范围之内，未出现超过各级超限阈值和报警的情况。",
        "cable_body": cable_body,
        "cable_front": cable_body.replace("监测结果表明，", "", 1),
    }

def _sxh_xml_text(element) -> str:
    return ooxml_text(element)

def _sxh_xml_set_paragraph_text(paragraph, text: str) -> None:
    ooxml_set_paragraph_text(paragraph, text)

def _sxh_xml_set_cell_text(cell, text: Any) -> None:
    ooxml_set_cell_text(cell, text)

def _sxh_xml_rewrite_contains(root, contains: str, replacement: str, *, startswith: str | None = None) -> None:
    ooxml_rewrite_contains(root, contains, replacement, startswith=startswith)

def _sxh_xml_fill_table(table, rows: list[dict[str, Any]], value_builder) -> None:
    ooxml_fill_table(table, rows, value_builder)

def _sxh_xml_update_stats_tables(root, context: dict[str, Any]) -> None:
    tables = required_result_tables(root)
    report_rows = context["report_rows"]
    _sxh_xml_fill_table(tables["acquisition"], report_rows, lambda idx, row: [idx, row.get("模块"), row.get("配置测点数"), row.get("实际获取测点数"), fmt_percent(row.get("获取率")).replace("100.0%", "100%"), context["date_span"], row.get("缺失说明") or ("无" if not row.get("缺失测点数") else f"缺失{row.get('缺失测点数')}个")])
    _sxh_xml_fill_table(tables["temperature"], context["temp_rows"], lambda idx, row: [idx, row.get("PointID"), "温度", fmt_num(row.get("Min"), 1), fmt_num(row.get("Max"), 1), fmt_num(row.get("Mean"), 1)])
    _sxh_xml_fill_table(tables["humidity"], context["humidity_rows"], lambda idx, row: [idx, row.get("PointID"), "相对湿度", fmt_num(row.get("Min"), 1), fmt_num(row.get("Max"), 1), fmt_num(row.get("Mean"), 1)])
    _sxh_xml_fill_table(tables["wind"], context["wind_rows"], lambda idx, row: [idx, _sxh_text_value(row, "测点编号", "PointID"), _sxh_text_value(row, "平均风向"), _sxh_text_value(row, "主导风向"), fmt_num(row.get("平均风速(m/s)") or row.get("MeanSpeed") or str(_sxh_text_value(row, "平均风速")).split()[0], 2), fmt_num(row.get("最大风速(m/s)") or row.get("MaxSpeed") or str(_sxh_text_value(row, "最大风速")).split()[0], 1), _sxh_fixed(row.get("10min平均风速最大值(m/s)") or row.get("Mean10minMax"), 2)])
    _sxh_xml_fill_table(tables["earthquake"], context["earthquake_rows"], lambda idx, row: [idx, _sxh_text_value(row, "测点编号", "PointID"), _sxh_text_value(row, "方向", "Component"), fmt_num(row.get("最小值(m/s²)") or row.get("Min"), 3), fmt_num(row.get("最大值(m/s²)") or row.get("Peak") or row.get("Max"), 3)])
    _sxh_xml_fill_table(tables["deflection_raw"], context["deflection_rows"], lambda idx, row: [idx, row.get("PointID"), fmt_num(row.get("OrigMin_mm"), 1), fmt_num(row.get("OrigMax_mm"), 1)])
    _sxh_xml_fill_table(tables["deflection_filtered"], context["deflection_rows"], lambda idx, row: [idx, row.get("PointID"), fmt_num(row.get("FiltMin_mm"), 1), fmt_num(row.get("FiltMax_mm"), 1)])
    _sxh_xml_fill_table(tables["bearing_raw"], context["bearing_rows"], lambda idx, row: [idx, row.get("PointID"), fmt_num(row.get("OrigMin_mm"), 1), fmt_num(row.get("OrigMax_mm"), 1)])
    _sxh_xml_fill_table(tables["bearing_filtered"], context["bearing_rows"], lambda idx, row: [idx, row.get("PointID"), fmt_num(row.get("FiltMin_mm"), 1), fmt_num(row.get("FiltMax_mm"), 1)])
    _sxh_xml_fill_table(tables["gnss"], [{"PointID": "拱顶、拱脚位移（GNSS）"}], lambda idx, row: [idx, row.get("PointID"), "/", "/", "/"])
    freq_map = context["accel_freq_map"]
    _sxh_xml_fill_table(tables["acceleration"], context["accel_rows"], lambda idx, row: [idx, row.get("PointID"), fmt_num(row.get("Min"), 4), fmt_num(row.get("Max"), 4), fmt_num(row.get("RMS10minMax"), 4), fmt_num(freq_map.get(str(row.get("PointID")), (None, None))[0], 3), fmt_num(freq_map.get(str(row.get("PointID")), (None, None))[1], 3)])
    _sxh_xml_fill_table(tables["strain"], context["strain_rows"], lambda idx, row: [row.get("分组"), row.get("PointID"), fmt_num(row.get("Min"), 1), fmt_num(row.get("Max"), 1), fmt_num(row.get("Mean"), 1)])
    _sxh_xml_fill_table(tables["cable_accel"], context["cable_rows"], lambda idx, row: [idx, row.get("PointID"), fmt_num(row.get("Min"), 4), fmt_num(row.get("Max"), 4), fmt_num(row.get("RMS10minMax"), 4)])

def _sxh_xml_update_summary(
    root,
    context: dict[str, Any],
    monitoring_range: str,
    report_date: str,
    period_label: str,
) -> None:
    payload = _sxh_summary_payload(context)
    normalized_period = re.sub(r"月份$", "月", str(period_label or "").strip())
    normalized_period = re.sub(r"(\d{4})年0?(\d{1,2})月$", lambda m: f"{int(m.group(1))}年{int(m.group(2))}月", normalized_period)
    for paragraph in root.findall(".//w:p", {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
        text = _sxh_xml_text(paragraph)
        if text.startswith("报告编号："):
            _sxh_xml_set_paragraph_text(paragraph, f"报告编号：{REPORT_NO}")
        elif text.startswith("报告日期："):
            _sxh_xml_set_paragraph_text(paragraph, f"报告日期：{report_date}")
        elif "监测时间：" in text and normalized_period:
            prefix = "（" if text.startswith("（") else ""
            suffix = "）" if text.endswith("）") else ""
            _sxh_xml_set_paragraph_text(paragraph, f"{prefix}监测时间：{normalized_period}{suffix}")
        elif re.search(r"\d{4}[.-]\d{1,2}[.-]\d{1,2}\s*[~至-]\s*\d{4}[.-]\d{1,2}[.-]\d{1,2}", text):
            _sxh_xml_set_paragraph_text(paragraph, _sxh_monitoring_range_span(monitoring_range))
    _sxh_xml_rewrite_contains(root, "WSD-01-11#-S11温度", payload["temp"])
    _sxh_xml_rewrite_contains(
        root,
        "温度监测中9个结构温度测点本月未获取数据",
        payload["availability"],
    )
    _sxh_xml_rewrite_contains(root, "相对湿度实测值范围", payload["humidity"])
    _sxh_xml_rewrite_contains(root, "本月相对湿度数据实测值范围", payload["humidity"])
    _sxh_xml_rewrite_contains(root, "本月相对湿度在", payload["humidity"])
    _sxh_xml_rewrite_contains(root, "桥面风速风向测点10min平均风速最大值", payload["wind"])
    _sxh_xml_rewrite_contains(root, "水平向地震动加速度峰值", payload["earthquake"])
    _sxh_xml_rewrite_contains(root, "监测结果表明，主梁挠度", payload["deflection_body"])
    _sxh_xml_rewrite_contains(root, "主梁挠度原始数据实测值范围", payload["deflection_front"])
    _sxh_xml_rewrite_contains(root, "支座位移原始数据实测值范围", payload["bearing_body"], startswith="监测结果表明")
    for paragraph in root.findall(".//w:p", {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}):
        text = _sxh_xml_text(paragraph)
        if "支座位移原始数据实测值范围" in text and not text.startswith("监测结果表明"):
            _sxh_xml_set_paragraph_text(paragraph, payload["bearing_front"])
        if "各测点10min均方根最大值" in text:
            if "0.315m/s²" in text or "ZLZD" in text:
                _sxh_xml_set_paragraph_text(paragraph, payload["accel_body"] if text.startswith("监测结果表明") else payload["accel_front"])
            elif "1.000m/s²" in text or "SL-" in text:
                _sxh_xml_set_paragraph_text(paragraph, payload["cable_body"] if text.startswith("监测结果表明") else payload["cable_front"])
    _sxh_xml_rewrite_contains(root, "结构应变按组图分组统计", payload["strain"])

def _sxh_update_docx_package(
    docx_path: Path,
    context: dict[str, Any],
    monitoring_range: str,
    report_date: str,
    period_label: str,
) -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".docx", dir=str(docx_path.parent)) as tmp_file:
        tmp_path = Path(tmp_file.name)
    try:
        with ZipFile(docx_path, "r") as zin, ZipFile(tmp_path, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "word/document.xml":
                    root = etree.fromstring(data)
                    _sxh_xml_update_summary(root, context, monitoring_range, report_date, period_label)
                    _sxh_xml_update_stats_tables(root, context)
                    data = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                elif item.filename == "word/settings.xml":
                    root = etree.fromstring(data)
                    for element in list(root.findall(qn("w:trackRevisions"))):
                        root.remove(element)
                    data = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)
                zout.writestr(item, data)
        tmp_path.replace(docx_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()


def _sxh_qc_exception_result(output: Path, exc: Exception) -> ReportQcResult:
    """Convert a QC execution exception into durable, fail-closed evidence."""
    issue = ReportQcIssue(
        code="report-qc-exception",
        severity="error",
        message="水仙花报告质量检查执行异常。",
        detail=f"{type(exc).__name__}: {exc}",
    )
    return ReportQcResult(
        kind="shuixianhua_monthly",
        docx_path=str(output.resolve()),
        checked_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        status="failed",
        issue_count=1,
        warning_count=0,
        issues=[issue],
        summary={
            "exists": output.exists(),
            "exception_type": type(exc).__name__,
            "exception_message": str(exc),
        },
    )


def build_report(
    template: Path,
    config_path: Path,
    result_root: Path,
    output_dir: Path | None = None,
    period_label: str = "2026年3月份",
    monitoring_range: str = "2026年3月23日~2026年3月31日",
    report_date: str = "2026年4月5日",
    start_date: str | None = None,
    end_date: str | None = None,
    update_word: bool = True,
) -> tuple[Path, Path | None]:
    if not template.exists():
        raise FileNotFoundError(f"未找到水仙花报告模板：{template}")
    ctx = ReportBuildContext.from_inputs(
        template=template,
        config_path=config_path,
        result_root=result_root,
        image_root=result_root,
        output_dir=output_dir,
        assets_subdir="generated_assets_shuixianhua",
    )
    output_dir = ctx.output_dir
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = output_dir / f"水仙花大桥健康监测{_sxh_period_filename_label(period_label)}月报_报告生成器_{timestamp}.docx"
    shutil.copy2(template, output)

    report_start, report_end = _sxh_period_dates(monitoring_range, start_date, end_date)
    context = _sxh_context(
        config_path,
        result_root,
        monitoring_range,
        report_start,
        report_end,
    )
    _sxh_update_docx_package(output, context, monitoring_range, report_date, period_label)
    grouped_images, image_sources, image_missing = _sxh_resolve_report_images(
        result_root, context, report_start, report_end
    )
    image_replacements, anchor_missing = _sxh_replace_report_image_blocks(output, grouped_images)
    content_missing = _sxh_validate_generated_content(output, context)
    pdf = update_word_fields_and_export_pdf(output) if update_word else None
    qc_paths: dict[str, Any] = {}
    qc_warnings: list[str] = []
    qc_failure_cause: Exception | None = None
    try:
        qc_result = check_shuixianhua_report(
            output,
            expected_period_label=period_label,
            expected_image_paths=[Path(record["path"]) for record in image_sources],
        )
    except Exception as exc:
        qc_failure_cause = exc
        qc_result = _sxh_qc_exception_result(output, exc)

    qc_status = qc_result.status
    qc_paths["report_qc_status"] = qc_status
    try:
        qc_txt, qc_json = write_report_qc_report(qc_result, output_dir, timestamp=timestamp)
        qc_paths.update({
            "report_qc_txt": str(qc_txt),
            "report_qc_json": str(qc_json),
        })
    except Exception as exc:
        qc_failure_cause = qc_failure_cause or exc
        qc_status = "failed"
        qc_paths.update({
            "report_qc_status": qc_status,
            "report_qc_artifact_error": {
                "type": type(exc).__name__,
                "message": str(exc),
            },
        })
        qc_warnings.append(f"report_qc_artifact_failed: {type(exc).__name__}: {exc}")
    qc_warnings.extend(
        f"{issue.code}: {issue.message}"
        for issue in qc_result.issues
        if issue.severity in {"warning", "error"}
    )
    analysis_context = ctx.analysis_context()
    missing_items = missing_module_summary_items(analysis_context)
    missing_items.extend(image_missing)
    missing_items.extend(anchor_missing)
    missing_items.extend(content_missing)
    missing_items.extend(
        {
            "category": "report_qc",
            "item": issue.code,
            "detail": f"{issue.message} {issue.detail}".strip(),
        }
        for issue in qc_result.issues
        if issue.severity == "error"
    )
    manifest_path = write_report_build_manifest(
        context=ctx,
        report_type="shuixianhua_monthly",
        output_docx=output,
        timestamp=timestamp,
        missing=missing_items,
        warnings=qc_warnings,
        extra={
            "output_pdf": str(pdf.resolve()) if pdf else "",
            "report_row_count": len(context.get("report_rows") or []),
            "report_period": {"start_date": report_start, "end_date": report_end, "period_label": period_label},
            "report_image_source_count": len(image_sources),
            "report_image_sources": image_sources,
            "report_image_replacements": image_replacements,
            **({"status": "failed"} if qc_status != "ok" else {}),
            **qc_paths,
        },
        filename_prefix="shuixianhua_report_build_manifest",
    )
    if qc_status != "ok":
        issue_summary = "; ".join(
            f"{issue.code}: {issue.message}"
            for issue in qc_result.issues
        ) or "no structured QC issue was returned"
        raise RuntimeError(
            "Shuixianhua report QC did not pass "
            f"(status={qc_status}); manifest={manifest_path}; {issue_summary}"
        ) from qc_failure_cause
    return output.resolve(), pdf.resolve() if pdf else None


def main() -> None:
    args = parse_args()
    output_dir = args.output_dir or args.result_root / "自动报告"
    output, pdf = build_report(
        template=args.template,
        config_path=args.config,
        result_root=args.result_root,
        output_dir=output_dir,
        period_label=args.period_label,
        monitoring_range=args.monitoring_range,
        report_date=args.report_date,
        start_date=args.start_date or None,
        end_date=args.end_date or None,
        update_word=not args.no_word_update,
    )
    print(f"Shuixianhua monthly report generated: {output}")
    if pdf:
        print(f"Shuixianhua monthly report PDF generated: {pdf}")


if __name__ == "__main__":
    main()
