from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from scripts.high_memory_recovery import (
    _infer_artifact_role,
    _spectrum_output_contract,
    build_baseline_evidence,
    compose_recovery_manifest,
    merge_cable_accel_stats,
    parse_recovery_started_at,
    prepare_recovery_requests,
)


POINTS = [f"SLCGQ-{index:02d}" for index in range(1, 16)]
BASE_MODULES = [
    "temperature",
    "humidity",
    "rainfall",
    "gnss",
    "wind",
    "earthquake",
    "deflection",
    "bearing_displacement",
    "tilt",
    "crack",
    "strain",
    "acceleration",
]


@pytest.mark.parametrize(
    ("relative_path", "expected"),
    [
        ("加速度_RMS/A1_RMS10.jpg", "rms10min"),
        ("频谱峰值曲线_加速度/SpecFreq_A1.jpg", "spectrum"),
        ("加速度箱线图/A1_boxplot.jpg", "boxplot"),
        ("风速风向结果/风玫瑰/W1_windrose_summary.txt", "wind_rose"),
        ("风速风向结果/风速10min/W1_speed10min.jpg", "wind_speed10min"),
        ("频次分布_湿度/HumidityFreq_H1.jpg", "frequency_distribution"),
        ("时程曲线_挠度_原始/Defl_A1_Orig.jpg", "raw"),
        ("时程曲线_挠度_滤波/Defl_A1_Filt.jpg", "filtered"),
        ("时程曲线_温度/T1.jpg", "time_history"),
    ],
)
def test_rebuilt_baseline_artifact_roles_match_normal_collector(
    relative_path: str, expected: str
) -> None:
    assert _infer_artifact_role(Path(relative_path)) == expected


@pytest.mark.parametrize(
    "value",
    [
        "2026-07-15T23:17:10.069879+00:00",
        "2026-07-15T23:17:10.069879Z",
        "'2026-07-15T23:17:10.0698799+00:00'",
        '"2026-07-15T23:17:10.069879+00:00"',
    ],
)
def test_parse_recovery_started_at_accepts_native_cli_compatibility_forms(
    value: str,
) -> None:
    parsed = parse_recovery_started_at(value)
    assert parsed.isoformat() == "2026-07-15T23:17:10.069879+00:00"


@pytest.mark.parametrize(
    "value",
    ["", "2026-07-15T23:17:10.069879", "'2026-07-15T23:17:10+00:00\""],
)
def test_parse_recovery_started_at_rejects_empty_naive_or_mismatched_values(
    value: str,
) -> None:
    with pytest.raises(ValueError):
        parse_recovery_started_at(value)


def test_current_jiulongjiang_spectrum_contract_has_46_formal_stubs() -> None:
    config_path = Path(__file__).resolve().parents[1] / "config" / "jiulongjiang_config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    acceleration = _spectrum_output_contract(
        config, "accel_spectrum", "2026-05-01", "2026-05-31"
    )
    cable = _spectrum_output_contract(
        config, "cable_accel_spectrum", "2026-05-01", "2026-05-31"
    )
    assert len(acceleration["configured_points"]) == 16
    assert len(acceleration["expected_formal_figure_stubs"]) == 16
    assert acceleration["groups"] == []
    assert len(cable["configured_points"]) == 15
    assert len(cable["valid_force_points"]) == 15
    assert len(cable["expected_formal_figure_stubs"]) == 30
    assert cable["groups"] == []
    assert cable["cable_force_engineering_valid"] is False
    assert cable["cable_force_engineering_status"] == "placeholder_parameters"
    assert {
        item["parameter_status"] for item in cable["force_parameter_evidence"]
    } == {"placeholder_parameters"}
    assert (
        len(acceleration["expected_formal_figure_stubs"])
        + len(cable["expected_formal_figure_stubs"])
    ) == 46


def _config_contract() -> dict:
    return {
        "vendor": "jiulongjiang",
        "data_adapter": {"time_series": {"source_mode": "mat_only"}},
        "plot_common": {
            "gap_mode": "connect",
            "dynamic_raw_modules": {
                "acceleration": {"sampling_mode": "full", "gap_mode": "connect"},
                "cable_accel": {"sampling_mode": "full", "gap_mode": "connect"},
            },
        },
    }


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _canonical_json_sha(payload: dict) -> str:
    raw = (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    return hashlib.sha256(raw).hexdigest().upper()


def _write_stats(path: Path, point: str, value: float = 1.0) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["PointID", "Min", "Max", "Mean", "RMS10minMax", "RMSStartTime"])
    sheet.append([point, -value, value, 0.0, value / 2, "2026-05-01 00:00:00"])
    workbook.save(path)
    workbook.close()


def _artifact(path: Path, kind: str = "figure") -> dict:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix == ".xlsx":
        _write_stats(path, "placeholder")
    else:
        path.write_bytes(f"artifact:{path.name}".encode())
    return {
        "kind": kind,
        "role": "stats" if kind == "stats" else "time_history",
        "path": str(path),
        "exists": True,
        "bytes": 1,
        "sha256": "stale",
    }


def _plot_provenance(stub: str, point_id: str = "") -> dict:
    series = {
        "sampling_mode": "full",
        "render_mode": "line",
        "input_count": 10,
        "finite_count": 10,
        "plotted_finite_count": 10,
        "reduction_applied": False,
        "source": {
            "source_sample_count": 10,
            "finite_source_sample_count": 10,
            "completeness_scope": "required_export_contribution",
            "internal_gap_coverage_assessed": True,
            "calendar_day_count_requested": 31,
            "complete_day_count": 31,
            "incomplete_day_count": 0,
            "incomplete_days": [],
            "missing_required_sources": [],
        },
    }
    if point_id:
        series["point_id"] = point_id
    return {"file_stub": stub, "series": series}


def _module_record(root: Path, key: str, suffix: str = "", point_id: str = "") -> dict:
    stats = root / "stats" / f"{key}{suffix}.xlsx"
    _write_stats(stats, key)
    figure = root / "figures" / key / f"{key}{suffix}.jpg"
    figure.parent.mkdir(parents=True, exist_ok=True)
    figure.write_bytes(f"{key}:{suffix}".encode())
    provenance = root / "figures" / key / f"{key}{suffix}.plot.json"
    _write_json(provenance, _plot_provenance(f"{key}{suffix}", point_id))
    return {
        "key": key,
        "label": key,
        "status": "ok",
        "stats_path": str(stats),
        "stats_exists": True,
        "artifacts": [
            {"kind": "stats", "role": "stats", "path": str(stats)},
            {"kind": "figure", "role": "time_history", "path": str(figure)},
            {"kind": "plot_provenance", "role": "plot_provenance", "path": str(provenance)},
        ],
    }


def _write_spectrum_stats(path: Path, points: list[str], *, cable: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for point in points:
        sheet = workbook.create_sheet(point)
        headers = ["Date", "Freq_1.000Hz", "Amp_1.000Hz"]
        if cable:
            headers.append("CableForce_kN")
        sheet.append(headers)
        sheet.append(["2026-05-01", 1.0, 2.0, *([1000.0] if cable else [])])
    workbook.save(path)
    workbook.close()


def _spectrum_record(
    root: Path, key: str, points: list[str], *, cable: bool = False,
) -> dict:
    stats = root / "stats" / f"{key}.xlsx"
    _write_spectrum_stats(stats, points, cable=cable)
    artifacts = [{"kind": "stats", "role": "stats", "path": str(stats)}]
    stubs = [f"SpecFreq_{point}_20260501_20260531" for point in points]
    if cable:
        stubs.extend(f"CableForce_{point}_20260501_20260531" for point in points)
    for stub in stubs:
        figure = root / "figures" / key / f"{stub}.jpg"
        figure.parent.mkdir(parents=True, exist_ok=True)
        figure.write_bytes(f"figure:{stub}".encode())
        artifacts.append({"kind": "figure", "role": "time_history", "path": str(figure)})
    return {
        "key": key,
        "label": key,
        "status": "ok",
        "stats_path": str(stats),
        "stats_exists": True,
        "artifacts": artifacts,
    }


def test_prepare_requests_is_dry_run_rc_bounded_and_disables_cable_groups(tmp_path: Path) -> None:
    base = tmp_path / "base_request.json"
    config = {
        **_config_contract(),
        "points": {"cable_force": POINTS, "acceleration": ["A1"]},
        "per_point": {"cable_accel": {point: {"thresholds": []} for point in POINTS}},
        "plot_common": {
            **_config_contract()["plot_common"],
            "dynamic_group_sampling_mode": "full",
        },
        "preprocessing": {
            "unzip": {
                "source_root": r"E:\九龙江数据\2026年5月",
                "output_root": r"F:\Guanbing_v1.8.1-rc1\validation\jlj",
            }
        },
    }
    _write_json(
        base,
        {
            "project_root": r"F:\Guanbing_v1.8.1-rc1\app",
            "data_root": r"F:\Guanbing_v1.8.1-rc1\validation\jlj",
            "start_date": "2026-05-01",
            "end_date": "2026-05-31",
            "config_path": r"F:\Guanbing_v1.8.1-rc1\old\config.json",
            "options": {
                "doTemp": True,
                "doCableAccel": True,
                "doAccelSpectrum": True,
                "doCableAccelSpectrum": True,
                "doLegacyTruth": "yes",
                "precheck_zip_count": 31,
                "input_mode": "mat_only",
            },
            "config": config,
        },
    )

    plan = prepare_recovery_requests(
        base,
        tmp_path / "bundle",
        rc_root=r"F:\Guanbing_v1.8.1-rc1",
        remote_bundle_root=r"F:\Guanbing_v1.8.1-rc1\run_logs\recovery",
    )

    assert plan["launch"] is False
    assert len(plan["jobs"]) == 17
    assert plan["cable_point_order"] == POINTS
    assert plan["cable_group_plot_policy"]["value"] == "off"
    assert plan["dynamic_plot_contract"]["cable_accel"] == {
        "sampling_mode": "full", "gap_mode": "connect"
    }
    assert plan["cable_group_recovery"]["required"] is False
    expectations = json.loads(
        (tmp_path / "bundle" / "recovery_plot_expectations.json").read_text(encoding="utf-8")
    )
    assert expectations["cable_group"]["expected_plot_provenance_count"] == 0
    assert [item["expected_plot_provenance_count"] for item in expectations["cable_points"]] \
        == [1] * 15
    assert expectations["accel_spectrum"]["configured_points"] == ["A1"]
    assert expectations["accel_spectrum"]["expected_formal_figure_stubs"] == [
        "SpecFreq_A1_20260501_20260531"
    ]
    assert expectations["cable_accel_spectrum"]["configured_points"] == POINTS
    assert expectations["cable_accel_spectrum"]["expected_plot_provenance_count"] == 0
    group_evidence = json.loads(
        (tmp_path / "bundle" / "cable_group_evidence.json").read_text(encoding="utf-8")
    )
    assert group_evidence["status"] == "ok"
    assert group_evidence["reason_code"] == "no_effective_cable_accel_groups"
    for index, point in enumerate(POINTS, start=1):
        request_path = tmp_path / "bundle" / "requests" / f"cable_accel_{index:02d}.json"
        request = json.loads(request_path.read_text(encoding="utf-8"))
        assert request["config"]["points"]["cable_accel"] == [point]
        assert request["config"]["points"]["cable_force"] == [point]
        assert request["config"]["plot_common"]["dynamic_group_sampling_mode"] == "off"
        assert "source_root" not in request["config"]["preprocessing"]["unzip"]
        assert request["options"]["doCableAccel"] is True
        assert request["options"]["doTemp"] is False
        assert request["options"]["doLegacyTruth"] is False
        assert request["options"]["precheck_zip_count"] is False
        assert request["options"]["input_mode"] == "mat_only"
        assert request["enabled_modules"] == ["cable_accel"]
        assert request["async_run_id"].startswith("high-memory-recovery-")
        config_bytes = (tmp_path / "bundle" / "configs" / f"cable_accel_{index:02d}.json").read_bytes()
        assert request["config_sha256"] == hashlib.sha256(config_bytes).hexdigest().upper()

    spectrum = json.loads(
        (tmp_path / "bundle" / "requests" / "accel_spectrum.json").read_text(encoding="utf-8")
    )
    assert spectrum["options"]["doAccelSpectrum"] is True
    assert spectrum["options"]["doCableAccel"] is False


def test_prepare_requests_rejects_data_root_outside_rc(tmp_path: Path) -> None:
    base = tmp_path / "bad.json"
    _write_json(
        base,
            {
                "project_root": r"F:\Guanbing_v1.8.1-rc1\app",
                "data_root": r"F:\Guanbing\production",
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "options": {},
            "config": {
                **_config_contract(),
                "points": {"cable_force": POINTS, "acceleration": ["A1"]},
            },
        },
    )
    with pytest.raises(ValueError, match="data_root"):
        prepare_recovery_requests(
            base,
            tmp_path / "bundle",
            rc_root=r"F:\Guanbing_v1.8.1-rc1",
            remote_bundle_root=r"F:\Guanbing_v1.8.1-rc1\run_logs\recovery",
        )


def test_prepare_requests_detects_fallback_groups_and_rejects_non_full_contract(
    tmp_path: Path,
) -> None:
    grouped_config = {
        **_config_contract(),
        "points": {"cable_force": POINTS, "acceleration": ["A1"]},
        "groups": {"cable_force": {"ALL": POINTS}},
        "per_point": {
            "cable_accel": {
                point: {"rho": 1.0, "L": 10.0} for point in POINTS
            }
        },
    }
    base = tmp_path / "grouped.json"
    _write_json(base, {
        "project_root": r"F:\Guanbing_v1.8.1-rc1\app",
        "data_root": r"F:\Guanbing_v1.8.1-rc1\validation\jlj",
        "start_date": "2026-05-01",
        "end_date": "2026-05-31",
        "options": {"input_mode": "mat_only"},
        "config": grouped_config,
    })
    plan = prepare_recovery_requests(
        base,
        tmp_path / "grouped_bundle",
        rc_root=r"F:\Guanbing_v1.8.1-rc1",
        remote_bundle_root=r"F:\Guanbing_v1.8.1-rc1\run_logs\grouped_recovery",
    )
    assert plan["cable_group_recovery"]["required"] is True
    assert plan["cable_group_recovery"]["resolved_groups"] == {"ALL": POINTS}
    expectations = json.loads(
        (tmp_path / "grouped_bundle" / "recovery_plot_expectations.json").read_text(
            encoding="utf-8"
        )
    )
    assert expectations["cable_group"]["operator_action_required"] is True
    assert expectations["cable_group"]["expected_plot_provenance_count"] is None
    assert expectations["cable_accel_spectrum"]["expected_group_stubs"] == [
        "CableForce_ALL_20260501_20260531"
    ]
    group_evidence = json.loads(
        (tmp_path / "grouped_bundle" / "cable_group_evidence.json").read_text(
            encoding="utf-8"
        )
    )
    assert group_evidence["status"] == "pending"
    assert group_evidence["resolved_groups"] == {"ALL": POINTS}

    broken = json.loads(base.read_text(encoding="utf-8"))
    broken["config"]["plot_common"]["dynamic_raw_modules"]["cable_accel"][
        "sampling_mode"
    ] = "capped"
    broken_path = tmp_path / "broken_plot_contract.json"
    _write_json(broken_path, broken)
    with pytest.raises(ValueError, match=r"full\+connect"):
        prepare_recovery_requests(
            broken_path,
            tmp_path / "broken_bundle",
            rc_root=r"F:\Guanbing_v1.8.1-rc1",
            remote_bundle_root=r"F:\Guanbing_v1.8.1-rc1\run_logs\broken_recovery",
        )


def test_merge_cable_stats_enforces_order_structure_and_receipt(tmp_path: Path) -> None:
    root = tmp_path / "rc"
    inputs = []
    for index, point in enumerate(POINTS, start=1):
        source = root / "snapshots" / f"{index:02d}" / "cable_accel_stats.xlsx"
        _write_stats(source, point, float(index))
        inputs.append(source)

    output = root / "stats" / "cable_accel_stats.xlsx"
    result = merge_cable_accel_stats(POINTS, inputs, output, allowed_root=root)

    workbook = load_workbook(output, read_only=True, data_only=False)
    sheet = workbook.active
    assert sheet.max_row == 16
    assert [sheet.cell(row, 1).value for row in range(2, 17)] == POINTS
    workbook.close()
    receipt = json.loads(Path(result["receipt_path"]).read_text(encoding="utf-8"))
    assert receipt["ordered_points"] == POINTS
    assert [item["point_id"] for item in receipt["inputs"]] == POINTS
    assert receipt["output"]["sha256"] == hashlib.sha256(output.read_bytes()).hexdigest().upper()


def test_merge_cable_stats_rejects_point_or_column_mismatch(tmp_path: Path) -> None:
    root = tmp_path / "rc"
    inputs = []
    for index, point in enumerate(POINTS, start=1):
        source = root / "snapshots" / f"{index:02d}.xlsx"
        _write_stats(source, "WRONG" if index == 8 else point)
        inputs.append(source)
    with pytest.raises(ValueError, match="point mismatch"):
        merge_cable_accel_stats(
            POINTS, inputs, root / "stats" / "merged.xlsx", allowed_root=root
        )


def test_compose_manifest_rehashes_and_rejects_failed_or_missing_evidence(tmp_path: Path) -> None:
    root = tmp_path / "rc"
    baseline_records = [_module_record(root, key) for key in BASE_MODULES]
    semantic_paths = {
        "wind_summary": root / "风速风向结果" / "风玫瑰" / "W1_windrose_summary.txt",
        "wind_rose": root / "风速风向结果" / "风玫瑰" / "W1_windrose.jpg",
        "wind_speed10min": root / "风速风向结果" / "风速10min" / "W1_speed10min.jpg",
        "rms10min": root / "加速度_RMS" / "A1_RMS10.jpg",
        "raw": root / "时程曲线_挠度_原始" / "Defl_A1_Orig.jpg",
        "filtered": root / "时程曲线_挠度_滤波" / "Defl_A1_Filt.jpg",
    }
    module_for_role = {
        "wind_summary": "wind",
        "wind_rose": "wind",
        "wind_speed10min": "wind",
        "rms10min": "acceleration",
        "raw": "deflection",
        "filtered": "deflection",
    }
    records_by_key = {record["key"]: record for record in baseline_records}
    for role_name, path in semantic_paths.items():
        kind = "summary" if role_name == "wind_summary" else "figure"
        records_by_key[module_for_role[role_name]]["artifacts"].append(
            _artifact(path, kind=kind)
        )
    config_path = root / "config" / "baseline.json"
    _write_json(config_path, {
        **_config_contract(),
        "points": {"cable_force": POINTS, "acceleration": ["A1"]},
        "per_point": {
            "cable_accel": {
                point: {"rho": 1.0, "L": 10.0} for point in POINTS
            }
        },
    })
    run_request = root / "run_logs" / "run_request.json"
    options = {field: False for field in (
        "doTemp", "doHumidity", "doRainfall", "doGNSS", "doWind", "doEq",
        "doDeflect", "doBearingDisplacement", "doTilt", "doCrack", "doStrain",
        "doAccel", "doCableAccel", "doAccelSpectrum", "doCableAccelSpectrum",
    )}
    for field in options:
        options[field] = True
    _write_json(run_request, {
        "data_root": str(root),
        "start_date": "2026-05-01",
        "end_date": "2026-05-31",
        "config_path": str(config_path),
        "config_sha256": hashlib.sha256(config_path.read_bytes()).hexdigest().upper(),
        "options": options,
    })
    inventory = root / "run_logs" / "baseline_inventory.json"
    inventory_payload = {"modules": [
        {
            "key": record["key"],
            "stats_path": record["stats_path"],
            "artifact_paths": [item["path"] for item in record["artifacts"] if item["kind"] != "stats"],
            "expected_plot_provenance_count": 1,
            "expected_plot_stubs": [record["key"]],
        }
        for record in baseline_records
    ]}
    _write_json(inventory, inventory_payload)
    bad_inventory = root / "run_logs" / "bad_baseline_inventory.json"
    bad_inventory_payload = json.loads(json.dumps(inventory_payload))
    bad_inventory_payload["modules"][0]["expected_plot_stubs"] = ["wrong_stub"]
    _write_json(bad_inventory, bad_inventory_payload)
    with pytest.raises(ValueError, match="plot stubs differ"):
        build_baseline_evidence(
            run_request,
            bad_inventory,
            root / "run_logs" / "bad_baseline.json",
            expected_modules=BASE_MODULES,
            allowed_root=root,
        )
    baseline = root / "run_logs" / "baseline.json"
    build_baseline_evidence(
        run_request,
        inventory,
        baseline,
        expected_modules=BASE_MODULES,
        allowed_root=root,
    )
    rebuilt_baseline = json.loads(baseline.read_text(encoding="utf-8"))
    assert rebuilt_baseline["bridge_profile"]["bridge_id"] == "jiulongjiang"
    rebuilt_roles = {
        Path(artifact["path"]).name: artifact["role"]
        for record in rebuilt_baseline["module_results"]
        for artifact in record["artifacts"]
    }
    assert rebuilt_roles[semantic_paths["wind_summary"].name] == "wind_rose"
    for role_name in ("wind_rose", "wind_speed10min", "rms10min", "raw", "filtered"):
        assert rebuilt_roles[semantic_paths[role_name].name] == role_name

    single_stats = []
    cable_manifests = []
    for index, point in enumerate(POINTS, start=1):
        stats = root / "snapshots" / f"{index:02d}" / "cable_accel_stats.xlsx"
        _write_stats(stats, point, float(index))
        single_stats.append(stats)
        record = _module_record(root, "cable_accel", f"_{point}", point)
        record["stats_path"] = str(stats)
        record["artifacts"][0] = {"kind": "stats", "role": "stats", "path": str(stats)}
        if index == 2:
            stale_neighbor = root / "figures" / "cable_accel" / "CableAccel_SLCGQ-01_stale.jpg"
            stale_neighbor.parent.mkdir(parents=True, exist_ok=True)
            stale_neighbor.write_bytes(b"stale neighboring point")
            record["artifacts"].append(
                {"kind": "figure", "role": "time_history", "path": str(stale_neighbor)}
            )
        point_config = root / "configs" / f"cable_{index:02d}.json"
        _write_json(point_config, {
            **_config_contract(),
            "points": {"cable_accel": [point], "cable_force": [point]},
            "plot_common": {
                **_config_contract()["plot_common"],
                "dynamic_group_sampling_mode": "off",
            },
        })
        manifest = root / "run_logs" / f"cable_{index:02d}.json"
        _write_json(
            manifest,
            {
                "status": "ok",
                "run_request": {
                    "data_root": str(root),
                    "start_date": "2026-05-01",
                    "end_date": "2026-05-31",
                    "enabled_modules": ["cable_accel"],
                    "config_path": str(point_config),
                    "config_sha256": hashlib.sha256(point_config.read_bytes()).hexdigest().upper(),
                },
                "module_results": [record],
            },
        )
        cable_manifests.append(manifest)

    merged = root / "stats" / "cable_accel_stats.xlsx"
    merge_result = merge_cable_accel_stats(POINTS, single_stats, merged, allowed_root=root)
    accel_manifest = root / "run_logs" / "accel_spectrum.json"
    cable_spectrum_manifest = root / "run_logs" / "cable_spectrum.json"
    accel_config = root / "configs" / "accel_spectrum.json"
    cable_spectrum_config = root / "configs" / "cable_spectrum.json"
    _write_json(accel_config, {
        **_config_contract(),
        "points": {"acceleration": ["A1"]},
    })
    _write_json(cable_spectrum_config, {
        **_config_contract(),
        "points": {"cable_force": POINTS},
        "per_point": {
            "cable_accel": {
                point: {"rho": 1.0, "L": 10.0} for point in POINTS
            }
        },
    })
    _write_json(accel_manifest, {
        "status": "ok",
        "run_request": {
            "data_root": str(root),
            "start_date": "2026-05-01",
            "end_date": "2026-05-31",
            "enabled_modules": ["accel_spectrum"],
            "config_path": str(accel_config),
            "config_sha256": hashlib.sha256(accel_config.read_bytes()).hexdigest().upper(),
        },
        "module_results": [_spectrum_record(root, "accel_spectrum", ["A1"])],
    })
    _write_json(
        cable_spectrum_manifest,
        {
            "status": "ok",
            "run_request": {
                "data_root": str(root),
                "start_date": "2026-05-01",
                "end_date": "2026-05-31",
                "enabled_modules": ["cable_accel_spectrum"],
                "config_path": str(cable_spectrum_config),
                "config_sha256": hashlib.sha256(cable_spectrum_config.read_bytes()).hexdigest().upper(),
            },
            "module_results": [
                _spectrum_record(
                    root, "cable_accel_spectrum", POINTS, cable=True
                )
            ],
        },
    )
    baseline_config_payload = json.loads(config_path.read_text(encoding="utf-8"))
    recovery_expectations = root / "run_logs" / "recovery_plot_expectations.json"
    _write_json(recovery_expectations, {
        "schema_version": 1,
        "expectation_type": "high_memory_recovery_plot_buckets",
        "baseline_config_canonical_sha256": _canonical_json_sha(baseline_config_payload),
        "cable_points": [
            {"point_id": point, "expected_plot_provenance_count": 1}
            for point in POINTS
        ],
        "cable_group": {
            "resolved_groups": {},
            "expected_plot_provenance_count": 0,
        },
        "accel_spectrum": {
            "expected_plot_provenance_count": 0,
            **_spectrum_output_contract(
                baseline_config_payload,
                "accel_spectrum",
                "2026-05-01",
                "2026-05-31",
            ),
        },
        "cable_accel_spectrum": {
            "expected_plot_provenance_count": 0,
            **_spectrum_output_contract(
                baseline_config_payload,
                "cable_accel_spectrum",
                "2026-05-01",
                "2026-05-31",
            ),
        },
    })
    group_evidence = root / "run_logs" / "cable_group_evidence.json"
    _write_json(group_evidence, {
        "evidence_type": "cable_accel_group_plot_resolution",
        "status": "ok",
        "mode": "not_applicable",
        "reason_code": "no_effective_cable_accel_groups",
        "reason": "Fixture config intentionally has no cable group figures.",
        "resolved_groups": {},
        "baseline_config_canonical_sha256": _canonical_json_sha(baseline_config_payload),
        "artifacts": [],
    })

    output = root / "run_logs" / "analysis_manifest_composite.json"
    result = compose_recovery_manifest(
        baseline,
        expected_baseline_modules=BASE_MODULES,
        cable_point_manifest_paths=cable_manifests,
        cable_merge_receipt_path=merge_result["receipt_path"],
        cable_group_evidence_path=group_evidence,
        recovery_expectations_path=recovery_expectations,
        accel_spectrum_manifest_path=accel_manifest,
        cable_spectrum_manifest_path=cable_spectrum_manifest,
        output_path=output,
        allowed_root=root,
        expected_plot_provenance_count=27,
    )
    payload = json.loads(output.read_text(encoding="utf-8"))
    assert result["module_count"] == 15
    assert payload["status"] == "ok"
    assert payload["plot_provenance_count"] == 27
    assert payload["formal_figure_stub_count"] == 58
    assert payload["formal_figure_stub_count_by_source"] == {
        "plot_provenance_backed": 27,
        "accel_spectrum": 1,
        "cable_accel_spectrum": 30,
    }
    assert payload["cable_force_engineering_valid"] is False
    assert payload["cable_force_engineering_status"] == "unverified_parameters"
    assert "Do not use" in payload["cable_force_engineering_note"]
    assert [item["key"] for item in payload["module_results"]] == [
        *BASE_MODULES,
        "cable_accel",
        "accel_spectrum",
        "cable_accel_spectrum",
    ]
    assert len(payload["source_chain"]["cable_accel_points"]) == 15
    assert payload["source_chain"]["cable_accel_points"][1]["ignored_neighbor_artifact_count"] == 1
    for record in payload["module_results"]:
        for artifact in record["artifacts"]:
            path = Path(artifact["path"])
            assert artifact["bytes"] == path.stat().st_size
            assert artifact["sha256"] == hashlib.sha256(path.read_bytes()).hexdigest().upper()

    output.unlink()
    original_cable_spectrum = json.loads(
        cable_spectrum_manifest.read_text(encoding="utf-8")
    )
    missing_figure = json.loads(json.dumps(original_cable_spectrum))
    missing_stub = f"SpecFreq_{POINTS[-1]}_20260501_20260531.jpg"
    missing_figure["module_results"][0]["artifacts"] = [
        item for item in missing_figure["module_results"][0]["artifacts"]
        if Path(item["path"]).name != missing_stub
    ]
    _write_json(cable_spectrum_manifest, missing_figure)
    with pytest.raises(ValueError, match="formal figure bundles differ"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(cable_spectrum_manifest, original_cable_spectrum)

    cable_stats_path = Path(
        original_cable_spectrum["module_results"][0]["stats_path"]
    )
    workbook = load_workbook(cable_stats_path)
    workbook.remove(workbook[POINTS[-1]])
    workbook.save(cable_stats_path)
    workbook.close()
    with pytest.raises(ValueError, match="statistics sheets differ"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_spectrum_stats(cable_stats_path, POINTS, cable=True)

    valid_group_evidence = json.loads(group_evidence.read_text(encoding="utf-8"))
    free_text_only_group = json.loads(group_evidence.read_text(encoding="utf-8"))
    free_text_only_group.pop("reason_code")
    _write_json(group_evidence, free_text_only_group)
    with pytest.raises(ValueError, match="reason_code"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(group_evidence, valid_group_evidence)

    valid_expectations = json.loads(recovery_expectations.read_text(encoding="utf-8"))
    compensating_expectations = json.loads(recovery_expectations.read_text(encoding="utf-8"))
    compensating_expectations["cable_points"][0]["expected_plot_provenance_count"] = 0
    compensating_expectations["cable_points"][1]["expected_plot_provenance_count"] = 2
    _write_json(recovery_expectations, compensating_expectations)
    with pytest.raises(ValueError, match="Cable point SLCGQ-01 plot count differs"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(recovery_expectations, valid_expectations)

    original_accel = json.loads(accel_manifest.read_text(encoding="utf-8"))
    foreign = json.loads(accel_manifest.read_text(encoding="utf-8"))
    foreign["module_results"].append(_module_record(root, "wind", "_foreign"))
    _write_json(accel_manifest, foreign)
    with pytest.raises(ValueError, match="unexpected module results"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(accel_manifest, original_accel)

    first_cable = json.loads(cable_manifests[0].read_text(encoding="utf-8"))
    cable_provenance = Path(next(
        item["path"] for item in first_cable["module_results"][0]["artifacts"]
        if item["kind"] == "plot_provenance"
    ))
    valid_cable_provenance = json.loads(cable_provenance.read_text(encoding="utf-8"))
    broken_cable_provenance = json.loads(cable_provenance.read_text(encoding="utf-8"))
    broken_cable_provenance["series"]["plotted_finite_count"] = 9
    _write_json(cable_provenance, broken_cable_provenance)
    with pytest.raises(ValueError, match="do not close"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(cable_provenance, valid_cable_provenance)

    capped_cable_provenance = json.loads(cable_provenance.read_text(encoding="utf-8"))
    capped_cable_provenance["series"]["sampling_mode"] = "capped"
    _write_json(cable_provenance, capped_cable_provenance)
    with pytest.raises(ValueError, match="requires full sampling"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(cable_provenance, valid_cable_provenance)

    wrong_scope = json.loads(accel_manifest.read_text(encoding="utf-8"))
    wrong_scope["run_request"]["end_date"] = "2026-06-30"
    _write_json(accel_manifest, wrong_scope)
    with pytest.raises(ValueError, match="scope differs"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
    _write_json(accel_manifest, original_accel)

    with pytest.raises(ValueError, match="Per-bucket plot expectations"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=88,
        )

    failed = json.loads(accel_manifest.read_text(encoding="utf-8"))
    failed["status"] = "failed"
    _write_json(accel_manifest, failed)
    with pytest.raises(ValueError, match="not successful"):
        compose_recovery_manifest(
            baseline,
            expected_baseline_modules=BASE_MODULES,
            cable_point_manifest_paths=cable_manifests,
            cable_merge_receipt_path=merge_result["receipt_path"],
            cable_group_evidence_path=group_evidence,
            recovery_expectations_path=recovery_expectations,
            accel_spectrum_manifest_path=accel_manifest,
            cable_spectrum_manifest_path=cable_spectrum_manifest,
            output_path=output,
            allowed_root=root,
            expected_plot_provenance_count=27,
        )
