from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from reporting.missing_summary import missing_summary_paths, normalize_missing_items, write_missing_summary


class MissingSummaryTests(unittest.TestCase):
    def test_normalize_prefixed_missing_items(self) -> None:
        rows = normalize_missing_items(
            [
                "section:strain:本周期未获取到结构应变有效数据",
                "warning:Missing WIM month 202601",
                "wim:202601:(a) 不同车道车辆数",
                "wind:W1风速图",
            ]
        )

        self.assertEqual(rows[0]["category"], "章节内容缺失")
        self.assertEqual(rows[0]["section"], "strain")
        self.assertEqual(rows[1]["category"], "生成警告")
        self.assertEqual(rows[1]["detail"], "Missing WIM month 202601")
        self.assertEqual(rows[2]["category"], "WIM图表缺失")
        self.assertEqual(rows[2]["item"], "202601")
        self.assertEqual(rows[3]["category"], "图表/资源缺失")
        self.assertEqual(rows[3]["section"], "wind")

    def test_write_missing_summary_outputs_txt_and_xlsx(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            report_path = Path(tmp) / "demo_report.docx"
            report_path.write_text("placeholder", encoding="utf-8")

            txt_path, xlsx_path = write_missing_summary(
                "测试报告",
                report_path,
                ["section:eq:本月未获取到地震动有效数据"],
                context={"result_root": "E:/data"},
            )

            self.assertEqual((txt_path, xlsx_path), missing_summary_paths(report_path))
            self.assertTrue(txt_path.exists())
            self.assertTrue(xlsx_path.exists())
            self.assertIn("本月未获取到地震动有效数据", txt_path.read_text(encoding="utf-8"))

            wb = load_workbook(xlsx_path)
            ws = wb["缺失内容清单"]
            self.assertEqual(ws["A6"].value, "序号")
            self.assertEqual(ws["C7"].value, "章节内容缺失")
            self.assertEqual(ws["F7"].value, "本月未获取到地震动有效数据")

    def test_write_missing_summary_no_missing_still_writes_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            report_path = Path(tmp) / "clean_report.docx"
            report_path.write_text("placeholder", encoding="utf-8")

            txt_path, xlsx_path = write_missing_summary("测试报告", report_path, [])

            self.assertIn("未发现缺失内容", txt_path.read_text(encoding="utf-8"))
            wb = load_workbook(xlsx_path)
            ws = wb["缺失内容清单"]
            self.assertEqual(ws["C7"].value, "无缺失")
            self.assertEqual(ws["G7"].value, "ok")


if __name__ == "__main__":
    unittest.main()
