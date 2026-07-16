from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from scripts.highfreq_regression import (
    ARTIFACT_CATEGORIES,
    MANIFEST_FORMAL_DIRS,
    build_source_inventory,
    compare_artifact_counts,
    compare_inventories,
    compare_workbook,
    prepare_job,
    validate_candidate_manifest,
    validate_config_binding,
    validate_candidate_plot_provenance,
)


def _write_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _write_cache_pair(root: Path, relative: str, marker: bytes = b"mat") -> None:
    mat = root / relative
    mat.parent.mkdir(parents=True, exist_ok=True)
    mat.write_bytes(marker)
    _write_json(
        Path(f"{mat}.meta.json"),
        {"schema_version": 1, "pair_id": relative, "mat_bytes": len(marker)},
    )


def _minimal_config() -> dict:
    return {
        "vendor": "hongtang",
        "data_adapter": {"time_series": {"source_mode": "auto"}},
        "points": {
            "acceleration": ["A1", "A10-X"],
            "accel_spectrum": ["A1", "A10-X"],
            "cable_accel": ["CS1", "CS6"],
            "cable_accel_spectrum": ["CS1", "CS6"],
        },
        "plot_common": {
            "dynamic_raw_sampling_mode": "full",
            "dynamic_raw_render_mode": "line",
            "gap_mode": "connect",
            "dynamic_raw_full_render_policy": "peak_preserving",
            "dynamic_raw_modules": {
                "acceleration": {
                    "sampling_mode": "full",
                    "render_mode": "line",
                    "gap_mode": "connect",
                },
                "cable_accel": {
                    "sampling_mode": "full",
                    "render_mode": "line",
                    "gap_mode": "connect",
                },
            },
        },
    }


def _write_row_workbook(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


def _write_sheet_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)
    workbook.close()


def _valid_source() -> dict:
    return {
        "source_sample_count": 10,
        "finite_source_sample_count": 9,
        "completeness_scope": "required_export_contribution",
        "internal_gap_coverage_assessed": True,
        "calendar_day_count_requested": 2,
        "complete_day_count": 2,
        "incomplete_day_count": 0,
        "incomplete_days": [],
        "missing_required_sources": [],
    }


def _valid_v2_plot(point: str) -> dict:
    return {
        "schema_version": 2,
        "file_stub": point,
        "series": [
            {
                "schema_version": 2,
                "point_id": point,
                "plot_scope": "point_time_history",
                "sampling_mode": "full",
                "render_mode": "line",
                "input_count": 10,
                "finite_count": 9,
                "plotted_finite_count": 5,
                "render_input_count": 6,
                "render_finite_input_count": 5,
                "render_vertex_count": 5,
                "reduction_applied": True,
                "reduction_scope": "render_only",
                "reduction_algorithm": "peak_preserving_bucket_minmax_v1",
                "extrema_preserved": True,
                "first_last_preserved": True,
                "source": _valid_source(),
            }
        ],
    }


class HighFrequencyRegressionPreparationTests(unittest.TestCase):
    def test_prepare_is_non_launching_mat_only_four_modules_and_date_junction_whitelist(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "source"
            for day in ("2026-04-01", "2026-04-03"):
                _write_cache_pair(source / day, "wave/cache/A1.mat")
            # These are realistic output/control names that must never be
            # linked into a fresh regression data root.
            non_date_names = (
                "stats",
                "run_logs",
                "lowfreq",
                "PSD_备查",
                "时程曲线_加速度",
            )
            for name in non_date_names:
                _write_cache_pair(source / name, "must_not_link.mat")

            config_path = root / "hongtang.json"
            _write_json(config_path, _minimal_config())
            links: list[tuple[Path, Path]] = []

            def fake_junction(link: Path, target: Path) -> None:
                links.append((link, target))
                link.mkdir()

            plan = prepare_job(
                "hongtang",
                root / "output",
                source_root=source,
                config_path=config_path,
                baseline_root=root / "baseline",
                mode="representative",
                run_id="case",
                junction_creator=fake_junction,
            )

            self.assertFalse(plan["launch_performed"])
            self.assertEqual(plan["available_dates"], ["2026-04-01", "2026-04-03"])
            self.assertEqual([link.name for link, _ in links], plan["available_dates"])
            self.assertTrue(all(link.name not in non_date_names for link, _ in links))
            self.assertEqual(plan["date_junction_count"], 2)
            inventory = json.loads(Path(plan["source_inventory_path"]).read_text(encoding="utf-8"))
            self.assertEqual(inventory["summary"]["mat_count"], 2)
            self.assertTrue(
                all(record["path"].split("/", 1)[0] in plan["available_dates"] for record in inventory["files"])
            )

            request = json.loads(Path(plan["run_request_path"]).read_text(encoding="utf-8"))
            self.assertEqual(request["enabled_modules"], [
                "acceleration", "cable_accel", "accel_spectrum", "cable_accel_spectrum"
            ])
            enabled = {key for key, value in request["options"].items() if value is True}
            self.assertEqual(
                enabled,
                {"doAccel", "doCableAccel", "doAccelSpectrum", "doCableAccelSpectrum"},
            )
            self.assertEqual(request["options"]["input_mode"], "mat_only")
            self.assertEqual(request["config"]["data_adapter"]["time_series"]["source_mode"], "mat_only")
            self.assertEqual(request["config"]["plot_common"]["dynamic_group_sampling_mode"], "off")
            self.assertEqual(request["config"]["points"]["acceleration"], ["A1", "A10-X"])
            self.assertEqual(request["config"]["points"]["cable_accel"], ["CS1", "CS6"])
            binding = validate_config_binding(plan)
            self.assertTrue(binding["passed"], binding)
            self.assertEqual(
                binding["dynamic_plot_contract"]["acceleration"]["full_render_policy"],
                "peak_preserving",
            )
            self.assertEqual(
                binding["dynamic_plot_contract"]["acceleration"]["render_max_points"],
                1200000,
            )
            self.assertEqual(
                binding["dynamic_plot_contract"]["acceleration"]["min_points_per_day"],
                12000,
            )

            snapshot = Path(plan["config_snapshot_path"])
            snapshot.write_text(snapshot.read_text(encoding="utf-8") + "\n", encoding="utf-8")
            self.assertFalse(validate_config_binding(plan)["passed"])

    def test_prepare_full_preserves_all_configured_points(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            source = root / "source"
            _write_cache_pair(source / "2026-04-01", "cache/A1.mat")
            config = _minimal_config()
            config["points"]["acceleration"].append("A2")
            config["points"]["accel_spectrum"].append("A2")
            config_path = root / "config.json"
            _write_json(config_path, config)

            def fake_junction(link: Path, target: Path) -> None:
                link.mkdir()

            plan = prepare_job(
                "hongtang",
                root / "output",
                source_root=source,
                config_path=config_path,
                mode="full",
                run_id="full",
                junction_creator=fake_junction,
            )
            self.assertEqual(plan["selected_points"]["acceleration"], ["A1", "A10-X", "A2"])
            request = json.loads(Path(plan["run_request_path"]).read_text(encoding="utf-8"))
            self.assertNotIn("dynamic_group_sampling_mode", request["config"]["plot_common"])

    def test_inventory_rejects_missing_or_orphan_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source"
            day = source / "2026-04-01"
            day.mkdir(parents=True)
            (day / "A1.mat").write_bytes(b"mat")
            with self.assertRaisesRegex(ValueError, "pairing is not closed"):
                build_source_inventory(source, [day])

    def test_inventory_comparison_detects_source_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source"
            day = source / "2026-04-01"
            _write_cache_pair(day, "cache/A1.mat")
            before = build_source_inventory(source, [day])
            after = build_source_inventory(source, [day])
            self.assertTrue(compare_inventories(before, after)["passed"])
            (day / "cache" / "A1.mat.meta.json").write_text("{}", encoding="utf-8")
            mutated = build_source_inventory(source, [day])
            result = compare_inventories(before, mutated)
            self.assertFalse(result["passed"])
            self.assertEqual(result["changed"], ["2026-04-01/cache/A1.mat.meta.json"])


class HighFrequencyRegressionComparisonTests(unittest.TestCase):
    def test_manifest_gate_requires_one_successful_exact_four_module_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            modules = [
                "acceleration",
                "cable_accel",
                "accel_spectrum",
                "cable_accel_spectrum",
            ]
            module_results = []
            for module in modules:
                stats = root / "stats" / {
                    "acceleration": "accel_stats.xlsx",
                    "cable_accel": "cable_accel_stats.xlsx",
                    "accel_spectrum": "accel_spec_stats.xlsx",
                    "cable_accel_spectrum": "cable_accel_spec_stats.xlsx",
                }[module]
                stats.parent.mkdir(parents=True, exist_ok=True)
                stats.write_bytes(module.encode("ascii"))
                module_results.append(
                    {
                        "key": module,
                        "status": "ok",
                        "stats_exists": True,
                        "stats_path": str(stats),
                        "artifact_count": 1,
                        "artifacts": [
                            {
                                "path": str(stats),
                                "exists": True,
                                "bytes": stats.stat().st_size,
                            }
                        ],
                    }
                )
            manifest = {
                "schema_version": 3,
                "status": "ok",
                "enabled_modules": modules,
                "module_results": module_results,
                "missing_expected_stats": [],
                "missing_stats_files": [],
            }
            path = root / "run_logs" / "analysis_manifest_20260716_120000.json"
            _write_json(path, manifest)

            good = validate_candidate_manifest(root)
            self.assertTrue(good["passed"], good)

            manifest["module_results"][0]["status"] = "failed"
            _write_json(path, manifest)
            failed = validate_candidate_manifest(root)
            self.assertFalse(failed["passed"])

            manifest["module_results"][0]["status"] = "ok"
            _write_json(path, manifest)
            duplicate = path.with_name("analysis_manifest_20260716_120001.json")
            _write_json(duplicate, manifest)
            self.assertFalse(validate_candidate_manifest(root)["passed"])

            path.unlink()
            duplicate.unlink()
            self.assertFalse(validate_candidate_manifest(root)["passed"])

    def test_manifest_gate_rejects_unlisted_formal_envelope_figure(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            modules = [
                "acceleration",
                "cable_accel",
                "accel_spectrum",
                "cable_accel_spectrum",
            ]
            module_results = []
            for module in modules:
                stats = root / "stats" / {
                    "acceleration": "accel_stats.xlsx",
                    "cable_accel": "cable_accel_stats.xlsx",
                    "accel_spectrum": "accel_spec_stats.xlsx",
                    "cable_accel_spectrum": "cable_accel_spec_stats.xlsx",
                }[module]
                stats.parent.mkdir(parents=True, exist_ok=True)
                stats.write_bytes(module.encode("ascii"))
                module_results.append(
                    {
                        "key": module,
                        "status": "ok",
                        "stats_exists": True,
                        "stats_path": str(stats),
                        "artifact_count": 1,
                        "artifacts": [
                            {
                                "path": str(stats),
                                "exists": True,
                                "bytes": stats.stat().st_size,
                            }
                        ],
                    }
                )

            envelope_dir = next(
                directory
                for category, directory, _, _ in ARTIFACT_CATEGORIES
                if category == "cable_accel_envelope"
            )
            envelope = root / envelope_dir / "CableAccelEnvelope30_CF-1.jpg"
            envelope.parent.mkdir(parents=True)
            envelope.write_bytes(b"figure")
            group_dir = "频谱峰值曲线_结构加速度_组图"
            self.assertIn(group_dir, MANIFEST_FORMAL_DIRS)
            self.assertNotIn("频谱峰值曲线_加速度_组图", MANIFEST_FORMAL_DIRS)
            group_figure = root / group_dir / "Accel_AZ_Group.jpg"
            group_figure.parent.mkdir(parents=True)
            group_figure.write_bytes(b"group-figure")
            manifest = {
                "schema_version": 3,
                "status": "ok",
                "enabled_modules": modules,
                "module_results": module_results,
                "missing_expected_stats": [],
                "missing_stats_files": [],
            }
            manifest_path = root / "run_logs" / "analysis_manifest_20260716_120000.json"
            _write_json(manifest_path, manifest)

            missing = validate_candidate_manifest(root)
            self.assertFalse(missing["passed"])
            coverage = next(
                item for item in missing["checks"] if item["name"] == "formal_artifact_coverage"
            )
            self.assertEqual(
                coverage["detail"]["missing"],
                sorted([str(envelope.resolve()), str(group_figure.resolve())]),
            )

            cable = next(item for item in module_results if item["key"] == "cable_accel")
            cable["artifacts"].append(
                {"path": str(envelope), "exists": True, "bytes": envelope.stat().st_size}
            )
            cable["artifact_count"] = 2
            acceleration = next(item for item in module_results if item["key"] == "acceleration")
            acceleration["artifacts"].append(
                {
                    "path": str(group_figure),
                    "exists": True,
                    "bytes": group_figure.stat().st_size,
                }
            )
            acceleration["artifact_count"] = 2
            _write_json(manifest_path, manifest)
            self.assertTrue(validate_candidate_manifest(root)["passed"])

    def test_row_and_sheet_workbooks_compare_selected_points_cell_by_cell(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            baseline_rows = root / "baseline_rows.xlsx"
            candidate_rows = root / "candidate_rows.xlsx"
            header = ["PointID", "Min", "Max"]
            _write_row_workbook(baseline_rows, [header, ["A1", -1.0, 1.0], ["A2", -2.0, 2.0]])
            _write_row_workbook(candidate_rows, [header, ["A1", -1.0, 1.0000001]])
            exact = compare_workbook(
                baseline_rows, candidate_rows, ["A1"], row_oriented=True
            )
            self.assertFalse(exact["passed"])
            tolerant = compare_workbook(
                baseline_rows,
                candidate_rows,
                ["A1"],
                row_oriented=True,
                abs_tol=1e-6,
            )
            self.assertTrue(tolerant["passed"])

            baseline_sheets = root / "baseline_sheets.xlsx"
            candidate_sheets = root / "candidate_sheets.xlsx"
            rows = [["Date", "Freq"], ["2026-04-01", 1.25]]
            _write_sheet_workbook(baseline_sheets, {"A1": rows, "A2": rows})
            _write_sheet_workbook(candidate_sheets, {"A1": rows})
            compared = compare_workbook(
                baseline_sheets, candidate_sheets, ["A1"], row_oriented=False
            )
            self.assertTrue(compared["passed"])

    def test_v2_candidate_plot_gate_accepts_render_only_peak_preserving_and_rejects_unknown(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            accel = root / "时程曲线_加速度" / "A1_20260401_20260430.plot.json"
            cable = root / "时程曲线_索力加速度" / "CS1_20260401_20260430.plot.json"
            _write_json(accel, _valid_v2_plot("A1"))
            _write_json(cable, _valid_v2_plot("CS1"))
            selected = {"acceleration": ["A1"], "cable_accel": ["CS1"]}
            good = validate_candidate_plot_provenance(root, selected)
            self.assertTrue(good["passed"])
            broken = _valid_v2_plot("CS1")
            broken["series"][0]["reduction_algorithm"] = "lttb"
            _write_json(cable, broken)
            bad = validate_candidate_plot_provenance(root, selected)
            self.assertFalse(bad["passed"])
            self.assertTrue(any("unsupported" in row.get("message", "") for row in bad["rows"]))

    def test_selected_artifact_counts_include_envelope_psd_fig_and_plot_json(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder)
            baseline = root / "baseline"
            candidate = root / "candidate"
            selected = {"acceleration": ["A1"], "cable_accel": ["CS1"]}
            for target in (baseline, candidate):
                for _, directory, family, kinds in ARTIFACT_CATEGORIES:
                    point = selected[family][0]
                    folder_path = target / directory
                    folder_path.mkdir(parents=True, exist_ok=True)
                    for kind in kinds:
                        suffix = {"image": ".jpg", "fig": ".fig", "plot_json": ".plot.json"}[kind]
                        (folder_path / f"prefix_{point}_20260401{suffix}").write_bytes(b"x")
            result = compare_artifact_counts(baseline, candidate, selected)
            self.assertTrue(result["passed"])
            (candidate / "时程曲线_索力加速度_包络30min" / "prefix_CS1_20260401.fig").unlink()
            failed = compare_artifact_counts(baseline, candidate, selected)
            self.assertFalse(failed["passed"])


if __name__ == "__main__":
    unittest.main()
